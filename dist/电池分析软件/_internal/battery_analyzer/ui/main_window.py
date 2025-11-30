#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Battery Analyzer ä¸»çª—å£ï¼ˆç®€ä½“ä¸­æ–‡ UI éª¨æ¶ï¼‰ã€‚

å¸ƒå±€ç›®æ ‡ï¼š
- å·¦ä¾§æ§åˆ¶é¢æ¿ï¼šæµ‹è¯•è®¾ç½®ã€å¯¼å‡ºæŠ¥å‘Šã€é‡‡é›†æ—¶é•¿/ç‚¹æ•°ã€ä¿å­˜/å¬å›æ³¢å½¢ã€äº§å“ä¿¡æ¯ã€å¼€å§‹æŒ‰é’®ã€‚
- ä¸­éƒ¨ï¼šåŒæ³¢å½¢åŒºåŸŸï¼ˆå·¦ï¼šä¸‰å…ƒç”µæ± ï¼›å³ï¼šåˆ€ç‰‡ç”µæ± ï¼‰ã€‚
- åº•éƒ¨ï¼šå…³é”®æŒ‡æ ‡ KPIï¼ˆç”µå‹ã€æ¸©åº¦ï¼‰ï¼Œå„è‡ªå¤§å·æ•°å­—æ˜¾ç¤ºã€‚

åç»­ä¼šæŠŠ LR8450 çš„å®æ—¶é‡‡é›†ä¸åˆ†æï¼ˆæ¸©å‡æ¯”å¯¹ã€ç”µå‹å‹é™ã€mX+bã€mAhï¼‰æ¥å…¥åˆ°ç›¸åº”æ§½å‡½æ•°ã€‚
"""

from __future__ import annotations

import time
from datetime import datetime
from typing import Optional, List

from PySide6.QtCore import Qt, Signal, QTimer
from PySide6.QtWidgets import (
    QWidget,
    QMainWindow,
    QHBoxLayout,
    QVBoxLayout,
    QGridLayout,
    QGroupBox,
    QLabel,
    QPushButton,
    QFormLayout,
    QLineEdit,
    QSpinBox,
    QComboBox,
    QApplication,
    QStatusBar,
    QDialog,
    QRadioButton,
    QMessageBox,
)

import pyqtgraph as pg
import numpy as np

from battery_analyzer.core.lr8450_client import LR8450Client
from battery_analyzer.core.analysis_engine import BatteryAnalysisEngine
from battery_analyzer.core.acquisition_thread import DataAcquisitionThread
from battery_analyzer.core.device_worker import DeviceConfigWorker, DeviceStopWorker, DeviceStartWorker
from battery_analyzer.ui.dialogs.channel_config_dialog import ChannelConfigDialog
from battery_analyzer.ui.dialogs.device_connect_dialog import DeviceConnectDialog


class KPIWidget(QWidget):
    """KPI æ•°å­—æ˜¾ç¤ºï¼ˆæ ‡é¢˜ + æ•°å€¼+å•ä½åœ¨åŒä¸€æ¡†å†…ï¼‰ã€‚"""

    def __init__(self, title: str, unit: str, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.unit = unit

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        self.title_label = QLabel(title)
        self.title_label.setObjectName("kpiTitle")
        self.title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.title_label.setProperty("class", "kpiTitle")

        # æ•°å€¼+å•ä½æ”¾åœ¨åŒä¸€ä¸ªæ ‡ç­¾å†…
        self.value_label = QLabel(f"0.00{unit}")
        self.value_label.setObjectName("valueLabel")
        self.value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.value_label.setProperty("class", "kpi")
        self.value_label.setStyleSheet(
            "background-color: #0b213f; "
            "border: 1px solid #2b4b7d; "
            "border-radius: 6px; "
            "padding: 8px 12px; "
            "color: #98d2ff; "
            "font-weight: bold; "
            "font-size: 28px;"
        )

        layout.addWidget(self.title_label)
        layout.addWidget(self.value_label)

    def set_value(self, value_text: str) -> None:
        """è®¾ç½®æ•°å€¼ï¼ˆè‡ªåŠ¨é™„åŠ å•ä½ï¼‰ã€‚"""
        self.value_label.setText(f"{value_text}{self.unit}")


class WaveformPair(QWidget):
    """åŒæ³¢å½¢æ˜¾ç¤ºé¢æ¿ï¼ˆæ¯ä¸ªå›¾è¡¨æœ‰åŒYè½´ï¼šå·¦ä¾§ç”µå‹ï¼Œå³ä¾§æ¸©åº¦ï¼‰ã€‚"""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        layout = QGridLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(12)

        # å·¦ï¼šä¸‰å…ƒç”µæ± 
        self.left_title = QLabel("ä¸‰å…ƒç”µæ± æ³¢å½¢ (Ternary Battery Waveform)")
        self.left_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.left_title.setStyleSheet("color: #6dd5ed; font-size: 14px; font-weight: bold;")
        
        self.left_plot = self._create_dual_axis_plot()
        
        # å³ï¼šåˆ€ç‰‡ç”µæ± 
        self.right_title = QLabel("åˆ€ç‰‡ç”µæ± æ³¢å½¢ (Blade Battery Waveform)")
        self.right_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.right_title.setStyleSheet("color: #6dd5ed; font-size: 14px; font-weight: bold;")
        
        self.right_plot = self._create_dual_axis_plot()

        layout.addWidget(self.left_title, 0, 0)
        layout.addWidget(self.right_title, 0, 1)
        layout.addWidget(self.left_plot, 1, 0)
        layout.addWidget(self.right_plot, 1, 1)
        layout.setColumnStretch(0, 1)
        layout.setColumnStretch(1, 1)

    def _create_dual_axis_plot(self) -> pg.PlotWidget:
        """åˆ›å»ºå¸¦åŒYè½´çš„å›¾è¡¨ï¼ˆå·¦ï¼šç”µå‹ï¼Œå³ï¼šæ¸©åº¦ï¼‰ã€‚"""
        plot_widget = pg.PlotWidget()
        plot_widget.setBackground("#0c1830")
        
        # éšè—å·¦ä¸‹è§’çš„ AutoRange æŒ‰é’®ï¼ˆAæŒ‰é’®ï¼‰
        plot_widget.plotItem.hideButtons()
        
        # å·¦Yè½´ï¼šç”µå‹ï¼ˆVï¼‰
        plot_widget.setLabel('left', 'ç”µå‹ V', color='#33c1ff', **{'font-size': '11pt'})
        plot_widget.setYRange(0, 10)
        plot_widget.getAxis('left').setPen(pg.mkPen('#33c1ff', width=2))
        plot_widget.getAxis('left').setTextPen('#33c1ff')
        
        # åº•éƒ¨Xè½´ï¼šæ—¶é—´ï¼ˆTimeï¼‰
        plot_widget.setLabel('bottom', 'æ—¶é—´ Time', units='s', color='#6dd5ed', **{'font-size': '11pt'})
        plot_widget.setXRange(0, 300)
        plot_widget.getAxis('bottom').setPen(pg.mkPen('#4a90e2', width=2))
        plot_widget.getAxis('bottom').setTextPen('#6dd5ed')
        
        # å³Yè½´ï¼šæ¸©åº¦ï¼ˆTï¼‰
        viewbox_temp = pg.ViewBox()
        plot_widget.plotItem.scene().addItem(viewbox_temp)
        plot_widget.plotItem.getAxis('right').linkToView(viewbox_temp)
        viewbox_temp.setXLink(plot_widget.plotItem)
        
        # é…ç½®å³è½´
        plot_widget.plotItem.showAxis('right')
        plot_widget.setLabel('right', 'æ¸©åº¦ T', units='Â°C', color='#ffb347', **{'font-size': '11pt'})
        plot_widget.getAxis('right').setPen(pg.mkPen('#ffb347', width=2))
        plot_widget.getAxis('right').setTextPen('#ffb347')
        
        # åŒæ­¥å³ä¾§ ViewBox çš„å‡ ä½•ä½ç½®
        def update_views():
            viewbox_temp.setGeometry(plot_widget.plotItem.vb.sceneBoundingRect())
            viewbox_temp.linkedViewChanged(plot_widget.plotItem.vb, viewbox_temp.XAxis)
        
        update_views()
        plot_widget.plotItem.vb.sigResized.connect(update_views)
        
        # è®¾ç½®æ¸©åº¦è½´èŒƒå›´
        viewbox_temp.setYRange(0, 260)
        
        # å­˜å‚¨ viewbox ä»¥ä¾¿åç»­ç»˜å›¾
        plot_widget.viewbox_temp = viewbox_temp
        
        # ç½‘æ ¼
        plot_widget.showGrid(x=True, y=True, alpha=0.15)
        
        return plot_widget


class ControlPanel(QWidget):
    """å·¦ä¾§æ§åˆ¶é¢æ¿ã€‚"""

    start_requested = Signal()
    voltage_color_changed = Signal(str)  # ç”µå‹é¢œè‰²æ”¹å˜ä¿¡å·
    temp_color_changed = Signal(str)     # æ¸©åº¦é¢œè‰²æ”¹å˜ä¿¡å·
    voltage_width_changed = Signal(int)  # ç”µå‹çº¿å®½æ”¹å˜ä¿¡å·
    temp_width_changed = Signal(int)     # æ¸©åº¦çº¿å®½æ”¹å˜ä¿¡å·

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(10)

        # æ“ä½œåŒº
        ops = QGroupBox("æ“ä½œ")
        ops_layout = QVBoxLayout(ops)
        self.btn_device_connect = QPushButton("è¿æ¥è®¾å¤‡")
        self.btn_device_connect.setStyleSheet("""
            QPushButton {
                background-color: #0d5aa7;
                color: #ffffff;
                font-weight: bold;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #0c4d91;
            }
        """)
        self.btn_channel_config = QPushButton("é€šé“é…ç½®")
        self.btn_report = QPushButton("å¯¼å‡ºæŠ¥å‘Š")
        self.btn_save = QPushButton("ä¿å­˜æ³¢å½¢")
        self.btn_recall = QPushButton("å¬å›æ³¢å½¢")
        ops_layout.addWidget(self.btn_device_connect)
        ops_layout.addWidget(self.btn_channel_config)
        ops_layout.addWidget(self.btn_report)
        ops_layout.addWidget(self.btn_save)
        ops_layout.addWidget(self.btn_recall)

        # é€šé“ä¸æ˜¾ç¤ºé€‰é¡¹
        vis = QGroupBox("æ˜¾ç¤ºè®¾ç½®")
        vis_layout = QVBoxLayout(vis)
        vis_layout.setSpacing(8)
        
        # ç”µå‹è¡Œï¼šæ ‡ç­¾ + åƒç´ é€‰æ‹© + é¢œè‰²å—
        volt_row = QHBoxLayout()
        volt_label = QLabel("ç”µå‹")
        volt_label.setFixedWidth(40)
        volt_row.addWidget(volt_label)
        self.combo_volt_pix = QComboBox()
        self.combo_volt_pix.addItems(["1åƒç´ ", "2åƒç´ ", "3åƒç´ ", "4åƒç´ "])
        self.combo_volt_pix.setCurrentIndex(2)  # é»˜è®¤3åƒç´ 
        self.combo_volt_pix.setFixedWidth(80)
        self.combo_volt_pix.currentIndexChanged.connect(self._on_volt_width_changed)
        volt_row.addWidget(self.combo_volt_pix)
        self.volt_color_btn = QPushButton()
        self.volt_color_btn.setFixedSize(50, 22)
        self.volt_color_btn.setStyleSheet("background-color: #33c1ff; border: 1px solid #2b4b7d; border-radius: 3px;")
        self.volt_color_btn.clicked.connect(lambda: self._choose_color("volt"))
        volt_row.addWidget(self.volt_color_btn)
        volt_row.addStretch()
        vis_layout.addLayout(volt_row)

        # æ¸©åº¦è¡Œï¼šæ ‡ç­¾ + åƒç´ é€‰æ‹© + é¢œè‰²å—
        temp_row = QHBoxLayout()
        temp_label = QLabel("æ¸©åº¦")
        temp_label.setFixedWidth(40)
        temp_row.addWidget(temp_label)
        self.combo_temp_pix = QComboBox()
        self.combo_temp_pix.addItems(["1åƒç´ ", "2åƒç´ ", "3åƒç´ ", "4åƒç´ "])
        self.combo_temp_pix.setCurrentIndex(2)  # é»˜è®¤3åƒç´ 
        self.combo_temp_pix.setFixedWidth(80)
        self.combo_temp_pix.currentIndexChanged.connect(self._on_temp_width_changed)
        temp_row.addWidget(self.combo_temp_pix)
        self.temp_color_btn = QPushButton()
        self.temp_color_btn.setFixedSize(50, 22)
        self.temp_color_btn.setStyleSheet("background-color: #ffb347; border: 1px solid #2b4b7d; border-radius: 3px;")
        self.temp_color_btn.clicked.connect(lambda: self._choose_color("temp"))
        temp_row.addWidget(self.temp_color_btn)
        temp_row.addStretch()
        vis_layout.addLayout(temp_row)

        # å­˜å‚¨å½“å‰é¢œè‰²ï¼ˆä¸Yè½´é¢œè‰²ä¸€è‡´ï¼‰
        self.volt_color = "#33c1ff"   # ç”µå‹ï¼šé’è“è‰²
        self.temp_color = "#ffb347"   # æ¸©åº¦ï¼šæ©™è‰²

        # äº§å“ä¿¡æ¯ï¼ˆä¸Šä¸‹æ’åˆ—ï¼Œç´§å‡‘å¸ƒå±€ï¼‰
        info = QGroupBox("äº§å“ä¿¡æ¯")
        info_layout = QVBoxLayout(info)
        info_layout.setSpacing(6)
        
        # äº§å“å‹å·
        info_layout.addWidget(QLabel("äº§å“å‹å·"))
        self.edit_model = QLineEdit("5mm")
        self.edit_model.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.edit_model.setStyleSheet("font-size: 18px; font-weight: bold; color: #33c1ff; padding: 6px;")
        info_layout.addWidget(self.edit_model)
        
        # äº§å“æµæ°´å·
        info_layout.addWidget(QLabel("äº§å“æµæ°´å·"))
        self.edit_sn = QLineEdit("25mm/s")
        self.edit_sn.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.edit_sn.setStyleSheet("font-size: 18px; font-weight: bold; color: #33c1ff; padding: 6px;")
        info_layout.addWidget(self.edit_sn)
        
        # æµ‹è¯•å‘˜
        info_layout.addWidget(QLabel("æµ‹è¯•å‘˜"))
        self.edit_tester = QLineEdit("1dot/s")
        self.edit_tester.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.edit_tester.setStyleSheet("font-size: 18px; font-weight: bold; color: #33c1ff; padding: 6px;")
        info_layout.addWidget(self.edit_tester)

        # æ—¶é—´ä¿¡æ¯åŒºåŸŸ
        time_info = QGroupBox("æ—¶é—´ä¿¡æ¯")
        time_info_layout = QVBoxLayout(time_info)
        time_info_layout.setSpacing(4)

        # ç³»ç»Ÿæ—¶é—´
        time_info_layout.addWidget(QLabel("ç³»ç»Ÿæ—¶é—´"))
        self.lbl_system_time = QLabel("--:--:--")
        self.lbl_system_time.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_system_time.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                color: #00ff88;
                background-color: #1a1a2e;
                border: 1px solid #00ff88;
                border-radius: 4px;
                padding: 8px;
            }
        """)
        time_info_layout.addWidget(self.lbl_system_time)

        # è½¯ä»¶è¿è¡Œæ—¶é—´
        time_info_layout.addWidget(QLabel("è½¯ä»¶è¿è¡Œæ—¶é—´"))
        self.lbl_running_time = QLabel("00:00:00")
        self.lbl_running_time.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_running_time.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                color: #ffaa00;
                background-color: #1a1a2e;
                border: 1px solid #ffaa00;
                border-radius: 4px;
                padding: 8px;
            }
        """)
        time_info_layout.addWidget(self.lbl_running_time)

        # å¼€å§‹æŒ‰é’®
        self.btn_start = QPushButton("å¼€å§‹")
        self.btn_start.setObjectName("startButton")
        self.btn_start.clicked.connect(self.start_requested.emit)

        # æ–°åŠŸèƒ½æŒ‰é’®åŒº
        functions = QGroupBox("åˆ†æåŠŸèƒ½")
        func_layout = QVBoxLayout(functions)
        func_layout.setSpacing(8)

        self.btn_mx_plus_b = QPushButton("mX+b")
        self.btn_mx_plus_b.setStyleSheet("""
            QPushButton {
                background-color: #1e3a8a;
                color: #ffffff;
                border: 2px solid #3b82f6;
                border-radius: 6px;
                padding: 10px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #1d4ed8;
                border-color: #60a5fa;
                color: #ffffff;
            }
        """)
        func_layout.addWidget(self.btn_mx_plus_b)

        self.btn_mah_test = QPushButton("mAh å®¹é‡æµ‹è¯•")
        self.btn_mah_test.setStyleSheet("""
            QPushButton {
                background-color: #059669;
                color: #ffffff;
                border: 2px solid #10b981;
                border-radius: 6px;
                padding: 10px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #047857;
                border-color: #34d399;
                color: #ffffff;
            }
        """)
        func_layout.addWidget(self.btn_mah_test)

        root.addWidget(ops)
        root.addWidget(functions)
        root.addWidget(vis)
        root.addWidget(info)
        root.addWidget(time_info)
        root.addStretch(1)
        root.addWidget(self.btn_start)
    
    def _choose_color(self, channel: str) -> None:
        """é€‰æ‹©é€šé“é¢œè‰²ã€‚"""
        from PySide6.QtGui import QColor
        from battery_analyzer.ui.color_dialog import SimpleColorDialog
        
        # è·å–å½“å‰é¢œè‰²
        current_color = QColor(self.volt_color if channel == "volt" else self.temp_color)
        
        # æ‰“å¼€è‡ªå®šä¹‰ä¸­æ–‡é¢œè‰²é€‰æ‹©å¯¹è¯æ¡†
        title = "é€‰æ‹©ç”µå‹é¢œè‰²" if channel == "volt" else "é€‰æ‹©æ¸©åº¦é¢œè‰²"
        color = SimpleColorDialog.get_color_from_dialog(current_color, title, self)
        
        if color and color.isValid():
            color_hex = color.name()
            if channel == "volt":
                self.volt_color = color_hex
                self.volt_color_btn.setStyleSheet(
                    f"background-color: {color_hex}; border: 1px solid #2b4b7d; border-radius: 3px;"
                )
                # å‘å°„ä¿¡å·é€šçŸ¥ä¸»çª—å£æ›´æ–°ç”µå‹æ›²çº¿é¢œè‰²
                self.voltage_color_changed.emit(color_hex)
            else:
                self.temp_color = color_hex
                self.temp_color_btn.setStyleSheet(
                    f"background-color: {color_hex}; border: 1px solid #2b4b7d; border-radius: 3px;"
                )
                # å‘å°„ä¿¡å·é€šçŸ¥ä¸»çª—å£æ›´æ–°æ¸©åº¦æ›²çº¿é¢œè‰²
                self.temp_color_changed.emit(color_hex)

    def _on_volt_width_changed(self, index: int) -> None:
        """ç”µå‹çº¿å®½æ”¹å˜"""
        width = index + 1  # 0->1, 1->2, 2->3, 3->4
        self.voltage_width_changed.emit(width)

    def _on_temp_width_changed(self, index: int) -> None:
        """æ¸©åº¦çº¿å®½æ”¹å˜"""
        width = index + 1  # 0->1, 1->2, 2->3, 3->4
        self.temp_width_changed.emit(width)


class MainWindow(QMainWindow):
    """ä¸»çª—å£ã€‚"""

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("ç”µæ± ç”µå‹ä¸æ¸©å‡åˆ†æè½¯ä»¶")
        self.resize(1280, 800)

        # ä¸­å¤®å¸ƒå±€ï¼ˆåŒ…å«é¡¶éƒ¨æ ‡é¢˜æ ï¼‰
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # é¡¶éƒ¨æ ‡é¢˜æ 
        title_bar = self._create_title_bar()
        main_layout.addWidget(title_bar)

        # ä¸»å†…å®¹åŒºåŸŸ
        content = QWidget()
        hbox = QHBoxLayout(content)
        hbox.setContentsMargins(8, 8, 8, 8)
        hbox.setSpacing(12)

        # å·¦ä¾§æ§åˆ¶
        self.control = ControlPanel()
        self.control.setFixedWidth(260)
        self.control.start_requested.connect(self._on_start)
        self.control.voltage_color_changed.connect(self.update_voltage_color)
        self.control.temp_color_changed.connect(self.update_temp_color)
        self.control.voltage_width_changed.connect(self.update_voltage_width)
        self.control.temp_width_changed.connect(self.update_temp_width)

        # è¿æ¥æ“ä½œæŒ‰é’®
        self.control.btn_device_connect.clicked.connect(self._show_device_connect_dialog)
        self.control.btn_channel_config.clicked.connect(self._show_channel_config_dialog)
        self.control.btn_report.clicked.connect(self._export_report)
        self.control.btn_save.clicked.connect(self._save_waveform)
        self.control.btn_recall.clicked.connect(self._recall_waveform)
        
        # è¿æ¥åˆ†æåŠŸèƒ½æŒ‰é’®
        self.control.btn_mx_plus_b.clicked.connect(self._show_mx_plus_b_dialog)
        self.control.btn_mah_test.clicked.connect(self._show_mah_test_dialog)

        # ä¸­éƒ¨åŒæ³¢å½¢
        self.waveforms = WaveformPair()

        # åº•éƒ¨ KPI åŒºåŸŸï¼ˆ4ä¸ªKPIï¼šå·¦ä¾§2ä¸ªä¸‰å…ƒç”µæ± ï¼Œå³ä¾§2ä¸ªåˆ€ç‰‡ç”µæ± ï¼‰
        bottom = QWidget()
        bottom_layout = QHBoxLayout(bottom)
        bottom_layout.setContentsMargins(0, 0, 0, 0)
        bottom_layout.setSpacing(12)
        
        # ä¸‰å…ƒç”µæ± ç”µå‹
        self.ternary_voltage_kpi = KPIWidget("ä¸‰å…ƒç”µæ± ç”µå‹\nTernary Battery Voltage", "V")
        bottom_layout.addWidget(self.ternary_voltage_kpi)
        
        # ä¸‰å…ƒç”µæ± æ¸©åº¦
        self.ternary_temp_kpi = KPIWidget("ä¸‰å…ƒç”µæ± æ¸©åº¦\nTernary Battery Temperature", "Â°C")
        bottom_layout.addWidget(self.ternary_temp_kpi)
        
        # åˆ€ç‰‡ç”µæ± ç”µå‹
        self.blade_voltage_kpi = KPIWidget("åˆ€ç‰‡ç”µæ± ç”µå‹\nBlade Battery Voltage", "V")
        bottom_layout.addWidget(self.blade_voltage_kpi)
        
        # åˆ€ç‰‡ç”µæ± æ¸©åº¦
        self.blade_temp_kpi = KPIWidget("åˆ€ç‰‡ç”µæ± æ¸©åº¦\nBlade Battery Temperature", "Â°C")
        bottom_layout.addWidget(self.blade_temp_kpi)

        # å°†ä¸­éƒ¨åŒºåŸŸç»„è£…ä¸ºä¸Šä¸‹ä¸¤å—
        center_col = QWidget()
        center_layout = QVBoxLayout(center_col)
        center_layout.setContentsMargins(0, 0, 0, 0)
        center_layout.setSpacing(12)
        center_layout.addWidget(self.waveforms, 1)
        center_layout.addWidget(bottom, 0)

        hbox.addWidget(self.control, 0)
        hbox.addWidget(center_col, 1)

        main_layout.addWidget(content, 1)

        # çŠ¶æ€æ 
        status = QStatusBar()
        self.setStatusBar(status)
        status.showMessage("å°±ç»ªâ€”â€”ç­‰å¾…è¿æ¥ä¸å¼€å§‹æµ‹è¯•")

        # å­˜å‚¨æ›²çº¿å¯¹è±¡ä»¥ä¾¿æ›´æ–°é¢œè‰²
        self.volt_curves = []  # [å·¦å›¾ç”µå‹æ›²çº¿, å³å›¾ç”µå‹æ›²çº¿]
        self.temp_curves = []  # [å·¦å›¾æ¸©åº¦æ›²çº¿, å³å›¾æ¸©åº¦æ›²çº¿]
        
        # LR8450è®¾å¤‡å®¢æˆ·ç«¯
        self.device_client: Optional[LR8450Client] = None
        self.device_connected = False

        # æ•°æ®é‡‡é›†çº¿ç¨‹
        self.acquisition_thread: Optional[DataAcquisitionThread] = None

        # åˆ†æå¼•æ“
        self.analysis_engine = BatteryAnalysisEngine()

        # å·²å®‰è£…çš„æ¨¡å—åˆ—è¡¨ï¼ˆè¿æ¥è®¾å¤‡åè‡ªåŠ¨æ£€æµ‹ï¼‰
        self.installed_modules: List[int] = []

        # ä¿å­˜çš„è¿æ¥é…ç½®ï¼ˆä»æ–‡ä»¶åŠ è½½ï¼‰
        self.saved_connection_config: dict = {}

        # é€šé“é…ç½®ï¼ˆåŒ…å«è¯¦ç»†å‚æ•°ï¼‰
        self.channel_config = {
            'ternary_voltage': {
                'channel': 'CH1_2',  # ä¸‰å…ƒç”µæ± ç”µå‹
                'type': 'VOLTAGE',
                'range': 20.0,
            },
            'ternary_temp': {
                'channel': 'CH1_1',  # ä¸‰å…ƒç”µæ± æ¸©åº¦
                'type': 'TEMPERATURE',
                'range': 100,
                'thermocouple': 'K',
                'int_ext': 'INT',
            },
            'blade_voltage': {
                'channel': 'CH1_3',  # åˆ€ç‰‡ç”µæ± ç”µå‹
                'type': 'VOLTAGE',
                'range': 20.0,
            },
            'blade_temp': {
                'channel': 'CH1_4',  # åˆ€ç‰‡ç”µæ± æ¸©åº¦
                'type': 'TEMPERATURE',
                'range': 100,
                'thermocouple': 'K',
                'int_ext': 'INT',
            },
        }

        # æ•°æ®é‡‡é›†çŠ¶æ€
        self.is_running = False
        self.data_index = 0
        self.max_points = 600  # æœ€å¤šæ˜¾ç¤º600ä¸ªç‚¹ï¼ˆ300ç§’ï¼‰

        # æ•°æ®ç¼“å†²åŒº
        self.x_data = []
        self.ternary_volt_data = []
        self.ternary_temp_data = []
        self.blade_volt_data = []
        self.blade_temp_data = []

        # å½“å‰é¢œè‰²å’Œçº¿å®½ï¼ˆä»æ§åˆ¶é¢æ¿è·å–ï¼‰
        # æ³¨æ„ï¼šè¿™äº›é¢œè‰²å¿…é¡»ä¸ _create_dual_axis_plot() ä¸­çš„Yè½´é¢œè‰²ä¸€è‡´
        self.current_volt_color = "#33c1ff"  # ç”µå‹ï¼šé’è“è‰²ï¼ˆä¸å·¦Yè½´ä¸€è‡´ï¼‰
        self.current_temp_color = "#ffb347"  # æ¸©åº¦ï¼šæ©™è‰²ï¼ˆä¸å³Yè½´ä¸€è‡´ï¼‰
        self.current_volt_width = 3  # é»˜è®¤3åƒç´ 
        self.current_temp_width = 3  # é»˜è®¤3åƒç´ 

        # å®šæ—¶å™¨ï¼šç”¨äºè™šæ‹Ÿæ•°æ®æ¨¡å¼
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self._update_waveform_virtual)
        self.update_interval_ms = 100  # æ›´æ–°é—´éš”ï¼ˆæ¯«ç§’ï¼‰

        # æ—¶é—´æ˜¾ç¤ºå®šæ—¶å™¨
        self.app_start_time = time.time()  # è½¯ä»¶å¯åŠ¨æ—¶é—´
        self.time_display_timer = QTimer()
        self.time_display_timer.timeout.connect(self._update_time_display)
        self.time_display_timer.start(1000)  # æ¯ç§’æ›´æ–°ä¸€æ¬¡
        self._update_time_display()  # ç«‹å³æ›´æ–°ä¸€æ¬¡

        # æ ‡è®°é€šé“æ˜¯å¦å·²é…ç½®ï¼ˆé¿å…æ¯æ¬¡å¼€å§‹éƒ½é‡æ–°é…ç½®ï¼‰
        self._channels_configured = False
        self._last_channel_config_hash = None

        # åå°å·¥ä½œçº¿ç¨‹
        self._config_worker: Optional[DeviceConfigWorker] = None
        self._stop_worker: Optional[DeviceStopWorker] = None
        self._start_worker: Optional[DeviceStartWorker] = None

        # é¢„ç½®ç¤ºä¾‹æ³¢å½¢ï¼ˆåˆå§‹é™æ€æ•°æ®ï¼‰
        self._plot_demo()

        # åŠ è½½ä¸Šæ¬¡ä¿å­˜çš„é…ç½®ï¼ˆé€šé“é…ç½®ã€è¿æ¥é…ç½®ã€äº§å“ä¿¡æ¯ï¼‰
        self._load_channel_config_from_file()

        # åˆå§‹åŒ–Yè½´èŒƒå›´ï¼ˆæ ¹æ®é…ç½®çš„é‡ç¨‹ï¼‰
        self._update_plot_ranges()

        # è¿æ¥äº§å“ä¿¡æ¯å­—æ®µçš„ä¿¡å·ï¼Œè‡ªåŠ¨ä¿å­˜é…ç½®ï¼ˆä½¿ç”¨ editingFinished è€Œä¸æ˜¯ textChanged é¿å…é¢‘ç¹ä¿å­˜ï¼‰
        self.control.edit_model.editingFinished.connect(self._save_channel_config_to_file)
        self.control.edit_sn.editingFinished.connect(self._save_channel_config_to_file)
        self.control.edit_tester.editingFinished.connect(self._save_channel_config_to_file)

    def _update_time_display(self) -> None:
        """æ›´æ–°æ—¶é—´æ˜¾ç¤ºï¼ˆç³»ç»Ÿæ—¶é—´å’Œè½¯ä»¶è¿è¡Œæ—¶é—´ï¼‰"""
        # æ›´æ–°ç³»ç»Ÿæ—¶é—´
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.control.lbl_system_time.setText(current_time)

        # æ›´æ–°è½¯ä»¶è¿è¡Œæ—¶é—´
        elapsed = int(time.time() - self.app_start_time)
        hours = elapsed // 3600
        minutes = (elapsed % 3600) // 60
        seconds = elapsed % 60
        running_time = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        self.control.lbl_running_time.setText(running_time)

    def _plot_demo(self) -> None:
        """ç»˜åˆ¶ç¤ºä¾‹æ³¢å½¢ï¼ˆç”µå‹åœ¨å·¦Yè½´ï¼Œæ¸©åº¦åœ¨å³Yè½´ï¼‰ã€‚"""
        import numpy as np

        x = np.linspace(0, 300, 600)
        
        # ä¸‰å…ƒç”µæ± æ•°æ®
        y_v_ternary = 5 + 0.3 * np.sin(0.05 * x) + 0.1 * np.random.randn(600)  # ç”µå‹ 0-10V
        y_t_ternary = 130 + 50 * np.sin(0.02 * x + 1.2) + 5 * np.random.randn(600)  # æ¸©åº¦ 0-260Â°C
        
        # åˆ€ç‰‡ç”µæ± æ•°æ®
        y_v_blade = 5.2 + 0.25 * np.sin(0.05 * x + 0.4) + 0.08 * np.random.randn(600)
        y_t_blade = 120 + 40 * np.sin(0.02 * x + 2.0) + 4 * np.random.randn(600)

        # å·¦å›¾ï¼šä¸‰å…ƒç”µæ± 
        # ç”µå‹æ›²çº¿ï¼ˆå·¦Yè½´ï¼Œä¸»ViewBoxï¼‰
        volt_curve_left = self.waveforms.left_plot.plot(x, y_v_ternary, pen=pg.mkPen(self.current_volt_color, width=self.current_volt_width), name="ç”µå‹")
        self.volt_curves.append(volt_curve_left)

        # æ¸©åº¦æ›²çº¿ï¼ˆå³Yè½´ï¼Œviewbox_tempï¼‰
        temp_curve_left = pg.PlotCurveItem(x, y_t_ternary, pen=pg.mkPen(self.current_temp_color, width=self.current_temp_width), name="æ¸©åº¦")
        self.waveforms.left_plot.viewbox_temp.addItem(temp_curve_left)
        self.temp_curves.append(temp_curve_left)

        # å³å›¾ï¼šåˆ€ç‰‡ç”µæ± 
        # ç”µå‹æ›²çº¿ï¼ˆå·¦Yè½´ï¼‰
        volt_curve_right = self.waveforms.right_plot.plot(x, y_v_blade, pen=pg.mkPen(self.current_volt_color, width=self.current_volt_width), name="ç”µå‹")
        self.volt_curves.append(volt_curve_right)

        # æ¸©åº¦æ›²çº¿ï¼ˆå³Yè½´ï¼‰
        temp_curve_right = pg.PlotCurveItem(x, y_t_blade, pen=pg.mkPen(self.current_temp_color, width=self.current_temp_width), name="æ¸©åº¦")
        self.waveforms.right_plot.viewbox_temp.addItem(temp_curve_right)
        self.temp_curves.append(temp_curve_right)

        # åˆå§‹åŒ– KPI
        self.ternary_voltage_kpi.set_value("0.00")
        self.ternary_temp_kpi.set_value("0.00")
        self.blade_voltage_kpi.set_value("0.00")
        self.blade_temp_kpi.set_value("0.00")

    def _create_title_bar(self) -> QWidget:
        """åˆ›å»ºé¡¶éƒ¨æ ‡é¢˜æ ï¼ˆç§‘æŠ€é£æ ¼ï¼‰ã€‚"""
        title_container = QWidget()
        title_container.setObjectName("titleBar")
        title_container.setFixedHeight(60)
        title_container.setStyleSheet("""
            QWidget#titleBar {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #0a1a35, stop:0.5 #0c1f3f, stop:1 #0a1a35);
                border-bottom: 2px solid #1a4d7d;
            }
            QLabel#mainTitle {
                color: #6dd5ed;
                font-size: 20px;
                font-weight: bold;
                letter-spacing: 2px;
            }
            QPushButton#titleBtn {
                background-color: transparent;
                color: #6dd5ed;
                border: 1px solid #2b5a8a;
                border-radius: 4px;
                padding: 6px 16px;
                font-size: 13px;
            }
            QPushButton#titleBtn:hover {
                background-color: #1a3a5f;
                border-color: #4a8fdd;
            }
        """)

        layout = QHBoxLayout(title_container)
        layout.setContentsMargins(20, 10, 20, 10)

        layout.addStretch()
        
        # ä¸»æ ‡é¢˜
        title = QLabel("ç”µæ± ç”µå‹ä¸æ¸©å‡é‡‡é›†è½¯ä»¶")
        title.setObjectName("mainTitle")
        layout.addWidget(title)

        layout.addStretch()

        # å¸®åŠ©æŒ‰é’®
        btn_help = QPushButton("â“ Help")
        btn_help.setObjectName("titleBtn")
        btn_help.clicked.connect(self._show_help)
        layout.addWidget(btn_help)

        # é€€å‡ºæŒ‰é’®
        btn_exit = QPushButton("âœ– Exit")
        btn_exit.setObjectName("titleBtn")
        btn_exit.clicked.connect(self.close)
        layout.addWidget(btn_exit)

        return title_container

    def _show_help(self) -> None:
        """æ˜¾ç¤ºå¸®åŠ©ä¿¡æ¯ã€‚"""
        from PySide6.QtWidgets import QMessageBox
        QMessageBox.information(
            self,
            "å¸®åŠ©",
            "ç”µæ± ç”µå‹ä¸æ¸©å‡åˆ†æè½¯ä»¶\n\n"
            "åŠŸèƒ½ï¼š\n"
            "â€¢ ä¸‰å…ƒç”µæ± ä¸åˆ€ç‰‡ç”µæ± æ¸©å‡æ¯”å¯¹\n"
            "â€¢ ç”µæ± å‹é™é‡‡é›†åˆ†æ\n"
            "â€¢ mX+b çº¿æ€§æ ¡å‡†\n"
            "â€¢ mAh å®¹é‡æµ‹è¯•\n\n"
            "å¼€å‘å®šåˆ¶è¯·è®¤å‡†è¿…å±¿ç§‘æŠ€\n"
            "https://www.xunyutek.com\n"
            
        )

    def update_voltage_color(self, color_hex: str) -> None:
        """æ›´æ–°ç”µå‹æ›²çº¿é¢œè‰²å’Œå·¦Yè½´é¢œè‰²ã€‚"""
        self.current_volt_color = color_hex  # æ›´æ–°å½“å‰é¢œè‰²
        
        # æ›´æ–°æ›²çº¿é¢œè‰²
        for curve in self.volt_curves:
            curve.setPen(pg.mkPen(color_hex, width=2))
        
        # æ›´æ–°å·¦Yè½´ï¼ˆç”µå‹è½´ï¼‰é¢œè‰²
        self.waveforms.left_plot.setLabel('left', 'ç”µå‹ V', color=color_hex, **{'font-size': '11pt'})
        self.waveforms.left_plot.getAxis('left').setPen(pg.mkPen(color_hex, width=2))
        self.waveforms.left_plot.getAxis('left').setTextPen(color_hex)
        
        self.waveforms.right_plot.setLabel('left', 'ç”µå‹ V', color=color_hex, **{'font-size': '11pt'})
        self.waveforms.right_plot.getAxis('left').setPen(pg.mkPen(color_hex, width=2))
        self.waveforms.right_plot.getAxis('left').setTextPen(color_hex)
        
        self.statusBar().showMessage(f"ç”µå‹æ›²çº¿å’Œåæ ‡è½´é¢œè‰²å·²æ›´æ–°ä¸º {color_hex}")
    
    def update_temp_color(self, color_hex: str) -> None:
        """æ›´æ–°æ¸©åº¦æ›²çº¿é¢œè‰²å’Œå³Yè½´é¢œè‰²ã€‚"""
        self.current_temp_color = color_hex  # æ›´æ–°å½“å‰é¢œè‰²
        
        # æ›´æ–°æ›²çº¿é¢œè‰²
        for curve in self.temp_curves:
            curve.setPen(pg.mkPen(color_hex, width=2))
        
        # æ›´æ–°å³Yè½´ï¼ˆæ¸©åº¦è½´ï¼‰é¢œè‰²
        self.waveforms.left_plot.setLabel('right', 'æ¸©åº¦ T', units='Â°C', color=color_hex, **{'font-size': '11pt'})
        self.waveforms.left_plot.getAxis('right').setPen(pg.mkPen(color_hex, width=2))
        self.waveforms.left_plot.getAxis('right').setTextPen(color_hex)
        
        self.waveforms.right_plot.setLabel('right', 'æ¸©åº¦ T', units='Â°C', color=color_hex, **{'font-size': '11pt'})
        self.waveforms.right_plot.getAxis('right').setPen(pg.mkPen(color_hex, width=2))
        self.waveforms.right_plot.getAxis('right').setTextPen(color_hex)
        
        self.statusBar().showMessage(f"æ¸©åº¦æ›²çº¿å’Œåæ ‡è½´é¢œè‰²å·²æ›´æ–°ä¸º {color_hex}")

    def update_voltage_width(self, width: int) -> None:
        """æ›´æ–°ç”µå‹æ›²çº¿çº¿å®½"""
        self.current_volt_width = width

        # æ›´æ–°æ›²çº¿çº¿å®½
        for curve in self.volt_curves:
            curve.setPen(pg.mkPen(self.current_volt_color, width=width))

        self.statusBar().showMessage(f"ç”µå‹æ›²çº¿çº¿å®½å·²æ›´æ–°ä¸º {width} åƒç´ ")

    def update_temp_width(self, width: int) -> None:
        """æ›´æ–°æ¸©åº¦æ›²çº¿çº¿å®½"""
        self.current_temp_width = width

        # æ›´æ–°æ›²çº¿çº¿å®½
        for curve in self.temp_curves:
            curve.setPen(pg.mkPen(self.current_temp_color, width=width))

        self.statusBar().showMessage(f"æ¸©åº¦æ›²çº¿çº¿å®½å·²æ›´æ–°ä¸º {width} åƒç´ ")

    def _show_mx_plus_b_dialog(self) -> None:
        """æ˜¾ç¤ºmX+bæ ¡å‡†å¯¹è¯æ¡†ã€‚"""
        dialog = MXPlusBDialog(self)
        dialog.exec()

    def _show_mah_test_dialog(self) -> None:
        """æ˜¾ç¤ºmAhå®¹é‡æµ‹è¯•å¯¹è¯æ¡†ã€‚"""
        dialog = MAHTestDialog(self)
        dialog.exec()

    def _on_data_acquired(self, timestamp: float, data: dict) -> None:
        """å¤„ç†ä»é‡‡é›†çº¿ç¨‹æ¥æ”¶åˆ°çš„æ•°æ®ï¼ˆçœŸå®è®¾å¤‡æ•°æ®ï¼‰

        Args:
            timestamp: æ—¶é—´æˆ³ï¼ˆç§’ï¼‰
            data: é€šé“æ•°æ®å­—å…¸
        """
        # è°ƒè¯•è¾“å‡ºï¼ˆæ¯10ä¸ªæ•°æ®ç‚¹è¾“å‡ºä¸€æ¬¡ï¼‰
        if self.data_index % 10 == 0:
            print(f"\nğŸ“Š æ•°æ®æ˜ å°„è°ƒè¯• (ç¬¬ {self.data_index} ä¸ªæ•°æ®ç‚¹):")
            print(f"  é…ç½®:")
            print(f"    ä¸‰å…ƒç”µæ± ç”µå‹ â† {self.channel_config['ternary_voltage']['channel']}")
            print(f"    ä¸‰å…ƒç”µæ± æ¸©åº¦ â† {self.channel_config['ternary_temp']['channel']}")
            print(f"    åˆ€ç‰‡ç”µæ± ç”µå‹ â† {self.channel_config['blade_voltage']['channel']}")
            print(f"    åˆ€ç‰‡ç”µæ± æ¸©åº¦ â† {self.channel_config['blade_temp']['channel']}")
            print(f"  æ¥æ”¶åˆ°çš„æ•°æ®: {data}")
            print(f"  æå–ç»“æœ:")
            print(f"    ä¸‰å…ƒç”µæ± ç”µå‹: {data.get(self.channel_config['ternary_voltage']['channel'], 0.0)}")
            print(f"    ä¸‰å…ƒç”µæ± æ¸©åº¦: {data.get(self.channel_config['ternary_temp']['channel'], 0.0)}")
            print(f"    åˆ€ç‰‡ç”µæ± ç”µå‹: {data.get(self.channel_config['blade_voltage']['channel'], 0.0)}")
            print(f"    åˆ€ç‰‡ç”µæ± æ¸©åº¦: {data.get(self.channel_config['blade_temp']['channel'], 0.0)}")

        # æå–åŸå§‹æ•°æ®ï¼ˆä½¿ç”¨é€šé“åç§°ï¼‰
        v_ternary_raw = data.get(self.channel_config['ternary_voltage']['channel'], 0.0)
        t_ternary_raw = data.get(self.channel_config['ternary_temp']['channel'], 0.0)
        v_blade_raw = data.get(self.channel_config['blade_voltage']['channel'], 0.0)
        t_blade_raw = data.get(self.channel_config['blade_temp']['channel'], 0.0)

        # æ£€æµ‹BURNOUTå¼‚å¸¸ï¼ˆæ¸©åº¦è¶…å‡ºé‡ç¨‹1.5å€è§†ä¸ºå¼‚å¸¸ï¼‰
        ternary_temp_range = self.channel_config['ternary_temp']['range']
        blade_temp_range = self.channel_config['blade_temp']['range']

        # å¦‚æœæ¸©åº¦è¶…å‡ºé‡ç¨‹1.5å€ï¼Œåˆ¤å®šä¸ºBURNOUTå¼‚å¸¸ï¼Œè®¾ç½®ä¸º0
        if abs(t_ternary_raw) > ternary_temp_range * 1.5:
            print(f"âš ï¸ ä¸‰å…ƒç”µæ± æ¸©åº¦å¼‚å¸¸ (BURNOUT): {t_ternary_raw:.2f}Â°C (é‡ç¨‹: {ternary_temp_range}Â°C)")
            t_ternary_raw = 0.0

        if abs(t_blade_raw) > blade_temp_range * 1.5:
            print(f"âš ï¸ åˆ€ç‰‡ç”µæ± æ¸©åº¦å¼‚å¸¸ (BURNOUT): {t_blade_raw:.2f}Â°C (é‡ç¨‹: {blade_temp_range}Â°C)")
            t_blade_raw = 0.0

        # æ£€æµ‹ç”µå‹å¼‚å¸¸ï¼ˆç”µå‹è¶…å‡ºé‡ç¨‹1.5å€è§†ä¸ºå¼‚å¸¸ï¼‰
        ternary_volt_range = self.channel_config['ternary_voltage']['range']
        blade_volt_range = self.channel_config['blade_voltage']['range']

        if abs(v_ternary_raw) > ternary_volt_range * 1.5:
            print(f"âš ï¸ ä¸‰å…ƒç”µæ± ç”µå‹å¼‚å¸¸: {v_ternary_raw:.2f}V (é‡ç¨‹: {ternary_volt_range}V)")
            v_ternary_raw = 0.0

        if abs(v_blade_raw) > blade_volt_range * 1.5:
            print(f"âš ï¸ åˆ€ç‰‡ç”µæ± ç”µå‹å¼‚å¸¸: {v_blade_raw:.2f}V (é‡ç¨‹: {blade_volt_range}V)")
            v_blade_raw = 0.0

        # åº”ç”¨mX+bæ ¡å‡†
        v_ternary = self.analysis_engine.apply_calibration("ternary", "voltage", v_ternary_raw)
        t_ternary = self.analysis_engine.apply_calibration("ternary", "temp", t_ternary_raw)
        v_blade = self.analysis_engine.apply_calibration("blade", "voltage", v_blade_raw)
        t_blade = self.analysis_engine.apply_calibration("blade", "temp", t_blade_raw)

        # æ·»åŠ åˆ°åˆ†æå¼•æ“
        self.analysis_engine.add_data_point(v_ternary, t_ternary, v_blade, t_blade, timestamp)

        # æ·»åŠ åˆ°ç¼“å†²åŒº
        self.x_data.append(timestamp)
        self.ternary_volt_data.append(v_ternary)
        self.ternary_temp_data.append(t_ternary)
        self.blade_volt_data.append(v_blade)
        self.blade_temp_data.append(t_blade)

        # é™åˆ¶æ•°æ®ç‚¹æ•°é‡ï¼ˆæ»šåŠ¨æ˜¾ç¤ºï¼‰
        if len(self.x_data) > self.max_points:
            self.x_data.pop(0)
            self.ternary_volt_data.pop(0)
            self.ternary_temp_data.pop(0)
            self.blade_volt_data.pop(0)
            self.blade_temp_data.pop(0)

        # æ›´æ–°æ›²çº¿
        if len(self.volt_curves) >= 2 and len(self.temp_curves) >= 2:
            self.volt_curves[0].setData(self.x_data, self.ternary_volt_data)
            self.temp_curves[0].setData(self.x_data, self.ternary_temp_data)
            self.volt_curves[1].setData(self.x_data, self.blade_volt_data)
            self.temp_curves[1].setData(self.x_data, self.blade_temp_data)

        # æ›´æ–°KPIæ˜¾ç¤º
        self.ternary_voltage_kpi.set_value(f"{v_ternary:.2f}")
        self.ternary_temp_kpi.set_value(f"{t_ternary:.2f}")
        self.blade_voltage_kpi.set_value(f"{v_blade:.2f}")
        self.blade_temp_kpi.set_value(f"{t_blade:.2f}")

    def _on_acquisition_error(self, error_msg: str) -> None:
        """å¤„ç†é‡‡é›†çº¿ç¨‹çš„é”™è¯¯

        Args:
            error_msg: é”™è¯¯æ¶ˆæ¯
        """
        print(f"âš ï¸ é‡‡é›†é”™è¯¯: {error_msg}")
        # å¯ä»¥é€‰æ‹©æ˜¾ç¤ºåœ¨çŠ¶æ€æ æˆ–å¼¹çª—
        # self.statusBar().showMessage(f"é‡‡é›†é”™è¯¯: {error_msg}")

    def _on_acquisition_status(self, status_msg: str) -> None:
        """å¤„ç†é‡‡é›†çº¿ç¨‹çš„çŠ¶æ€å˜åŒ–

        Args:
            status_msg: çŠ¶æ€æ¶ˆæ¯
        """
        print(f"â„¹ï¸ é‡‡é›†çŠ¶æ€: {status_msg}")

    def _update_waveform_virtual(self) -> None:
        """å®šæ—¶æ›´æ–°æ³¢å½¢ï¼ˆè™šæ‹Ÿæ•°æ®æ¨¡å¼ï¼‰"""
        # ä¿®æ­£æ—¶é—´æˆ³è®¡ç®—ï¼š100msé—´éš” = 0.1ç§’
        t = self.data_index * (self.update_interval_ms / 1000.0)

        # ç”Ÿæˆè™šæ‹Ÿæ•°æ®
        v_ternary, t_ternary, v_blade, t_blade = self._generate_virtual_data(t)

        # æ·»åŠ åˆ°åˆ†æå¼•æ“
        self.analysis_engine.add_data_point(v_ternary, t_ternary, v_blade, t_blade, t)

        # æ·»åŠ åˆ°ç¼“å†²åŒº
        self.x_data.append(t)
        self.ternary_volt_data.append(v_ternary)
        self.ternary_temp_data.append(t_ternary)
        self.blade_volt_data.append(v_blade)
        self.blade_temp_data.append(t_blade)

        # é™åˆ¶æ•°æ®ç‚¹æ•°é‡ï¼ˆæ»šåŠ¨æ˜¾ç¤ºï¼‰
        if len(self.x_data) > self.max_points:
            self.x_data.pop(0)
            self.ternary_volt_data.pop(0)
            self.ternary_temp_data.pop(0)
            self.blade_volt_data.pop(0)
            self.blade_temp_data.pop(0)

        # æ›´æ–°æ›²çº¿
        if len(self.volt_curves) >= 2 and len(self.temp_curves) >= 2:
            self.volt_curves[0].setData(self.x_data, self.ternary_volt_data)
            self.temp_curves[0].setData(self.x_data, self.ternary_temp_data)
            self.volt_curves[1].setData(self.x_data, self.blade_volt_data)
            self.temp_curves[1].setData(self.x_data, self.blade_temp_data)

        # æ›´æ–°KPIæ˜¾ç¤º
        self.ternary_voltage_kpi.set_value(f"{v_ternary:.2f}")
        self.ternary_temp_kpi.set_value(f"{t_ternary:.2f}")
        self.blade_voltage_kpi.set_value(f"{v_blade:.2f}")
        self.blade_temp_kpi.set_value(f"{t_blade:.2f}")

        self.data_index += 1
    
    def _generate_virtual_data(self, t: float) -> tuple[float, float, float, float]:
        """ç”Ÿæˆè™šæ‹Ÿæ•°æ®ï¼ˆè®¾å¤‡æœªè¿æ¥æ—¶ä½¿ç”¨ï¼‰"""
        v_ternary = 5 + 0.3 * np.sin(0.05 * t) + 0.1 * np.random.randn()
        t_ternary = 130 + 50 * np.sin(0.02 * t + 1.2) + 5 * np.random.randn()
        v_blade = 5.2 + 0.25 * np.sin(0.05 * t + 0.4) + 0.08 * np.random.randn()
        t_blade = 120 + 40 * np.sin(0.02 * t + 2.0) + 4 * np.random.randn()
        return v_ternary, t_ternary, v_blade, t_blade
    
    # æ§½å‡½æ•°ï¼šå¼€å§‹/åœæ­¢æ•°æ®é‡‡é›†
    def _on_start(self) -> None:
        """å¼€å§‹æ•°æ®é‡‡é›†ï¼ˆçœŸå®è®¾å¤‡æˆ–è™šæ‹Ÿæ•°æ®ï¼‰"""
        if not self.is_running:
            self.is_running = True
            self.data_index = 0
            self.x_data.clear()
            self.ternary_volt_data.clear()
            self.ternary_temp_data.clear()
            self.blade_volt_data.clear()
            self.blade_temp_data.clear()

            # æ¸…ç©ºåˆ†æå¼•æ“æ•°æ®
            self.analysis_engine.clear_data()

            # å¦‚æœè®¾å¤‡å·²è¿æ¥ï¼Œä½¿ç”¨åå°çº¿ç¨‹é‡‡é›†çœŸå®æ•°æ®
            if self.device_connected and self.device_client:
                # è·å–é€šé“åˆ—è¡¨ï¼ˆæŒ‰é€šé“å·æ’åºï¼Œç¡®ä¿ä¸è®¾å¤‡å†…éƒ¨é¡ºåºä¸€è‡´ï¼‰
                channel_map = {
                    self.channel_config['ternary_voltage']['channel']: 'ternary_voltage',
                    self.channel_config['ternary_temp']['channel']: 'ternary_temp',
                    self.channel_config['blade_voltage']['channel']: 'blade_voltage',
                    self.channel_config['blade_temp']['channel']: 'blade_temp',
                }
                self._current_channels = sorted(channel_map.keys())  # ä¿å­˜å½“å‰é€šé“åˆ—è¡¨

                print(f"\nğŸš€ å¼€å§‹æ•°æ®é‡‡é›†...")
                print(f"ğŸ“‹ é€šé“è¯»å–é¡ºåºï¼ˆå·²æ’åºï¼‰: {self._current_channels}")

                # ã€å…³é”®ä¿®å¤ã€‘æ¯æ¬¡å¼€å§‹å‰éƒ½å¼ºåˆ¶é‡æ–°é…ç½®é€šé“ï¼Œç¡®ä¿è®¾å¤‡æ•°æ®ç¼“å†²åŒºé¡ºåºæ­£ç¡®
                # è¿™å¯ä»¥è§£å†³å¤šæ¬¡å¼€å§‹åœæ­¢åæ•°æ®é”™ä¹±çš„é—®é¢˜
                print("ğŸ”§ é‡æ–°é…ç½®é€šé“ï¼ˆç¡®ä¿æ•°æ®é¡ºåºæ­£ç¡®ï¼‰...")

                # å‡†å¤‡é€šé“é…ç½®
                channel_configs = []
                for ch in self._current_channels:
                    for key in ['ternary_voltage', 'ternary_temp', 'blade_voltage', 'blade_temp']:
                        if self.channel_config[key]['channel'] == ch:
                            channel_configs.append(self.channel_config[key])
                            break

                # ç¦ç”¨å¼€å§‹æŒ‰é’®ï¼Œé˜²æ­¢é‡å¤ç‚¹å‡»
                self.control.btn_start.setEnabled(False)
                self.control.btn_start.setText("é…ç½®ä¸­...")
                self.statusBar().showMessage("æ­£åœ¨é…ç½®é€šé“...")

                # ä½¿ç”¨åå°çº¿ç¨‹é…ç½®é€šé“ï¼ˆé¿å…UIå¡é¡¿ï¼‰
                self._config_worker = DeviceConfigWorker(
                    device_client=self.device_client,
                    channels=self._current_channels,
                    channel_configs=channel_configs
                )
                self._config_worker.progress_updated.connect(self._on_config_progress)
                self._config_worker.config_finished.connect(self._on_config_finished)
                self._config_worker.start()
            else:
                # ä½¿ç”¨è™šæ‹Ÿæ•°æ®æ¨¡å¼
                self.update_timer.start(self.update_interval_ms)
                self.statusBar().showMessage("è™šæ‹Ÿæ•°æ®é‡‡é›†è¿›è¡Œä¸­...")
                self.control.btn_start.setText("åœæ­¢")
                self.control.btn_start.clicked.disconnect()
                self.control.btn_start.clicked.connect(self._on_stop)

    def _on_config_progress(self, percent: int, message: str) -> None:
        """é…ç½®è¿›åº¦æ›´æ–°"""
        self.statusBar().showMessage(f"é…ç½®ä¸­ ({percent}%): {message}")

    def _on_config_finished(self, success: bool, message: str) -> None:
        """é…ç½®å®Œæˆå›è°ƒ"""
        print(message)

        if success:
            # æ›´æ–°é…ç½®æ ‡è®°
            self._channels_configured = True
            current_config_hash = str(sorted(self._current_channels)) + str(self.channel_config)
            self._last_channel_config_hash = current_config_hash

            # é‡ç½®è®¾å¤‡çš„è°ƒè¯•æ‰“å°æ ‡è®°ï¼Œç¡®ä¿ä¸‹æ¬¡èƒ½æ‰“å°è°ƒè¯•ä¿¡æ¯
            if hasattr(self.device_client, '_debug_printed'):
                delattr(self.device_client, '_debug_printed')

            # åˆ›å»ºå¹¶å¯åŠ¨é‡‡é›†çº¿ç¨‹
            self.acquisition_thread = DataAcquisitionThread(
                device_client=self.device_client,
                channels=self._current_channels,
                interval_ms=self.update_interval_ms
            )

            # è¿æ¥ä¿¡å·
            self.acquisition_thread.data_acquired.connect(self._on_data_acquired)
            self.acquisition_thread.error_occurred.connect(self._on_acquisition_error)
            self.acquisition_thread.status_changed.connect(self._on_acquisition_status)

            # å¯åŠ¨çº¿ç¨‹
            self.acquisition_thread.start()

            self.statusBar().showMessage("âœ“ çœŸå®è®¾å¤‡æ•°æ®é‡‡é›†è¿›è¡Œä¸­...")
            self.control.btn_start.setEnabled(True)
            self.control.btn_start.setText("åœæ­¢")
            self.control.btn_start.clicked.disconnect()
            self.control.btn_start.clicked.connect(self._on_stop)
        else:
            # é…ç½®å¤±è´¥ï¼Œæ¢å¤æŒ‰é’®çŠ¶æ€
            self.is_running = False
            self.control.btn_start.setEnabled(True)
            self.control.btn_start.setText("å¼€å§‹")
            self.statusBar().showMessage(f"é…ç½®å¤±è´¥: {message}")

    def _on_stop(self) -> None:
        """åœæ­¢æ•°æ®é‡‡é›†"""
        if self.is_running:
            self.is_running = False
            print("\nğŸ›‘ åœæ­¢æ•°æ®é‡‡é›†...")

            # ç¦ç”¨æŒ‰é’®ï¼Œé˜²æ­¢é‡å¤ç‚¹å‡»
            self.control.btn_start.setEnabled(False)
            self.control.btn_start.setText("åœæ­¢ä¸­...")

            # 1. å…ˆæ–­å¼€ä¿¡å·è¿æ¥ï¼Œé˜²æ­¢åœæ­¢è¿‡ç¨‹ä¸­è¿˜æœ‰æ•°æ®å›è°ƒ
            if self.acquisition_thread:
                print("   æ–­å¼€ä¿¡å·è¿æ¥...")
                try:
                    self.acquisition_thread.data_acquired.disconnect(self._on_data_acquired)
                    self.acquisition_thread.error_occurred.disconnect(self._on_acquisition_error)
                    self.acquisition_thread.status_changed.disconnect(self._on_acquisition_status)
                except Exception:
                    pass  # ä¿¡å·å¯èƒ½å·²æ–­å¼€

            # 2. ã€é‡è¦ã€‘å…ˆåœæ­¢è®¾å¤‡é‡‡é›†ï¼Œå†åœæ­¢çº¿ç¨‹
            # ç¡®ä¿è®¾å¤‡åœæ­¢åï¼Œçº¿ç¨‹ä¸ä¼šç»§ç»­è¯»å–åˆ°æ–°æ•°æ®
            if self.device_connected and self.device_client:
                print("   å‘é€ :STOP å‘½ä»¤åˆ°è®¾å¤‡...")
                self._stop_worker = DeviceStopWorker(self.device_client)
                self._stop_worker.stop_finished.connect(self._on_device_stopped)
                self._stop_worker.start()
            else:
                # æ— è®¾å¤‡è¿æ¥ï¼Œç›´æ¥åœæ­¢çº¿ç¨‹
                self._stop_acquisition_thread()
                self._finalize_stop()

            # 3. åœæ­¢è™šæ‹Ÿæ•°æ®å®šæ—¶å™¨
            if self.update_timer.isActive():
                self.update_timer.stop()
                print("   âœ“ è™šæ‹Ÿæ•°æ®å®šæ—¶å™¨å·²åœæ­¢")

    def _stop_acquisition_thread(self) -> None:
        """åœæ­¢é‡‡é›†çº¿ç¨‹"""
        if self.acquisition_thread:
            print("   åœæ­¢é‡‡é›†çº¿ç¨‹...")
            try:
                if self.acquisition_thread.is_running():
                    self.acquisition_thread.stop()
                print("   âœ“ é‡‡é›†çº¿ç¨‹å·²åœæ­¢")
            except Exception as e:
                print(f"   âš ï¸ çº¿ç¨‹åœæ­¢å¼‚å¸¸: {e}")
            finally:
                self.acquisition_thread = None

    def _on_device_stopped(self, success: bool, message: str) -> None:
        """è®¾å¤‡åœæ­¢å›è°ƒ"""
        print(f"   {message}")
        # è®¾å¤‡åœæ­¢åï¼Œå†åœæ­¢é‡‡é›†çº¿ç¨‹
        self._stop_acquisition_thread()
        self._finalize_stop()

        if not success:
            QMessageBox.warning(self, "åœæ­¢è­¦å‘Š", message)

    def _finalize_stop(self) -> None:
        """å®Œæˆåœæ­¢æ“ä½œ"""
        self.statusBar().showMessage("æ•°æ®é‡‡é›†å·²åœæ­¢")
        print("âœ“ æ•°æ®é‡‡é›†å·²å®Œå…¨åœæ­¢\n")

        self.control.btn_start.setEnabled(True)
        self.control.btn_start.setText("å¼€å§‹")
        try:
            self.control.btn_start.clicked.disconnect()
        except Exception:
            pass
        self.control.btn_start.clicked.connect(self._on_start)
    
    def _show_device_connect_dialog(self) -> None:
        """æ˜¾ç¤ºè®¾å¤‡è¿æ¥å¯¹è¯æ¡†"""
        # ä¼ é€’ä¿å­˜çš„è¿æ¥é…ç½®ä½œä¸ºé»˜è®¤å€¼
        dialog = DeviceConnectDialog(self, saved_config=self.saved_connection_config)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            params = dialog.get_connection_params()
            self._connect_to_device(params)
    
    def _connect_to_device(self, params: dict) -> None:
        """è¿æ¥åˆ°LR8450è®¾å¤‡ï¼ˆæ”¯æŒTCPå’ŒUSBï¼‰

        Args:
            params: è¿æ¥å‚æ•°å­—å…¸ï¼ŒåŒ…å«ï¼š
                - connection_type: "TCP" æˆ– "USB"
                - ip_address: TCP IPåœ°å€
                - port: TCPç«¯å£
                - com_port: COMç«¯å£
        """
        connection_type = params['connection_type']

        if connection_type == "TCP":
            ip = params['ip_address']
            port = params['port']
            connection_info = f"IPåœ°å€: {ip}\nç«¯å£: {port}"
            progress_text = f"æ­£åœ¨é€šè¿‡TCPè¿æ¥è®¾å¤‡ {ip}:{port}..."
        else:  # USB
            com_port = params['com_port']
            connection_info = f"COMç«¯å£: {com_port}"
            progress_text = f"æ­£åœ¨é€šè¿‡USBè¿æ¥è®¾å¤‡ {com_port}..."

        # æ˜¾ç¤ºåŠ è½½å¯¹è¯æ¡†
        from PySide6.QtWidgets import QProgressDialog
        progress = QProgressDialog(progress_text, None, 0, 0, self)
        progress.setWindowTitle("è¿æ¥ä¸­")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        progress.show()
        QApplication.processEvents()

        try:
            # æ–­å¼€æ—§è¿æ¥
            if self.device_client:
                self.device_client.disconnect()

            # åˆ›å»ºæ–°å®¢æˆ·ç«¯
            self.device_client = LR8450Client(
                connection_type=connection_type,
                ip_address=params['ip_address'],
                port=params['port'],
                com_port=params['com_port']
            )

            # è¿æ¥
            if self.device_client.connect():
                self.device_connected = True
                conn_method = "TCP/IP" if connection_type == "TCP" else "USBä¸²å£"

                # æ›´æ–°è¿›åº¦å¯¹è¯æ¡† - æ£€æµ‹æ¨¡å—
                progress.setLabelText(f"âœ“ è®¾å¤‡å·²è¿æ¥ï¼Œæ­£åœ¨æ£€æµ‹å·²å®‰è£…çš„æ¨¡å—...")
                QApplication.processEvents()

                # æ£€æµ‹å·²å®‰è£…çš„æ¨¡å—
                self.installed_modules = self.device_client.detect_installed_modules()

                # å¦‚æœæ²¡æœ‰æ£€æµ‹åˆ°ä»»ä½•æ¨¡å—ï¼Œæç¤ºç”¨æˆ·
                if not self.installed_modules:
                    progress.close()
                    QMessageBox.warning(
                        self,
                        "æœªæ£€æµ‹åˆ°æ¨¡å—",
                        f"è®¾å¤‡å·²é€šè¿‡{conn_method}è¿æ¥æˆåŠŸï¼Œä½†æœªæ£€æµ‹åˆ°ä»»ä½•å·²å®‰è£…çš„æ¨¡å—ã€‚\n\n"
                        f"è¯·æ£€æŸ¥ï¼š\n"
                        f"1. æ¨¡å—æ˜¯å¦æ­£ç¡®æ’å…¥è®¾å¤‡çš„UNITæ’æ§½\n"
                        f"2. è®¾å¤‡æ˜¯å¦å·²è¯†åˆ«æ¨¡å—\n\n"
                        f"è®¾å¤‡å·²è¿æ¥ï¼Œä½†æ— æ³•è¿›è¡Œæ•°æ®é‡‡é›†ã€‚"
                    )
                    self.statusBar().showMessage("âš ï¸ è®¾å¤‡å·²è¿æ¥ï¼Œä½†æœªæ£€æµ‹åˆ°æ¨¡å—")
                    return

                # æ›´æ–°è¿›åº¦å¯¹è¯æ¡† - é…ç½®é€šé“
                progress.setLabelText(f"âœ“ æ£€æµ‹åˆ°æ¨¡å— UNIT{self.installed_modules}ï¼Œæ­£åœ¨é…ç½®é€šé“...")
                QApplication.processEvents()

                # å‡†å¤‡é€šé“é…ç½®ï¼ˆæŒ‰é€šé“å·æ’åºï¼Œç¡®ä¿ä¸è®¾å¤‡å†…éƒ¨é¡ºåºä¸€è‡´ï¼‰
                channel_configs_dict = {
                    self.channel_config['ternary_voltage']['channel']: {
                        'channel': self.channel_config['ternary_voltage']['channel'],
                        'type': self.channel_config['ternary_voltage']['type'],
                        'range': self.channel_config['ternary_voltage']['range'],
                    },
                    self.channel_config['ternary_temp']['channel']: {
                        'channel': self.channel_config['ternary_temp']['channel'],
                        'type': self.channel_config['ternary_temp']['type'],
                        'range': self.channel_config['ternary_temp']['range'],
                        'thermocouple': self.channel_config['ternary_temp']['thermocouple'],
                        'int_ext': self.channel_config['ternary_temp']['int_ext'],
                    },
                    self.channel_config['blade_voltage']['channel']: {
                        'channel': self.channel_config['blade_voltage']['channel'],
                        'type': self.channel_config['blade_voltage']['type'],
                        'range': self.channel_config['blade_voltage']['range'],
                    },
                    self.channel_config['blade_temp']['channel']: {
                        'channel': self.channel_config['blade_temp']['channel'],
                        'type': self.channel_config['blade_temp']['type'],
                        'range': self.channel_config['blade_temp']['range'],
                        'thermocouple': self.channel_config['blade_temp']['thermocouple'],
                        'int_ext': self.channel_config['blade_temp']['int_ext'],
                    },
                }

                # æŒ‰é€šé“å·æ’åº
                sorted_channels = sorted(channel_configs_dict.keys())
                channel_configs = [channel_configs_dict[ch] for ch in sorted_channels]
                channels = sorted_channels

                print(f"ğŸ“‹ é€šé“é…ç½®é¡ºåºï¼ˆå·²æ’åºï¼‰: {channels}")

                # é…ç½®é€šé“ï¼ˆå…ˆç¦ç”¨æ‰€æœ‰é€šé“ï¼Œç„¶ååªå¯ç”¨éœ€è¦çš„é€šé“ï¼‰
                config_success = self.device_client.configure_channels(
                    channels=channels,
                    disable_others=True,  # å…ˆç¦ç”¨å…¶ä»–é€šé“ï¼Œé˜²æ­¢æ•°æ®é”™ä¹±
                    channel_configs=channel_configs
                )

                # å…³é—­è¿›åº¦å¯¹è¯æ¡†
                progress.close()

                if config_success:
                    # æ ‡è®°é€šé“å·²é…ç½®ï¼Œåç»­å¼€å§‹é‡‡é›†æ—¶æ— éœ€é‡æ–°é…ç½®
                    self._channels_configured = True
                    self._last_channel_config_hash = str(sorted(channels)) + str(self.channel_config)
                    self.statusBar().showMessage(f"âœ“ è®¾å¤‡å·²é€šè¿‡{conn_method}è¿æ¥ï¼Œé€šé“å·²é…ç½®")

                    # æ„å»ºé€šé“é…ç½®ä¿¡æ¯ï¼ˆæŒ‰å®é™…é…ç½®é¡ºåºï¼‰
                    channel_info_lines = []
                    for ch in channels:
                        if ch == self.channel_config['ternary_voltage']['channel']:
                            channel_info_lines.append(
                                f"â€¢ {ch} - ä¸‰å…ƒç”µæ± ç”µå‹ ({self.channel_config['ternary_voltage']['range']}V)"
                            )
                        elif ch == self.channel_config['ternary_temp']['channel']:
                            channel_info_lines.append(
                                f"â€¢ {ch} - ä¸‰å…ƒç”µæ± æ¸©åº¦ ({self.channel_config['ternary_temp']['range']}Â°C, "
                                f"{self.channel_config['ternary_temp']['thermocouple']}å‹)"
                            )
                        elif ch == self.channel_config['blade_voltage']['channel']:
                            channel_info_lines.append(
                                f"â€¢ {ch} - åˆ€ç‰‡ç”µæ± ç”µå‹ ({self.channel_config['blade_voltage']['range']}V)"
                            )
                        elif ch == self.channel_config['blade_temp']['channel']:
                            channel_info_lines.append(
                                f"â€¢ {ch} - åˆ€ç‰‡ç”µæ± æ¸©åº¦ ({self.channel_config['blade_temp']['range']}Â°C, "
                                f"{self.channel_config['blade_temp']['thermocouple']}å‹)"
                            )

                    QMessageBox.information(
                        self,
                        "è¿æ¥æˆåŠŸ",
                        f"æˆåŠŸé€šè¿‡{conn_method}è¿æ¥åˆ°LR8450è®¾å¤‡\n\n"
                        f"{connection_info}\n\n"
                        f"å·²è‡ªåŠ¨é…ç½®ä»¥ä¸‹é€šé“ï¼š\n"
                        + "\n".join(channel_info_lines) + "\n\n"
                        f"ç°åœ¨å¯ä»¥å¼€å§‹æµ‹è¯•äº†ï¼"
                    )
                else:
                    self.statusBar().showMessage(f"âš ï¸ è®¾å¤‡å·²è¿æ¥ï¼Œä½†é€šé“é…ç½®å¤±è´¥")
                    QMessageBox.warning(
                        self,
                        "é€šé“é…ç½®è­¦å‘Š",
                        f"è®¾å¤‡è¿æ¥æˆåŠŸï¼Œä½†é€šé“é…ç½®å¤±è´¥ã€‚\n\n"
                        f"è¯·åœ¨è®¾å¤‡ä¸Šæ‰‹åŠ¨å¯ç”¨ä»¥ä¸‹é€šé“ï¼š\n"
                        f"â€¢ {channels[0]}\n"
                        f"â€¢ {channels[1]}\n"
                        f"â€¢ {channels[2]}\n"
                        f"â€¢ {channels[3]}\n\n"
                        f"æˆ–è€…å°è¯•é‡æ–°è¿æ¥ã€‚"
                    )
            else:
                # å…³é—­è¿›åº¦å¯¹è¯æ¡†
                progress.close()

                self.device_connected = False
                self.statusBar().showMessage("âœ— è®¾å¤‡è¿æ¥å¤±è´¥")

                if connection_type == "TCP":
                    error_msg = (
                        f"æ— æ³•è¿æ¥åˆ°è®¾å¤‡ {ip}:{port}\n\n"
                        f"è¯·æ£€æŸ¥ï¼š\n"
                        f"1. è®¾å¤‡ç”µæºæ˜¯å¦å¼€å¯\n"
                        f"2. ç½‘ç»œè¿æ¥æ˜¯å¦æ­£å¸¸\n"
                        f"3. IPåœ°å€å’Œç«¯å£æ˜¯å¦æ­£ç¡®\n"
                        f"4. ç«¯å£8802ï¼ˆSCPIæ§åˆ¶ç«¯å£ï¼‰"
                    )
                else:
                    error_msg = (
                        f"æ— æ³•è¿æ¥åˆ°è®¾å¤‡ {com_port}\n\n"
                        f"è¯·æ£€æŸ¥ï¼š\n"
                        f"1. è®¾å¤‡ç”µæºæ˜¯å¦å¼€å¯\n"
                        f"2. USBçº¿æ˜¯å¦è¿æ¥\n"
                        f"3. æ˜¯å¦å·²å®‰è£…HIOKI USBé©±åŠ¨\n"
                        f"4. COMç«¯å£æ˜¯å¦æ­£ç¡®"
                    )

                QMessageBox.warning(self, "è¿æ¥å¤±è´¥", error_msg)
        except Exception as e:
            # å…³é—­è¿›åº¦å¯¹è¯æ¡†
            progress.close()

            self.device_connected = False
            self.statusBar().showMessage("âœ— è¿æ¥é”™è¯¯")
            QMessageBox.critical(self, "é”™è¯¯", f"è¿æ¥å¤±è´¥ï¼š{str(e)}")
    
    def _show_channel_config_dialog(self) -> None:
        """æ˜¾ç¤ºé€šé“é…ç½®å¯¹è¯æ¡†"""
        # å¦‚æœè®¾å¤‡å·²è¿æ¥ä¸”æ£€æµ‹åˆ°æ¨¡å—ï¼Œä¼ é€’æ¨¡å—ä¿¡æ¯ï¼›å¦åˆ™æ˜¾ç¤ºæ‰€æœ‰æ¨¡å—
        installed_modules = self.installed_modules if self.installed_modules else None
        dialog = ChannelConfigDialog(
            self,
            current_config=self.channel_config,
            installed_modules=installed_modules
        )

        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.channel_config = dialog.get_config()
            self.statusBar().showMessage("âœ“ é€šé“é…ç½®å·²æ›´æ–°")

            # æ ‡è®°éœ€è¦é‡æ–°é…ç½®é€šé“
            self._channels_configured = False
            self._last_channel_config_hash = None

            # æ›´æ–°Yè½´èŒƒå›´ï¼ˆæ ¹æ®æ–°çš„é‡ç¨‹é…ç½®ï¼‰
            self._update_plot_ranges()

            # æ ¼å¼åŒ–é…ç½®ä¿¡æ¯
            config_info = (
                f"ä¸‰å…ƒç”µæ± ç”µå‹: {self.channel_config['ternary_voltage']['channel']} "
                f"({self.channel_config['ternary_voltage']['range']}V)\n"
                f"ä¸‰å…ƒç”µæ± æ¸©åº¦: {self.channel_config['ternary_temp']['channel']} "
                f"({self.channel_config['ternary_temp']['range']}Â°C, "
                f"{self.channel_config['ternary_temp']['thermocouple']}å‹, "
                f"{self.channel_config['ternary_temp']['int_ext']})\n"
                f"åˆ€ç‰‡ç”µæ± ç”µå‹: {self.channel_config['blade_voltage']['channel']} "
                f"({self.channel_config['blade_voltage']['range']}V)\n"
                f"åˆ€ç‰‡ç”µæ± æ¸©åº¦: {self.channel_config['blade_temp']['channel']} "
                f"({self.channel_config['blade_temp']['range']}Â°C, "
                f"{self.channel_config['blade_temp']['thermocouple']}å‹, "
                f"{self.channel_config['blade_temp']['int_ext']})"
            )

            # å¦‚æœè®¾å¤‡å·²è¿æ¥ï¼Œè¯¢é—®æ˜¯å¦é‡æ–°è¿æ¥è®¾å¤‡ä»¥åº”ç”¨æ–°é…ç½®
            if self.device_connected and self.device_client:
                reply = QMessageBox.question(
                    self,
                    "åº”ç”¨é…ç½®",
                    f"é€šé“é…ç½®å·²æ›´æ–°ï¼š\n\n{config_info}\n\n"
                    f"æ˜¯å¦é‡æ–°è¿æ¥è®¾å¤‡ä»¥åº”ç”¨æ–°é…ç½®ï¼Ÿ\n"
                    f"ï¼ˆæ¨èï¼šè¿™å°†ç¡®ä¿é€šé“é…ç½®æ­£ç¡®ç”Ÿæ•ˆï¼‰",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes
                )

                if reply == QMessageBox.StandardButton.Yes:
                    # é‡æ–°è¿æ¥è®¾å¤‡ä»¥åº”ç”¨æ–°é…ç½®
                    self._reconnect_device_with_new_config()
            else:
                QMessageBox.information(
                    self,
                    "é…ç½®æˆåŠŸ",
                    f"é€šé“é…ç½®å·²æ›´æ–°ï¼š\n\n{config_info}\n\n"
                    f"ä¸‹æ¬¡è¿æ¥è®¾å¤‡æ—¶å°†è‡ªåŠ¨åº”ç”¨æ­¤é…ç½®ã€‚"
                )

            # ä¿å­˜é…ç½®åˆ°æœ¬åœ°æ–‡ä»¶
            self._save_channel_config_to_file()

    def _apply_channel_config_to_device(self) -> None:
        """åº”ç”¨é€šé“é…ç½®åˆ°å·²è¿æ¥çš„è®¾å¤‡"""
        if not self.device_connected or not self.device_client:
            return

        try:
            # æ˜¾ç¤ºåŠ è½½å¯¹è¯æ¡†
            from PySide6.QtWidgets import QProgressDialog
            progress = QProgressDialog("æ­£åœ¨åº”ç”¨é€šé“é…ç½®åˆ°è®¾å¤‡...", None, 0, 0, self)
            progress.setWindowTitle("é…ç½®ä¸­")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(0)
            progress.show()
            QApplication.processEvents()

            # å‡†å¤‡é€šé“é…ç½®ï¼ˆæŒ‰é€šé“å·æ’åºï¼Œç¡®ä¿ä¸è®¾å¤‡å†…éƒ¨é¡ºåºä¸€è‡´ï¼‰
            channel_configs_dict = {
                self.channel_config['ternary_voltage']['channel']: {
                    'channel': self.channel_config['ternary_voltage']['channel'],
                    'type': self.channel_config['ternary_voltage']['type'],
                    'range': self.channel_config['ternary_voltage']['range'],
                },
                self.channel_config['ternary_temp']['channel']: {
                    'channel': self.channel_config['ternary_temp']['channel'],
                    'type': self.channel_config['ternary_temp']['type'],
                    'range': self.channel_config['ternary_temp']['range'],
                    'thermocouple': self.channel_config['ternary_temp']['thermocouple'],
                    'int_ext': self.channel_config['ternary_temp']['int_ext'],
                },
                self.channel_config['blade_voltage']['channel']: {
                    'channel': self.channel_config['blade_voltage']['channel'],
                    'type': self.channel_config['blade_voltage']['type'],
                    'range': self.channel_config['blade_voltage']['range'],
                },
                self.channel_config['blade_temp']['channel']: {
                    'channel': self.channel_config['blade_temp']['channel'],
                    'type': self.channel_config['blade_temp']['type'],
                    'range': self.channel_config['blade_temp']['range'],
                    'thermocouple': self.channel_config['blade_temp']['thermocouple'],
                    'int_ext': self.channel_config['blade_temp']['int_ext'],
                },
            }

            # æŒ‰é€šé“å·æ’åº
            sorted_channels = sorted(channel_configs_dict.keys())
            channel_configs = [channel_configs_dict[ch] for ch in sorted_channels]
            channels = sorted_channels

            print(f"ğŸ“‹ é€šé“é…ç½®é¡ºåºï¼ˆå·²æ’åºï¼‰: {channels}")

            # é…ç½®é€šé“
            config_success = self.device_client.configure_channels(
                channels=channels,
                disable_others=True,
                channel_configs=channel_configs
            )

            progress.close()

            if config_success:
                self.statusBar().showMessage("âœ“ é€šé“é…ç½®å·²æˆåŠŸåº”ç”¨åˆ°è®¾å¤‡")
                QMessageBox.information(
                    self,
                    "é…ç½®æˆåŠŸ",
                    f"é€šé“é…ç½®å·²æˆåŠŸåº”ç”¨åˆ°è®¾å¤‡ï¼\n\n"
                    f"å·²é…ç½®é€šé“ï¼š\n"
                    f"â€¢ {channels[0]} - ä¸‰å…ƒç”µæ± ç”µå‹\n"
                    f"â€¢ {channels[1]} - ä¸‰å…ƒç”µæ± æ¸©åº¦\n"
                    f"â€¢ {channels[2]} - åˆ€ç‰‡ç”µæ± ç”µå‹\n"
                    f"â€¢ {channels[3]} - åˆ€ç‰‡ç”µæ± æ¸©åº¦"
                )
            else:
                QMessageBox.warning(
                    self,
                    "é…ç½®å¤±è´¥",
                    "é€šé“é…ç½®åº”ç”¨å¤±è´¥ï¼Œè¯·å°è¯•é‡æ–°è¿æ¥è®¾å¤‡ã€‚"
                )
        except Exception as e:
            QMessageBox.critical(self, "é”™è¯¯", f"åº”ç”¨é…ç½®æ—¶å‡ºé”™ï¼š{str(e)}")

    def _reconnect_device_with_new_config(self) -> None:
        """é‡æ–°è¿æ¥è®¾å¤‡ä»¥åº”ç”¨æ–°é…ç½®"""
        if not self.device_client:
            return

        try:
            # æ˜¾ç¤ºåŠ è½½å¯¹è¯æ¡†
            from PySide6.QtWidgets import QProgressDialog
            progress = QProgressDialog("æ­£åœ¨é‡æ–°è¿æ¥è®¾å¤‡ä»¥åº”ç”¨æ–°é…ç½®...", None, 0, 0, self)
            progress.setWindowTitle("é‡æ–°è¿æ¥ä¸­")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(0)
            progress.show()
            QApplication.processEvents()

            # ä¿å­˜å½“å‰è¿æ¥å‚æ•°
            connection_type = self.device_client.connection_type
            ip_address = self.device_client.ip_address
            port = self.device_client.port
            com_port = self.device_client.com_port

            # æ–­å¼€å½“å‰è¿æ¥
            progress.setLabelText("æ­£åœ¨æ–­å¼€å½“å‰è¿æ¥...")
            QApplication.processEvents()
            self.device_client.disconnect()
            self.device_connected = False

            import time
            time.sleep(0.5)  # ç­‰å¾…è®¾å¤‡å®Œå…¨æ–­å¼€

            # é‡æ–°è¿æ¥
            progress.setLabelText("æ­£åœ¨é‡æ–°è¿æ¥è®¾å¤‡...")
            QApplication.processEvents()

            # ä½¿ç”¨ä¿å­˜çš„è¿æ¥å‚æ•°é‡æ–°è¿æ¥
            params = {
                'connection_type': connection_type,
                'ip_address': ip_address,
                'port': port,
                'com_port': com_port,
            }

            # å…³é—­è¿›åº¦å¯¹è¯æ¡†ï¼Œè®© _connect_to_device æ˜¾ç¤ºè‡ªå·±çš„è¿›åº¦å¯¹è¯æ¡†
            progress.close()

            # è°ƒç”¨è¿æ¥æ–¹æ³•ï¼ˆä¼šè‡ªåŠ¨åº”ç”¨æ–°é…ç½®ï¼‰
            self._connect_to_device(params)

        except Exception as e:
            progress.close()
            QMessageBox.critical(self, "é”™è¯¯", f"é‡æ–°è¿æ¥è®¾å¤‡æ—¶å‡ºé”™ï¼š{str(e)}")
            self.statusBar().showMessage("âœ— é‡æ–°è¿æ¥å¤±è´¥")

    def _update_plot_ranges(self) -> None:
        """æ ¹æ®é…ç½®çš„é‡ç¨‹æ›´æ–°Yè½´èŒƒå›´"""
        # è·å–ç”µå‹é‡ç¨‹ï¼ˆä¸¤ä¸ªç”µæ± å–æœ€å¤§å€¼ï¼‰
        voltage_range = max(
            self.channel_config['ternary_voltage']['range'],
            self.channel_config['blade_voltage']['range']
        )

        # è·å–æ¸©åº¦é‡ç¨‹ï¼ˆä¸¤ä¸ªç”µæ± å–æœ€å¤§å€¼ï¼‰
        temp_range = max(
            self.channel_config['ternary_temp']['range'],
            self.channel_config['blade_temp']['range']
        )

        # æ›´æ–°ä¸‰å…ƒç”µæ± æ³¢å½¢çš„Yè½´èŒƒå›´ï¼ˆå·¦å›¾ï¼‰
        self.waveforms.left_plot.setYRange(0, voltage_range * 1.1, padding=0)  # å·¦Yè½´ï¼ˆç”µå‹ï¼‰
        self.waveforms.left_plot.viewbox_temp.setYRange(0, temp_range * 1.1)  # å³Yè½´ï¼ˆæ¸©åº¦ï¼‰

        # æ›´æ–°åˆ€ç‰‡ç”µæ± æ³¢å½¢çš„Yè½´èŒƒå›´ï¼ˆå³å›¾ï¼‰
        self.waveforms.right_plot.setYRange(0, voltage_range * 1.1, padding=0)  # å·¦Yè½´ï¼ˆç”µå‹ï¼‰
        self.waveforms.right_plot.viewbox_temp.setYRange(0, temp_range * 1.1)  # å³Yè½´ï¼ˆæ¸©åº¦ï¼‰

        # æ›´æ–°Yè½´æ ‡ç­¾ï¼ˆä¿æŒå½“å‰é¢œè‰²ï¼‰
        self.waveforms.left_plot.setLabel('left', f'ç”µå‹ (V, é‡ç¨‹: {voltage_range}V)', color=self.current_volt_color, **{'font-size': '11pt'})
        self.waveforms.left_plot.setLabel('right', f'æ¸©åº¦ (Â°C, é‡ç¨‹: {temp_range}Â°C)', color=self.current_temp_color, **{'font-size': '11pt'})
        self.waveforms.right_plot.setLabel('left', f'ç”µå‹ (V, é‡ç¨‹: {voltage_range}V)', color=self.current_volt_color, **{'font-size': '11pt'})
        self.waveforms.right_plot.setLabel('right', f'æ¸©åº¦ (Â°C, é‡ç¨‹: {temp_range}Â°C)', color=self.current_temp_color, **{'font-size': '11pt'})

        self.statusBar().showMessage(f"âœ“ Yè½´èŒƒå›´å·²æ›´æ–°ï¼šç”µå‹ 0-{voltage_range}Vï¼Œæ¸©åº¦ 0-{temp_range}Â°C")

    def _save_channel_config_to_file(self) -> None:
        """ä¿å­˜æ‰€æœ‰é…ç½®åˆ°æœ¬åœ°æ–‡ä»¶ï¼ˆé€šé“é…ç½®ã€è¿æ¥é…ç½®ã€äº§å“ä¿¡æ¯ï¼‰"""
        try:
            import json
            import os

            # é…ç½®æ–‡ä»¶è·¯å¾„ï¼ˆä¿å­˜åœ¨ç”¨æˆ·ç›®å½•ï¼‰
            config_dir = os.path.expanduser("~/.battery_analyzer")
            os.makedirs(config_dir, exist_ok=True)
            config_file = os.path.join(config_dir, "app_config.json")

            # æ”¶é›†æ‰€æœ‰é…ç½®
            all_config = {
                # é€šé“é…ç½®
                'channel_config': self.channel_config,

                # è¿æ¥é…ç½®
                'connection': {
                    'ip_address': getattr(self.device_client, 'ip_address', '192.168.2.44') if self.device_client else '192.168.2.44',
                    'port': getattr(self.device_client, 'port', 8802) if self.device_client else 8802,
                    'com_port': getattr(self.device_client, 'com_port', 'COM3') if self.device_client else 'COM3',
                },

                # äº§å“ä¿¡æ¯
                'product_info': {
                    'model': self.control.edit_model.text(),
                    'serial_number': self.control.edit_sn.text(),
                    'tester': self.control.edit_tester.text(),
                },

                # mX+bæ ¡å‡†å‚æ•°
                'calibration': self.analysis_engine.get_calibration_params(),
            }

            # ä¿å­˜é…ç½®
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(all_config, f, indent=4, ensure_ascii=False)

            print(f"âœ“ é…ç½®å·²ä¿å­˜åˆ°: {config_file}")

        except Exception as e:
            print(f"âœ— ä¿å­˜é…ç½®å¤±è´¥: {str(e)}")

    def _load_channel_config_from_file(self) -> None:
        """ä»æœ¬åœ°æ–‡ä»¶åŠ è½½æ‰€æœ‰é…ç½®ï¼ˆé€šé“é…ç½®ã€è¿æ¥é…ç½®ã€äº§å“ä¿¡æ¯ï¼‰"""
        try:
            import json
            import os

            # é…ç½®æ–‡ä»¶è·¯å¾„
            config_dir = os.path.expanduser("~/.battery_analyzer")
            config_file = os.path.join(config_dir, "app_config.json")
            old_config_file = os.path.join(config_dir, "channel_config.json")

            # ä¼˜å…ˆåŠ è½½æ–°ç‰ˆé…ç½®æ–‡ä»¶
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    all_config = json.load(f)

                    # åŠ è½½é€šé“é…ç½®
                    if 'channel_config' in all_config:
                        self.channel_config = all_config['channel_config']

                    # åŠ è½½è¿æ¥é…ç½®ï¼ˆä¿å­˜åˆ°å®ä¾‹å˜é‡ï¼Œè¿æ¥æ—¶ä½¿ç”¨ï¼‰
                    if 'connection' in all_config:
                        self.saved_connection_config = all_config['connection']

                    # åŠ è½½äº§å“ä¿¡æ¯ï¼ˆé˜»æ­¢ä¿¡å·è§¦å‘ï¼Œé¿å…é‡å¤ä¿å­˜ï¼‰
                    if 'product_info' in all_config:
                        product_info = all_config['product_info']

                        # ä¸´æ—¶é˜»æ­¢ä¿¡å·
                        self.control.edit_model.blockSignals(True)
                        self.control.edit_sn.blockSignals(True)
                        self.control.edit_tester.blockSignals(True)

                        self.control.edit_model.setText(product_info.get('model', ''))
                        self.control.edit_sn.setText(product_info.get('serial_number', ''))
                        self.control.edit_tester.setText(product_info.get('tester', ''))

                        # æ¢å¤ä¿¡å·
                        self.control.edit_model.blockSignals(False)
                        self.control.edit_sn.blockSignals(False)
                        self.control.edit_tester.blockSignals(False)

                    # åŠ è½½mX+bæ ¡å‡†å‚æ•°
                    if 'calibration' in all_config:
                        self.analysis_engine.set_calibration_params(all_config['calibration'])

                    print(f"âœ“ å·²åŠ è½½ä¸Šæ¬¡ä¿å­˜çš„é…ç½®")
                    self.statusBar().showMessage("âœ“ å·²åŠ è½½ä¸Šæ¬¡ä¿å­˜çš„é…ç½®")

            # å…¼å®¹æ—§ç‰ˆé…ç½®æ–‡ä»¶ï¼ˆåªæœ‰é€šé“é…ç½®ï¼‰
            elif os.path.exists(old_config_file):
                with open(old_config_file, 'r', encoding='utf-8') as f:
                    loaded_config = json.load(f)

                    # éªŒè¯é…ç½®å®Œæ•´æ€§
                    required_keys = ['ternary_voltage', 'ternary_temp', 'blade_voltage', 'blade_temp']
                    if all(key in loaded_config for key in required_keys):
                        self.channel_config = loaded_config
                        print(f"âœ“ å·²åŠ è½½ä¸Šæ¬¡ä¿å­˜çš„é€šé“é…ç½®ï¼ˆæ—§ç‰ˆæ ¼å¼ï¼‰")
                        self.statusBar().showMessage("âœ“ å·²åŠ è½½ä¸Šæ¬¡ä¿å­˜çš„é€šé“é…ç½®")
                    else:
                        print("âœ— é…ç½®æ–‡ä»¶æ ¼å¼ä¸å®Œæ•´ï¼Œä½¿ç”¨é»˜è®¤é…ç½®")
            else:
                print("â„¹ æœªæ‰¾åˆ°ä¿å­˜çš„é…ç½®æ–‡ä»¶ï¼Œä½¿ç”¨é»˜è®¤é…ç½®")

        except Exception as e:
            print(f"âœ— åŠ è½½é…ç½®å¤±è´¥: {str(e)}")

    def _export_report(self) -> None:
        """å¯¼å‡ºæµ‹è¯•æŠ¥å‘Š"""
        if not self.analysis_engine.ternary_data.timestamps:
            QMessageBox.warning(self, "æ— æ•°æ®", "è¯·å…ˆè¿›è¡Œæµ‹è¯•ï¼Œé‡‡é›†æ•°æ®åå†å¯¼å‡ºæŠ¥å‘Š")
            return

        # ç”ŸæˆæŠ¥å‘Šæ•°æ®
        report_data = self.analysis_engine.generate_report_data()

        # æ˜¾ç¤ºåˆ†æç»“æœ
        temp_compare = report_data.get('å¯¹æ¯”åˆ†æ', {}).get('å¯¹æ¯”', {})

        result_text = "=== ç”µæ± æµ‹è¯•åˆ†ææŠ¥å‘Š ===\n\n"
        result_text += f"æµ‹è¯•æ—¶é•¿: {report_data['æµ‹è¯•æ—¶é•¿']:.1f} ç§’\n\n"

        result_text += "ã€æ¸©å‡å¯¹æ¯”åˆ†æã€‘\n"
        if temp_compare:
            result_text += f"ä¸‰å…ƒç”µæ± æ¸©å‡: {temp_compare['ä¸‰å…ƒæ¸©å‡']:.2f}Â°C\n"
            result_text += f"åˆ€ç‰‡ç”µæ± æ¸©å‡: {temp_compare['åˆ€ç‰‡æ¸©å‡']:.2f}Â°C\n"
            result_text += f"æ¸©å‡å·®å¼‚: {temp_compare['æ¸©å‡å·®å¼‚']:.2f}Â°C\n"
            result_text += f"ä¼˜åŠ¿ç”µæ± : {temp_compare['ä¼˜åŠ¿ç”µæ± ']}\n\n"

        result_text += f"ã€mAhå®¹é‡ã€‘\n"
        result_text += f"ç´¯è®¡å®¹é‡: {report_data['mAhå®¹é‡']:.2f} mAh\n"

        # è¯¢é—®æ˜¯å¦å¯¼å‡ºåˆ°æ–‡ä»¶
        from PySide6.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            self,
            "åˆ†ææŠ¥å‘Š",
            result_text + "\n\næ˜¯å¦å¯¼å‡ºæŠ¥å‘Šåˆ°æ–‡ä»¶ï¼Ÿ",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes
        )

        if reply == QMessageBox.StandardButton.Yes:
            self._export_report_to_file(report_data, result_text)

    def _save_waveform(self) -> None:
        """ä¿å­˜æ³¢å½¢æ•°æ®åˆ°æ–‡ä»¶"""
        if not self.x_data:
            QMessageBox.warning(self, "æ— æ•°æ®", "å½“å‰æ²¡æœ‰æ³¢å½¢æ•°æ®å¯ä»¥ä¿å­˜")
            return

        from PySide6.QtWidgets import QFileDialog
        import json
        from datetime import datetime

        # é€‰æ‹©ä¿å­˜è·¯å¾„
        default_name = f"waveform_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "ä¿å­˜æ³¢å½¢æ•°æ®",
            default_name,
            "JSONæ–‡ä»¶ (*.json);;æ‰€æœ‰æ–‡ä»¶ (*.*)"
        )

        if not file_path:
            return

        try:
            # å‡†å¤‡ä¿å­˜çš„æ•°æ®
            save_data = {
                'timestamp': datetime.now().isoformat(),
                'product_model': self.control.edit_model.text(),
                'product_sn': self.control.edit_sn.text(),
                'tester': self.control.edit_tester.text(),
                'x_data': self.x_data,
                'ternary_voltage': self.ternary_volt_data,
                'ternary_temp': self.ternary_temp_data,
                'blade_voltage': self.blade_volt_data,
                'blade_temp': self.blade_temp_data,
                'channel_config': self.channel_config,
                'analysis_data': self.analysis_engine.generate_report_data() if self.analysis_engine.ternary_data.timestamps else None,
            }

            # å†™å…¥æ–‡ä»¶
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2)

            self.statusBar().showMessage(f"âœ“ æ³¢å½¢æ•°æ®å·²ä¿å­˜åˆ°: {file_path}")
            QMessageBox.information(self, "ä¿å­˜æˆåŠŸ", f"æ³¢å½¢æ•°æ®å·²ä¿å­˜åˆ°:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "ä¿å­˜å¤±è´¥", f"ä¿å­˜æ³¢å½¢æ•°æ®æ—¶å‡ºé”™:\n{str(e)}")

    def _recall_waveform(self) -> None:
        """ä»æ–‡ä»¶å¬å›æ³¢å½¢æ•°æ®"""
        from PySide6.QtWidgets import QFileDialog
        import json

        # é€‰æ‹©æ–‡ä»¶
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "å¬å›æ³¢å½¢æ•°æ®",
            "",
            "JSONæ–‡ä»¶ (*.json);;æ‰€æœ‰æ–‡ä»¶ (*.*)"
        )

        if not file_path:
            return

        try:
            # è¯»å–æ–‡ä»¶
            with open(file_path, 'r', encoding='utf-8') as f:
                load_data = json.load(f)

            # åœæ­¢å½“å‰é‡‡é›†
            if self.is_running:
                self._on_stop()

            # æ¢å¤æ•°æ®
            self.x_data = load_data.get('x_data', [])
            self.ternary_volt_data = load_data.get('ternary_voltage', [])
            self.ternary_temp_data = load_data.get('ternary_temp', [])
            self.blade_volt_data = load_data.get('blade_voltage', [])
            self.blade_temp_data = load_data.get('blade_temp', [])

            # æ¢å¤äº§å“ä¿¡æ¯
            if 'product_model' in load_data:
                self.control.edit_model.setText(load_data['product_model'])
            if 'product_sn' in load_data:
                self.control.edit_sn.setText(load_data['product_sn'])
            if 'tester' in load_data:
                self.control.edit_tester.setText(load_data['tester'])

            # æ¢å¤é€šé“é…ç½®
            if 'channel_config' in load_data:
                self.channel_config = load_data['channel_config']

            # æ›´æ–°æ³¢å½¢æ˜¾ç¤º
            if len(self.volt_curves) >= 2 and len(self.temp_curves) >= 2:
                self.volt_curves[0].setData(self.x_data, self.ternary_volt_data)
                self.temp_curves[0].setData(self.x_data, self.ternary_temp_data)
                self.volt_curves[1].setData(self.x_data, self.blade_volt_data)
                self.temp_curves[1].setData(self.x_data, self.blade_temp_data)

            # æ›´æ–°KPIæ˜¾ç¤º
            if self.ternary_volt_data:
                self.ternary_voltage_kpi.set_value(f"{self.ternary_volt_data[-1]:.2f}")
                self.ternary_temp_kpi.set_value(f"{self.ternary_temp_data[-1]:.2f}")
                self.blade_voltage_kpi.set_value(f"{self.blade_volt_data[-1]:.2f}")
                self.blade_temp_kpi.set_value(f"{self.blade_temp_data[-1]:.2f}")

            self.statusBar().showMessage(f"âœ“ æ³¢å½¢æ•°æ®å·²å¬å›: {file_path}")
            QMessageBox.information(
                self,
                "å¬å›æˆåŠŸ",
                f"æ³¢å½¢æ•°æ®å·²æˆåŠŸå¬å›ï¼\n\n"
                f"æ–‡ä»¶: {file_path}\n"
                f"æ—¶é—´: {load_data.get('timestamp', 'æœªçŸ¥')}\n"
                f"æ•°æ®ç‚¹æ•°: {len(self.x_data)}"
            )

        except Exception as e:
            QMessageBox.critical(self, "å¬å›å¤±è´¥", f"å¬å›æ³¢å½¢æ•°æ®æ—¶å‡ºé”™:\n{str(e)}")

    def _export_report_to_file(self, report_data: dict, result_text: str) -> None:
        """å¯¼å‡ºæŠ¥å‘Šåˆ°æ–‡ä»¶ï¼ˆæ”¯æŒPDFã€Excelã€HTMLã€TXTæ ¼å¼ï¼‰"""
        from PySide6.QtWidgets import QFileDialog, QMessageBox

        # é€‰æ‹©å¯¼å‡ºæ ¼å¼å’Œè·¯å¾„
        default_name = f"battery_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "å¯¼å‡ºæµ‹è¯•æŠ¥å‘Š",
            default_name,
            "PDFæ–‡ä»¶ (*.pdf);;Excelæ–‡ä»¶ (*.xlsx);;HTMLæ–‡ä»¶ (*.html);;æ–‡æœ¬æ–‡ä»¶ (*.txt);;æ‰€æœ‰æ–‡ä»¶ (*.*)"
        )

        if not file_path:
            return

        try:
            if file_path.endswith('.pdf'):
                self._export_pdf_report(file_path, report_data)
            elif file_path.endswith('.xlsx'):
                self._export_excel_report(file_path, report_data)
            elif file_path.endswith('.html'):
                self._export_html_report(file_path, report_data)
            elif file_path.endswith('.txt'):
                self._export_txt_report(file_path, report_data)
            else:
                # é»˜è®¤å¯¼å‡ºä¸ºPDF
                self._export_pdf_report(file_path + '.pdf', report_data)

            self.statusBar().showMessage(f"âœ“ æŠ¥å‘Šå·²å¯¼å‡º: {file_path}")
            QMessageBox.information(self, "å¯¼å‡ºæˆåŠŸ", f"æµ‹è¯•æŠ¥å‘Šå·²å¯¼å‡ºåˆ°:\n{file_path}")

        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "å¯¼å‡ºå¤±è´¥", f"å¯¼å‡ºæŠ¥å‘Šæ—¶å‡ºé”™:\n{str(e)}")

    def _export_pdf_report(self, file_path: str, report_data: dict) -> None:
        """å¯¼å‡ºPDFæ ¼å¼æŠ¥å‘Šï¼ˆä¸“ä¸šç‰ˆï¼‰"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import tempfile
            import os
        except ImportError:
            QMessageBox.warning(
                self, "ç¼ºå°‘ä¾èµ–",
                "å¯¼å‡ºPDFéœ€è¦å®‰è£… reportlab åº“ã€‚\n\n"
                "è¯·è¿è¡Œ: pip install reportlab"
            )
            return

        # æ³¨å†Œä¸­æ–‡å­—ä½“
        font_registered = False
        try:
            # å°è¯•ä½¿ç”¨ç³»ç»Ÿä¸­æ–‡å­—ä½“
            font_paths = [
                "C:/Windows/Fonts/simhei.ttf",  # é»‘ä½“
                "C:/Windows/Fonts/msyh.ttf",    # å¾®è½¯é›…é»‘
                "C:/Windows/Fonts/simsun.ttc",  # å®‹ä½“
            ]
            for font_path in font_paths:
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                    font_registered = True
                    print(f"âœ“ å·²æ³¨å†Œä¸­æ–‡å­—ä½“: {font_path}")
                    break
            if not font_registered:
                print("âš ï¸ æœªæ‰¾åˆ°ä¸­æ–‡å­—ä½“ï¼Œä½¿ç”¨é»˜è®¤å­—ä½“")
        except Exception as e:
            print(f"âš ï¸ æ³¨å†Œä¸­æ–‡å­—ä½“å¤±è´¥: {e}")
            font_registered = False

        # åˆ›å»ºPDFæ–‡æ¡£
        doc = SimpleDocTemplate(
            file_path,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )

        # æ ·å¼
        styles = getSampleStyleSheet()
        try:
            title_style = ParagraphStyle(
                'ChineseTitle',
                parent=styles['Title'],
                fontName='ChineseFont',
                fontSize=18,
                alignment=1,  # å±…ä¸­
                spaceAfter=20
            )
            heading_style = ParagraphStyle(
                'ChineseHeading',
                parent=styles['Heading2'],
                fontName='ChineseFont',
                fontSize=14,
                textColor=colors.HexColor('#1e3a8a'),
                spaceBefore=15,
                spaceAfter=10
            )
            normal_style = ParagraphStyle(
                'ChineseNormal',
                parent=styles['Normal'],
                fontName='ChineseFont',
                fontSize=10,
                leading=14
            )
        except Exception:
            # å¦‚æœä¸­æ–‡å­—ä½“æ³¨å†Œå¤±è´¥ï¼Œä½¿ç”¨é»˜è®¤æ ·å¼
            title_style = styles['Title']
            heading_style = styles['Heading2']
            normal_style = styles['Normal']

        # æ„å»ºPDFå†…å®¹
        story = []

        # æ ‡é¢˜
        story.append(Paragraph("ç”µæ± ç”µå‹ä¸æ¸©å‡æµ‹è¯•åˆ†ææŠ¥å‘Š", title_style))
        story.append(Spacer(1, 10*mm))

        # åŸºæœ¬ä¿¡æ¯è¡¨æ ¼
        info_data = [
            ["æŠ¥å‘Šç”Ÿæˆæ—¶é—´", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["äº§å“å‹å·", self.control.edit_model.text() or "-"],
            ["äº§å“æµæ°´å·", self.control.edit_sn.text() or "-"],
            ["æµ‹è¯•å‘˜", self.control.edit_tester.text() or "-"],
            ["æµ‹è¯•æ—¶é•¿", f"{report_data.get('æµ‹è¯•æ—¶é•¿', 0):.1f} ç§’"],
            ["æ•°æ®ç‚¹æ•°", f"{report_data.get('æ•°æ®ç‚¹æ•°', 0)} ä¸ª"],
        ]
        info_table = Table(info_data, colWidths=[50*mm, 100*mm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont' if font_registered else 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 10*mm))

        # ä¸‰å…ƒç”µæ± æµ‹è¯•ç»“æœ
        story.append(Paragraph("ä¸€ã€ä¸‰å…ƒç”µæ± æµ‹è¯•ç»“æœ", heading_style))
        ternary = report_data.get('ä¸‰å…ƒç”µæ± ', {})
        temp_rise = ternary.get('æ¸©å‡åˆ†æ', {})
        voltage_drop = ternary.get('å‹é™åˆ†æ', {})

        ternary_data = [
            ["æµ‹è¯•é¡¹ç›®", "æ¸©å‡åˆ†æ", "ç”µå‹åˆ†æ"],
            ["åˆå§‹å€¼", f"{temp_rise.get('åˆå§‹æ¸©åº¦', 0):.2f} Â°C", f"{voltage_drop.get('åˆå§‹ç”µå‹', 0):.2f} V"],
            ["å½“å‰å€¼", f"{temp_rise.get('å½“å‰æ¸©åº¦', 0):.2f} Â°C", f"{voltage_drop.get('å½“å‰ç”µå‹', 0):.2f} V"],
            ["æœ€é«˜å€¼", f"{temp_rise.get('å³°å€¼æ¸©åº¦', 0):.2f} Â°C", f"{voltage_drop.get('æœ€é«˜ç”µå‹', 0):.2f} V"],
            ["æœ€ä½å€¼", f"{temp_rise.get('æœ€ä½æ¸©åº¦', 0):.2f} Â°C", f"{voltage_drop.get('æœ€ä½ç”µå‹', 0):.2f} V"],
            ["å˜åŒ–é‡", f"{temp_rise.get('æ¸©å‡', 0):.2f} Â°C", f"{voltage_drop.get('ç”µå‹é™', 0):.2f} V"],
            ["å¹³å‡å€¼", f"{temp_rise.get('å¹³å‡æ¸©åº¦', 0):.2f} Â°C", f"{voltage_drop.get('å¹³å‡ç”µå‹', 0):.2f} V"],
        ]
        ternary_table = Table(ternary_data, colWidths=[40*mm, 55*mm, 55*mm])
        ternary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#dbeafe')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont' if font_registered else 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(ternary_table)
        story.append(Spacer(1, 8*mm))

        # åˆ€ç‰‡ç”µæ± æµ‹è¯•ç»“æœ
        story.append(Paragraph("äºŒã€åˆ€ç‰‡ç”µæ± æµ‹è¯•ç»“æœ", heading_style))
        blade = report_data.get('åˆ€ç‰‡ç”µæ± ', {})
        temp_rise_b = blade.get('æ¸©å‡åˆ†æ', {})
        voltage_drop_b = blade.get('å‹é™åˆ†æ', {})

        blade_data = [
            ["æµ‹è¯•é¡¹ç›®", "æ¸©å‡åˆ†æ", "ç”µå‹åˆ†æ"],
            ["åˆå§‹å€¼", f"{temp_rise_b.get('åˆå§‹æ¸©åº¦', 0):.2f} Â°C", f"{voltage_drop_b.get('åˆå§‹ç”µå‹', 0):.2f} V"],
            ["å½“å‰å€¼", f"{temp_rise_b.get('å½“å‰æ¸©åº¦', 0):.2f} Â°C", f"{voltage_drop_b.get('å½“å‰ç”µå‹', 0):.2f} V"],
            ["æœ€é«˜å€¼", f"{temp_rise_b.get('å³°å€¼æ¸©åº¦', 0):.2f} Â°C", f"{voltage_drop_b.get('æœ€é«˜ç”µå‹', 0):.2f} V"],
            ["æœ€ä½å€¼", f"{temp_rise_b.get('æœ€ä½æ¸©åº¦', 0):.2f} Â°C", f"{voltage_drop_b.get('æœ€ä½ç”µå‹', 0):.2f} V"],
            ["å˜åŒ–é‡", f"{temp_rise_b.get('æ¸©å‡', 0):.2f} Â°C", f"{voltage_drop_b.get('ç”µå‹é™', 0):.2f} V"],
            ["å¹³å‡å€¼", f"{temp_rise_b.get('å¹³å‡æ¸©åº¦', 0):.2f} Â°C", f"{voltage_drop_b.get('å¹³å‡ç”µå‹', 0):.2f} V"],
        ]
        blade_table = Table(blade_data, colWidths=[40*mm, 55*mm, 55*mm])
        blade_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#d1fae5')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont' if font_registered else 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(blade_table)
        story.append(Spacer(1, 8*mm))

        # å¯¹æ¯”åˆ†æ
        story.append(Paragraph("ä¸‰ã€å¯¹æ¯”åˆ†æ", heading_style))
        compare = report_data.get('å¯¹æ¯”åˆ†æ', {}).get('å¯¹æ¯”', {})

        compare_data = [
            ["å¯¹æ¯”é¡¹ç›®", "ä¸‰å…ƒç”µæ± ", "åˆ€ç‰‡ç”µæ± ", "å·®å¼‚"],
            ["æ¸©å‡", f"{compare.get('ä¸‰å…ƒæ¸©å‡', 0):.2f} Â°C", f"{compare.get('åˆ€ç‰‡æ¸©å‡', 0):.2f} Â°C", f"{compare.get('æ¸©å‡å·®å¼‚', 0):.2f} Â°C"],
        ]
        compare_table = Table(compare_data, colWidths=[35*mm, 40*mm, 40*mm, 35*mm])
        compare_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont' if font_registered else 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(compare_table)
        story.append(Spacer(1, 5*mm))

        # ç»“è®º
        winner = compare.get('ä¼˜åŠ¿ç”µæ± ', 'æœªçŸ¥')
        conclusion_text = f"<b>æµ‹è¯•ç»“è®ºï¼š</b>æ ¹æ®æ¸©å‡å¯¹æ¯”åˆ†æï¼Œ<font color='#dc2626'><b>{winner}</b></font>åœ¨æœ¬æ¬¡æµ‹è¯•ä¸­è¡¨ç°æ›´ä¼˜ã€‚"
        story.append(Paragraph(conclusion_text, normal_style))
        story.append(Spacer(1, 8*mm))

        # å®¹é‡æµ‹è¯•
        story.append(Paragraph("å››ã€å®¹é‡æµ‹è¯•", heading_style))
        capacity_data = [
            ["æµ‹è¯•é¡¹ç›®", "æ•°å€¼"],
            ["ç´¯è®¡å®¹é‡", f"{report_data.get('mAhå®¹é‡', 0):.2f} mAh"],
        ]
        capacity_table = Table(capacity_data, colWidths=[50*mm, 100*mm])
        capacity_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont' if font_registered else 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(capacity_table)
        story.append(Spacer(1, 15*mm))

        # é¡µè„š
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontName='ChineseFont' if font_registered else 'Helvetica',
            fontSize=8,
            textColor=colors.grey,
            alignment=1
        )
        story.append(Paragraph("â€”â€” ç”µæ± ç”µå‹ä¸æ¸©å‡é‡‡é›†è½¯ä»¶ v1.0 â€”â€”", footer_style))
        story.append(Paragraph(f"æŠ¥å‘Šç”Ÿæˆæ—¶é—´: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))

        # ç”ŸæˆPDF
        doc.build(story)

    def _export_txt_report(self, file_path: str, report_data: dict) -> None:
        """å¯¼å‡ºTXTæ ¼å¼æŠ¥å‘Š"""
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("ç”µæ± ç”µå‹ä¸æ¸©å‡æµ‹è¯•åˆ†ææŠ¥å‘Š\n")
            f.write("=" * 60 + "\n\n")

            f.write(f"ç”Ÿæˆæ—¶é—´: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"äº§å“å‹å·: {self.control.edit_model.text()}\n")
            f.write(f"äº§å“æµæ°´å·: {self.control.edit_sn.text()}\n")
            f.write(f"æµ‹è¯•å‘˜: {self.control.edit_tester.text()}\n")
            f.write(f"æµ‹è¯•æ—¶é•¿: {report_data['æµ‹è¯•æ—¶é•¿']:.1f} ç§’\n\n")

            f.write("-" * 60 + "\n")
            f.write("ä¸€ã€ä¸‰å…ƒç”µæ± æµ‹è¯•ç»“æœ\n")
            f.write("-" * 60 + "\n")
            ternary = report_data.get('ä¸‰å…ƒç”µæ± ', {})
            temp_rise = ternary.get('æ¸©å‡åˆ†æ', {})
            voltage_drop = ternary.get('å‹é™åˆ†æ', {})

            if temp_rise:
                f.write("\nã€æ¸©å‡åˆ†æã€‘\n")
                f.write(f"  åˆå§‹æ¸©åº¦: {temp_rise.get('åˆå§‹æ¸©åº¦', 0):.2f} Â°C\n")
                f.write(f"  å½“å‰æ¸©åº¦: {temp_rise.get('å½“å‰æ¸©åº¦', 0):.2f} Â°C\n")
                f.write(f"  å³°å€¼æ¸©åº¦: {temp_rise.get('å³°å€¼æ¸©åº¦', 0):.2f} Â°C\n")
                f.write(f"  æœ€ä½æ¸©åº¦: {temp_rise.get('æœ€ä½æ¸©åº¦', 0):.2f} Â°C\n")
                f.write(f"  æ¸©å‡: {temp_rise.get('æ¸©å‡', 0):.2f} Â°C\n")
                f.write(f"  å¹³å‡æ¸©åº¦: {temp_rise.get('å¹³å‡æ¸©åº¦', 0):.2f} Â°C\n")

            if voltage_drop:
                f.write("\nã€ç”µå‹åˆ†æã€‘\n")
                f.write(f"  åˆå§‹ç”µå‹: {voltage_drop.get('åˆå§‹ç”µå‹', 0):.2f} V\n")
                f.write(f"  å½“å‰ç”µå‹: {voltage_drop.get('å½“å‰ç”µå‹', 0):.2f} V\n")
                f.write(f"  æœ€é«˜ç”µå‹: {voltage_drop.get('æœ€é«˜ç”µå‹', 0):.2f} V\n")
                f.write(f"  æœ€ä½ç”µå‹: {voltage_drop.get('æœ€ä½ç”µå‹', 0):.2f} V\n")
                f.write(f"  ç”µå‹é™: {voltage_drop.get('ç”µå‹é™', 0):.2f} V\n")
                f.write(f"  å‹é™ç‡: {voltage_drop.get('å‹é™ç‡', 0):.2f} %\n")
                f.write(f"  å¹³å‡ç”µå‹: {voltage_drop.get('å¹³å‡ç”µå‹', 0):.2f} V\n")

            f.write("\n" + "-" * 60 + "\n")
            f.write("äºŒã€åˆ€ç‰‡ç”µæ± æµ‹è¯•ç»“æœ\n")
            f.write("-" * 60 + "\n")
            blade = report_data.get('åˆ€ç‰‡ç”µæ± ', {})
            temp_rise = blade.get('æ¸©å‡åˆ†æ', {})
            voltage_drop = blade.get('å‹é™åˆ†æ', {})

            if temp_rise:
                f.write("\nã€æ¸©å‡åˆ†æã€‘\n")
                f.write(f"  åˆå§‹æ¸©åº¦: {temp_rise.get('åˆå§‹æ¸©åº¦', 0):.2f} Â°C\n")
                f.write(f"  å½“å‰æ¸©åº¦: {temp_rise.get('å½“å‰æ¸©åº¦', 0):.2f} Â°C\n")
                f.write(f"  å³°å€¼æ¸©åº¦: {temp_rise.get('å³°å€¼æ¸©åº¦', 0):.2f} Â°C\n")
                f.write(f"  æœ€ä½æ¸©åº¦: {temp_rise.get('æœ€ä½æ¸©åº¦', 0):.2f} Â°C\n")
                f.write(f"  æ¸©å‡: {temp_rise.get('æ¸©å‡', 0):.2f} Â°C\n")
                f.write(f"  å¹³å‡æ¸©åº¦: {temp_rise.get('å¹³å‡æ¸©åº¦', 0):.2f} Â°C\n")

            if voltage_drop:
                f.write("\nã€ç”µå‹åˆ†æã€‘\n")
                f.write(f"  åˆå§‹ç”µå‹: {voltage_drop.get('åˆå§‹ç”µå‹', 0):.2f} V\n")
                f.write(f"  å½“å‰ç”µå‹: {voltage_drop.get('å½“å‰ç”µå‹', 0):.2f} V\n")
                f.write(f"  æœ€é«˜ç”µå‹: {voltage_drop.get('æœ€é«˜ç”µå‹', 0):.2f} V\n")
                f.write(f"  æœ€ä½ç”µå‹: {voltage_drop.get('æœ€ä½ç”µå‹', 0):.2f} V\n")
                f.write(f"  ç”µå‹é™: {voltage_drop.get('ç”µå‹é™', 0):.2f} V\n")
                f.write(f"  å‹é™ç‡: {voltage_drop.get('å‹é™ç‡', 0):.2f} %\n")
                f.write(f"  å¹³å‡ç”µå‹: {voltage_drop.get('å¹³å‡ç”µå‹', 0):.2f} V\n")

            f.write("\n" + "-" * 60 + "\n")
            f.write("ä¸‰ã€å¯¹æ¯”åˆ†æ\n")
            f.write("-" * 60 + "\n")
            compare = report_data.get('å¯¹æ¯”åˆ†æ', {}).get('å¯¹æ¯”', {})
            if compare:
                f.write(f"\nä¸‰å…ƒç”µæ± æ¸©å‡: {compare.get('ä¸‰å…ƒæ¸©å‡', 0):.2f} Â°C\n")
                f.write(f"åˆ€ç‰‡ç”µæ± æ¸©å‡: {compare.get('åˆ€ç‰‡æ¸©å‡', 0):.2f} Â°C\n")
                f.write(f"æ¸©å‡å·®å¼‚: {compare.get('æ¸©å‡å·®å¼‚', 0):.2f} Â°C\n")
                f.write(f"ä¼˜åŠ¿ç”µæ± : {compare.get('ä¼˜åŠ¿ç”µæ± ', 'æœªçŸ¥')}\n")

            f.write("\n" + "-" * 60 + "\n")
            f.write("å››ã€å®¹é‡æµ‹è¯•\n")
            f.write("-" * 60 + "\n")
            f.write(f"\nç´¯è®¡å®¹é‡: {report_data.get('mAhå®¹é‡', 0):.2f} mAh\n")

            f.write("\n" + "=" * 60 + "\n")
            f.write("æŠ¥å‘Šç»“æŸ\n")
            f.write("=" * 60 + "\n")

    def _export_excel_report(self, file_path: str, report_data: dict) -> None:
        """å¯¼å‡ºExcelæ ¼å¼æŠ¥å‘Š"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from datetime import datetime
        except ImportError:
            QMessageBox.warning(
                self,
                "ç¼ºå°‘ä¾èµ–",
                "å¯¼å‡ºExceléœ€è¦å®‰è£… openpyxl åº“\n\nè¯·è¿è¡Œ: pip install openpyxl"
            )
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ç”µæ± æµ‹è¯•æŠ¥å‘Š"

        # è®¾ç½®åˆ—å®½
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        # æ ‡é¢˜æ ·å¼
        title_font = Font(name='å¾®è½¯é›…é»‘', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_alignment = Alignment(horizontal='center', vertical='center')

        # è¡¨å¤´æ ·å¼
        header_font = Font(name='å¾®è½¯é›…é»‘', size=12, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # è¾¹æ¡†
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1

        # æ ‡é¢˜
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "ç”µæ± ç”µå‹ä¸æ¸©å‡æµ‹è¯•åˆ†ææŠ¥å‘Š"
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_alignment
        row += 2

        # åŸºæœ¬ä¿¡æ¯
        ws[f'A{row}'] = "ç”Ÿæˆæ—¶é—´:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        ws[f'A{row}'] = "äº§å“å‹å·:"
        ws[f'B{row}'] = self.control.edit_model.text()
        row += 1
        ws[f'A{row}'] = "äº§å“æµæ°´å·:"
        ws[f'B{row}'] = self.control.edit_sn.text()
        row += 1
        ws[f'A{row}'] = "æµ‹è¯•å‘˜:"
        ws[f'B{row}'] = self.control.edit_tester.text()
        row += 1
        ws[f'A{row}'] = "æµ‹è¯•æ—¶é•¿:"
        ws[f'B{row}'] = f"{report_data['æµ‹è¯•æ—¶é•¿']:.1f} ç§’"
        row += 2

        # ä¸‰å…ƒç”µæ± æ•°æ®
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "ä¸‰å…ƒç”µæ± æµ‹è¯•ç»“æœ"
        cell.font = header_font
        cell.fill = header_fill
        row += 1

        ternary = report_data.get('ä¸‰å…ƒç”µæ± ', {})
        temp_rise = ternary.get('æ¸©å‡åˆ†æ', {})
        voltage_drop = ternary.get('å‹é™åˆ†æ', {})

        ws[f'A{row}'] = "é¡¹ç›®"
        ws[f'B{row}'] = "æ¸©å‡åˆ†æ"
        ws[f'C{row}'] = "é¡¹ç›®"
        ws[f'D{row}'] = "ç”µå‹åˆ†æ"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = header_font
        row += 1

        metrics = [
            ('åˆå§‹æ¸©åº¦', 'åˆå§‹ç”µå‹'),
            ('å½“å‰æ¸©åº¦', 'å½“å‰ç”µå‹'),
            ('å³°å€¼æ¸©åº¦', 'æœ€é«˜ç”µå‹'),
            ('æœ€ä½æ¸©åº¦', 'æœ€ä½ç”µå‹'),
            ('æ¸©å‡', 'ç”µå‹é™'),
            ('å¹³å‡æ¸©åº¦', 'å¹³å‡ç”µå‹'),
        ]

        for temp_key, volt_key in metrics:
            ws[f'A{row}'] = temp_key
            ws[f'B{row}'] = f"{temp_rise.get(temp_key, 0):.2f}"
            ws[f'C{row}'] = volt_key
            ws[f'D{row}'] = f"{voltage_drop.get(volt_key, 0):.2f}"
            row += 1

        row += 1

        # åˆ€ç‰‡ç”µæ± æ•°æ®
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "åˆ€ç‰‡ç”µæ± æµ‹è¯•ç»“æœ"
        cell.font = header_font
        cell.fill = header_fill
        row += 1

        blade = report_data.get('åˆ€ç‰‡ç”µæ± ', {})
        temp_rise = blade.get('æ¸©å‡åˆ†æ', {})
        voltage_drop = blade.get('å‹é™åˆ†æ', {})

        ws[f'A{row}'] = "é¡¹ç›®"
        ws[f'B{row}'] = "æ¸©å‡åˆ†æ"
        ws[f'C{row}'] = "é¡¹ç›®"
        ws[f'D{row}'] = "ç”µå‹åˆ†æ"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = header_font
        row += 1

        for temp_key, volt_key in metrics:
            ws[f'A{row}'] = temp_key
            ws[f'B{row}'] = f"{temp_rise.get(temp_key, 0):.2f}"
            ws[f'C{row}'] = volt_key
            ws[f'D{row}'] = f"{voltage_drop.get(volt_key, 0):.2f}"
            row += 1

        row += 1

        # å¯¹æ¯”åˆ†æ
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "å¯¹æ¯”åˆ†æ"
        cell.font = header_font
        cell.fill = header_fill
        row += 1

        compare = report_data.get('å¯¹æ¯”åˆ†æ', {}).get('å¯¹æ¯”', {})
        if compare:
            ws[f'A{row}'] = "ä¸‰å…ƒç”µæ± æ¸©å‡"
            ws[f'B{row}'] = f"{compare.get('ä¸‰å…ƒæ¸©å‡', 0):.2f} Â°C"
            row += 1
            ws[f'A{row}'] = "åˆ€ç‰‡ç”µæ± æ¸©å‡"
            ws[f'B{row}'] = f"{compare.get('åˆ€ç‰‡æ¸©å‡', 0):.2f} Â°C"
            row += 1
            ws[f'A{row}'] = "æ¸©å‡å·®å¼‚"
            ws[f'B{row}'] = f"{compare.get('æ¸©å‡å·®å¼‚', 0):.2f} Â°C"
            row += 1
            ws[f'A{row}'] = "ä¼˜åŠ¿ç”µæ± "
            ws[f'B{row}'] = compare.get('ä¼˜åŠ¿ç”µæ± ', 'æœªçŸ¥')
            row += 1

        row += 1
        ws[f'A{row}'] = "ç´¯è®¡å®¹é‡"
        ws[f'B{row}'] = f"{report_data.get('mAhå®¹é‡', 0):.2f} mAh"

        # åº”ç”¨è¾¹æ¡†
        for row_cells in ws.iter_rows(min_row=1, max_row=row, min_col=1, max_col=4):
            for cell in row_cells:
                cell.border = thin_border

        wb.save(file_path)

    def _export_html_report(self, file_path: str, report_data: dict) -> None:
        """å¯¼å‡ºHTMLæ ¼å¼æŠ¥å‘Š"""
        html_content = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ç”µæ± æµ‹è¯•åˆ†ææŠ¥å‘Š</title>
    <style>
        body {{
            font-family: "Microsoft YaHei", Arial, sans-serif;
            max-width: 1000px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            background-color: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2c3e50;
            text-align: center;
            border-bottom: 3px solid #3498db;
            padding-bottom: 15px;
        }}
        h2 {{
            color: #34495e;
            margin-top: 30px;
            border-left: 4px solid #3498db;
            padding-left: 10px;
        }}
        .info-table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        .info-table th, .info-table td {{
            padding: 12px;
            text-align: left;
            border: 1px solid #ddd;
        }}
        .info-table th {{
            background-color: #3498db;
            color: white;
            font-weight: bold;
        }}
        .info-table tr:nth-child(even) {{
            background-color: #f2f2f2;
        }}
        .highlight {{
            background-color: #fff3cd;
            padding: 15px;
            border-left: 4px solid #ffc107;
            margin: 15px 0;
        }}
        .footer {{
            text-align: center;
            margin-top: 30px;
            color: #7f8c8d;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>ç”µæ± ç”µå‹ä¸æ¸©å‡æµ‹è¯•åˆ†ææŠ¥å‘Š</h1>

        <table class="info-table">
            <tr>
                <th>é¡¹ç›®</th>
                <th>å†…å®¹</th>
            </tr>
            <tr>
                <td>ç”Ÿæˆæ—¶é—´</td>
                <td>{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</td>
            </tr>
            <tr>
                <td>äº§å“å‹å·</td>
                <td>{self.control.edit_model.text()}</td>
            </tr>
            <tr>
                <td>äº§å“æµæ°´å·</td>
                <td>{self.control.edit_sn.text()}</td>
            </tr>
            <tr>
                <td>æµ‹è¯•å‘˜</td>
                <td>{self.control.edit_tester.text()}</td>
            </tr>
            <tr>
                <td>æµ‹è¯•æ—¶é•¿</td>
                <td>{report_data['æµ‹è¯•æ—¶é•¿']:.1f} ç§’</td>
            </tr>
        </table>

        <h2>ä¸€ã€ä¸‰å…ƒç”µæ± æµ‹è¯•ç»“æœ</h2>
        <table class="info-table">
            <tr>
                <th colspan="2">æ¸©å‡åˆ†æ</th>
                <th colspan="2">ç”µå‹åˆ†æ</th>
            </tr>
"""

        ternary = report_data.get('ä¸‰å…ƒç”µæ± ', {})
        temp_rise = ternary.get('æ¸©å‡åˆ†æ', {})
        voltage_drop = ternary.get('å‹é™åˆ†æ', {})

        metrics = [
            ('åˆå§‹æ¸©åº¦', 'åˆå§‹ç”µå‹'),
            ('å½“å‰æ¸©åº¦', 'å½“å‰ç”µå‹'),
            ('å³°å€¼æ¸©åº¦', 'æœ€é«˜ç”µå‹'),
            ('æœ€ä½æ¸©åº¦', 'æœ€ä½ç”µå‹'),
            ('æ¸©å‡', 'ç”µå‹é™'),
            ('å¹³å‡æ¸©åº¦', 'å¹³å‡ç”µå‹'),
        ]

        for temp_key, volt_key in metrics:
            html_content += f"""
            <tr>
                <td>{temp_key}</td>
                <td>{temp_rise.get(temp_key, 0):.2f} Â°C</td>
                <td>{volt_key}</td>
                <td>{voltage_drop.get(volt_key, 0):.2f} V</td>
            </tr>
"""

        html_content += """
        </table>

        <h2>äºŒã€åˆ€ç‰‡ç”µæ± æµ‹è¯•ç»“æœ</h2>
        <table class="info-table">
            <tr>
                <th colspan="2">æ¸©å‡åˆ†æ</th>
                <th colspan="2">ç”µå‹åˆ†æ</th>
            </tr>
"""

        blade = report_data.get('åˆ€ç‰‡ç”µæ± ', {})
        temp_rise = blade.get('æ¸©å‡åˆ†æ', {})
        voltage_drop = blade.get('å‹é™åˆ†æ', {})

        for temp_key, volt_key in metrics:
            html_content += f"""
            <tr>
                <td>{temp_key}</td>
                <td>{temp_rise.get(temp_key, 0):.2f} Â°C</td>
                <td>{volt_key}</td>
                <td>{voltage_drop.get(volt_key, 0):.2f} V</td>
            </tr>
"""

        compare = report_data.get('å¯¹æ¯”åˆ†æ', {}).get('å¯¹æ¯”', {})

        html_content += f"""
        </table>

        <h2>ä¸‰ã€å¯¹æ¯”åˆ†æ</h2>
        <div class="highlight">
            <p><strong>ä¸‰å…ƒç”µæ± æ¸©å‡:</strong> {compare.get('ä¸‰å…ƒæ¸©å‡', 0):.2f} Â°C</p>
            <p><strong>åˆ€ç‰‡ç”µæ± æ¸©å‡:</strong> {compare.get('åˆ€ç‰‡æ¸©å‡', 0):.2f} Â°C</p>
            <p><strong>æ¸©å‡å·®å¼‚:</strong> {compare.get('æ¸©å‡å·®å¼‚', 0):.2f} Â°C</p>
            <p><strong>ä¼˜åŠ¿ç”µæ± :</strong> {compare.get('ä¼˜åŠ¿ç”µæ± ', 'æœªçŸ¥')}</p>
        </div>

        <h2>å››ã€å®¹é‡æµ‹è¯•</h2>
        <table class="info-table">
            <tr>
                <th>é¡¹ç›®</th>
                <th>æ•°å€¼</th>
            </tr>
            <tr>
                <td>ç´¯è®¡å®¹é‡</td>
                <td>{report_data.get('mAhå®¹é‡', 0):.2f} mAh</td>
            </tr>
        </table>

        <div class="footer">
            <p>ç”µæ± ç”µå‹ä¸æ¸©å‡é‡‡é›†è½¯ä»¶ v0.1.0</p>
            <p>æŠ¥å‘Šç”Ÿæˆæ—¶é—´: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>
"""

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)


class MXPlusBDialog(QDialog):
    """mX+bçº¿æ€§æ ¡å‡†å¯¹è¯æ¡†ã€‚"""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("mX+b çº¿æ€§æ ¡å‡†")
        self.setFixedSize(450, 400)

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # è¯´æ˜æ–‡æœ¬
        desc = QLabel("çº¿æ€§æ ¡å‡†å…¬å¼: Y = mX + b")
        desc.setStyleSheet("color: #6dd5ed; font-size: 14px; font-weight: bold;")
        layout.addWidget(desc)

        sub_desc = QLabel("æ ¡å‡†å°†åº”ç”¨åˆ°æ³¢å½¢æ˜¾ç¤ºå’ŒKPIæ•°æ®")
        sub_desc.setStyleSheet("color: #888; font-size: 11px;")
        layout.addWidget(sub_desc)

        # é€šé“é€‰æ‹©
        channel_group = QGroupBox("é€‰æ‹©é€šé“")
        channel_layout = QVBoxLayout(channel_group)
        channel_layout.setSpacing(8)

        self.radio_ternary = QRadioButton("ä¸‰å…ƒç”µæ± ")
        self.radio_ternary.setChecked(True)
        self.radio_ternary.toggled.connect(self._on_channel_changed)
        self.radio_blade = QRadioButton("åˆ€ç‰‡ç”µæ± ")
        self.radio_blade.toggled.connect(self._on_channel_changed)

        channel_layout.addWidget(self.radio_ternary)
        channel_layout.addWidget(self.radio_blade)

        layout.addWidget(channel_group)

        # æ•°æ®ç±»å‹é€‰æ‹©
        type_group = QGroupBox("æ ¡å‡†æ•°æ®ç±»å‹")
        type_layout = QVBoxLayout(type_group)
        type_layout.setSpacing(8)

        self.radio_voltage = QRadioButton("ç”µå‹æ ¡å‡†")
        self.radio_voltage.setChecked(True)
        self.radio_voltage.toggled.connect(self._on_type_changed)
        self.radio_temp = QRadioButton("æ¸©åº¦æ ¡å‡†")
        self.radio_temp.toggled.connect(self._on_type_changed)

        type_layout.addWidget(self.radio_voltage)
        type_layout.addWidget(self.radio_temp)

        layout.addWidget(type_group)

        # å‚æ•°è¾“å…¥åŒºåŸŸ
        params_group = QGroupBox("æ ¡å‡†å‚æ•°")
        params_layout = QFormLayout(params_group)

        self.edit_m = QLineEdit("1.0")
        self.edit_m.setPlaceholderText("æ–œç‡ m")
        params_layout.addRow("æ–œç‡ m:", self.edit_m)

        self.edit_b = QLineEdit("0.0")
        self.edit_b.setPlaceholderText("æˆªè· b")
        params_layout.addRow("æˆªè· b:", self.edit_b)

        layout.addWidget(params_group)

        # æŒ‰é’®åŒºåŸŸ
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        btn_cancel = QPushButton("å–æ¶ˆ")
        btn_cancel.clicked.connect(self.reject)
        button_layout.addWidget(btn_cancel)

        btn_apply = QPushButton("åº”ç”¨æ ¡å‡†")
        btn_apply.clicked.connect(self._apply_calibration)
        btn_apply.setStyleSheet("""
            QPushButton {
                background-color: #059669;
                color: #ffffff;
                border: none;
                padding: 10px 20px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #047857;
            }
        """)
        button_layout.addWidget(btn_apply)

        layout.addLayout(button_layout)

        # åŠ è½½å½“å‰æ ¡å‡†å‚æ•°
        self._load_current_params()

    def _get_current_key(self) -> str:
        """è·å–å½“å‰é€‰æ‹©çš„æ ¡å‡†å‚æ•°é”®å"""
        battery = "ternary" if self.radio_ternary.isChecked() else "blade"
        data_type = "voltage" if self.radio_voltage.isChecked() else "temp"
        return f"{battery}_{data_type}"

    def _load_current_params(self) -> None:
        """åŠ è½½å½“å‰é€‰æ‹©é€šé“çš„æ ¡å‡†å‚æ•°"""
        main_window = self.parent()
        if not main_window or not hasattr(main_window, 'analysis_engine'):
            return

        params = main_window.analysis_engine.get_calibration_params()
        key = self._get_current_key()

        if key in params:
            self.edit_m.setText(str(params[key]['m']))
            self.edit_b.setText(str(params[key]['b']))

    def _on_channel_changed(self, checked: bool) -> None:
        """é€šé“é€‰æ‹©æ”¹å˜æ—¶æ›´æ–°å‚æ•°æ˜¾ç¤º"""
        if checked:
            self._load_current_params()

    def _on_type_changed(self, checked: bool) -> None:
        """æ•°æ®ç±»å‹é€‰æ‹©æ”¹å˜æ—¶æ›´æ–°å‚æ•°æ˜¾ç¤º"""
        if checked:
            self._load_current_params()

    def _apply_calibration(self) -> None:
        """åº”ç”¨æ ¡å‡†å‚æ•°ã€‚"""
        try:
            m = float(self.edit_m.text())
            b = float(self.edit_b.text())

            battery_type = "ternary" if self.radio_ternary.isChecked() else "blade"
            data_type = "voltage" if self.radio_voltage.isChecked() else "temp"

            battery_name = "ä¸‰å…ƒç”µæ± " if battery_type == "ternary" else "åˆ€ç‰‡ç”µæ± "
            type_name = "ç”µå‹" if data_type == "voltage" else "æ¸©åº¦"

            main_window = self.parent()
            if main_window and hasattr(main_window, 'analysis_engine'):
                # åº”ç”¨æ ¡å‡†åˆ°åˆ†æå¼•æ“
                main_window.analysis_engine.set_mx_plus_b(battery_type, data_type, m, b)

                # ä¿å­˜æ ¡å‡†å‚æ•°åˆ°é…ç½®æ–‡ä»¶
                if hasattr(main_window, '_save_channel_config_to_file'):
                    main_window._save_channel_config_to_file()

            QMessageBox.information(
                self,
                "æ ¡å‡†åº”ç”¨æˆåŠŸ",
                f"å·²å¯¹ {battery_name} {type_name} åº”ç”¨çº¿æ€§æ ¡å‡†:\nY = {m}X + {b}"
            )

            self.accept()

        except ValueError:
            QMessageBox.warning(self, "è¾“å…¥é”™è¯¯", "è¯·è¾“å…¥æœ‰æ•ˆçš„æ•°å­—å‚æ•°")


class MAHTestDialog(QDialog):
    """mAhå®¹é‡æµ‹è¯•å¯¹è¯æ¡† - åŸºäºå®é™…é‡‡é›†æ•°æ®çš„æ’æµå®¹é‡æµ‹è¯•ã€‚"""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("mAh å®¹é‡æµ‹è¯•")
        self.setFixedSize(500, 500)

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # æ ‡é¢˜
        title = QLabel("æ¯«å®‰æ—¶å®¹é‡æµ‹è¯•ï¼ˆæ’æµæ¨¡å¼ï¼‰")
        title.setStyleSheet("color: #6dd5ed; font-size: 16px; font-weight: bold;")
        layout.addWidget(title)

        # è¯´æ˜
        desc = QLabel("è¯´æ˜ï¼šè¾“å…¥æ’æµå……/æ”¾ç”µç”µæµå€¼ï¼Œç³»ç»Ÿå°†æ ¹æ®å®é™…é‡‡é›†æ—¶é—´è®¡ç®—ç´¯è®¡å®¹é‡ã€‚\n"
                      "è¯·ç¡®ä¿å·²å¼€å§‹æ•°æ®é‡‡é›†åå†å¯åŠ¨å®¹é‡æµ‹è¯•ã€‚")
        desc.setStyleSheet("color: #aaa; font-size: 11px;")
        desc.setWordWrap(True)
        layout.addWidget(desc)

        # æµ‹è¯•å‚æ•°
        params_group = QGroupBox("æµ‹è¯•å‚æ•°")
        params_layout = QFormLayout(params_group)

        self.edit_test_current = QLineEdit("1000")
        self.edit_test_current.setPlaceholderText("æ’æµç”µæµ (mA)")
        params_layout.addRow("æ’æµç”µæµ (mA):", self.edit_test_current)

        layout.addWidget(params_group)

        # é€šé“é€‰æ‹©
        channel_group = QGroupBox("æµ‹è¯•é€šé“")
        channel_layout = QVBoxLayout(channel_group)
        channel_layout.setSpacing(8)

        self.radio_ternary_mah = QRadioButton("ä¸‰å…ƒç”µæ± ")
        self.radio_ternary_mah.setChecked(True)
        self.radio_blade_mah = QRadioButton("åˆ€ç‰‡ç”µæ± ")

        channel_layout.addWidget(self.radio_ternary_mah)
        channel_layout.addWidget(self.radio_blade_mah)

        layout.addWidget(channel_group)

        # å®æ—¶æ•°æ®æ˜¾ç¤º
        display_group = QGroupBox("å®æ—¶æ•°æ®ï¼ˆæ¥è‡ªé‡‡é›†ï¼‰")
        display_layout = QFormLayout(display_group)

        self.label_status = QLabel("æœªå¼€å§‹")
        self.label_status.setStyleSheet("color: #888;")
        self.label_current = QLabel("0.00 mA")
        self.label_capacity = QLabel("0.00 mAh")
        self.label_voltage = QLabel("0.00 V")
        self.label_time = QLabel("00:00:00")
        self.label_points = QLabel("0")

        display_layout.addRow("æµ‹è¯•çŠ¶æ€:", self.label_status)
        display_layout.addRow("è®¾å®šç”µæµ:", self.label_current)
        display_layout.addRow("å®æ—¶ç”µå‹:", self.label_voltage)
        display_layout.addRow("ç´¯è®¡å®¹é‡:", self.label_capacity)
        display_layout.addRow("æµ‹è¯•æ—¶é•¿:", self.label_time)
        display_layout.addRow("æ•°æ®ç‚¹æ•°:", self.label_points)

        layout.addWidget(display_group)

        # æŒ‰é’®åŒºåŸŸ
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        btn_close = QPushButton("å…³é—­")
        btn_close.clicked.connect(self.reject)
        button_layout.addWidget(btn_close)

        self.btn_stop_test = QPushButton("åœæ­¢æµ‹è¯•")
        self.btn_stop_test.setStyleSheet("""
            QPushButton {
                background-color: #f59e0b;
                color: #ffffff;
                border: none;
                padding: 10px 20px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #d97706;
            }
        """)
        self.btn_stop_test.clicked.connect(self._stop_capacity_test)
        self.btn_stop_test.setEnabled(False)
        button_layout.addWidget(self.btn_stop_test)

        self.btn_start_test = QPushButton("å¼€å§‹æµ‹è¯•")
        self.btn_start_test.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: #ffffff;
                border: none;
                padding: 10px 20px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #059669;
            }
        """)
        self.btn_start_test.clicked.connect(self._start_capacity_test)
        button_layout.addWidget(self.btn_start_test)

        layout.addLayout(button_layout)

        # å®šæ—¶å™¨ç”¨äºæ›´æ–°æ˜¾ç¤º
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self._update_test_display)

    def _start_capacity_test(self) -> None:
        """å¼€å§‹å®¹é‡æµ‹è¯•ã€‚"""
        try:
            current = float(self.edit_test_current.text())

            if current <= 0:
                raise ValueError("ç”µæµå¿…é¡»ä¸ºæ­£æ•°")

            # æ£€æŸ¥æ˜¯å¦æ­£åœ¨é‡‡é›†æ•°æ®
            main_window = self.parent()
            if not main_window or not hasattr(main_window, 'is_running') or not main_window.is_running:
                QMessageBox.warning(
                    self,
                    "æç¤º",
                    "è¯·å…ˆç‚¹å‡»ã€å¼€å§‹ã€‘æŒ‰é’®å¼€å§‹æ•°æ®é‡‡é›†ï¼Œ\nç„¶åå†å¯åŠ¨å®¹é‡æµ‹è¯•ã€‚"
                )
                return

            channel = "ternary" if self.radio_ternary_mah.isChecked() else "blade"
            channel_name = "ä¸‰å…ƒç”µæ± " if channel == "ternary" else "åˆ€ç‰‡ç”µæ± "

            # å¯åŠ¨åˆ†æå¼•æ“çš„mAhæµ‹è¯•
            if hasattr(main_window, 'analysis_engine'):
                main_window.analysis_engine.start_mah_test(current, channel)

            # æ›´æ–°UI
            self.btn_start_test.setEnabled(False)
            self.btn_stop_test.setEnabled(True)
            self.edit_test_current.setEnabled(False)
            self.radio_ternary_mah.setEnabled(False)
            self.radio_blade_mah.setEnabled(False)

            self.label_status.setText("æµ‹è¯•ä¸­...")
            self.label_status.setStyleSheet("color: #10b981; font-weight: bold;")
            self.label_current.setText(f"{current:.2f} mA")

            # å¼€å§‹å®šæ—¶æ›´æ–°æ˜¾ç¤ºï¼ˆæ¯500msï¼‰
            self.update_timer.start(500)

            self.statusBar_message = f"mAhå®¹é‡æµ‹è¯•å·²å¯åŠ¨ ({channel_name}, {current}mA)"

        except ValueError as e:
            QMessageBox.warning(self, "å‚æ•°é”™è¯¯", str(e))

    def _stop_capacity_test(self) -> None:
        """åœæ­¢å®¹é‡æµ‹è¯•ã€‚"""
        self.update_timer.stop()

        main_window = self.parent()
        final_capacity = 0.0

        if main_window and hasattr(main_window, 'analysis_engine'):
            final_capacity = main_window.analysis_engine.stop_mah_test()

        # æ›´æ–°UI
        self.btn_start_test.setEnabled(True)
        self.btn_stop_test.setEnabled(False)
        self.edit_test_current.setEnabled(True)
        self.radio_ternary_mah.setEnabled(True)
        self.radio_blade_mah.setEnabled(True)

        self.label_status.setText("å·²å®Œæˆ")
        self.label_status.setStyleSheet("color: #6dd5ed; font-weight: bold;")

        QMessageBox.information(
            self,
            "æµ‹è¯•å®Œæˆ",
            f"å®¹é‡æµ‹è¯•å·²å®Œæˆï¼\n\næœ€ç»ˆç´¯è®¡å®¹é‡: {final_capacity:.3f} mAh"
        )

    def _update_test_display(self) -> None:
        """æ›´æ–°æµ‹è¯•æ˜¾ç¤ºï¼ˆä»å®é™…é‡‡é›†æ•°æ®è·å–ï¼‰ã€‚"""
        main_window = self.parent()
        if not main_window or not hasattr(main_window, 'analysis_engine'):
            return

        info = main_window.analysis_engine.get_mah_test_info()

        # æ›´æ–°æ˜¾ç¤º
        self.label_capacity.setText(f"{info['capacity_mah']:.3f} mAh")
        self.label_voltage.setText(f"{info['current_voltage']:.3f} V")
        self.label_points.setText(f"{info['data_points']}")

        # æ ¼å¼åŒ–æ—¶é—´
        elapsed = int(info['elapsed_time'])
        hours = elapsed // 3600
        minutes = (elapsed % 3600) // 60
        seconds = elapsed % 60
        self.label_time.setText(f"{hours:02d}:{minutes:02d}:{seconds:02d}")

        # å¦‚æœæµ‹è¯•ä¸å†æ´»è·ƒï¼Œè‡ªåŠ¨åœæ­¢æ›´æ–°
        if not info['active']:
            self._stop_capacity_test()
