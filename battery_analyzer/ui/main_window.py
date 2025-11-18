#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Battery Analyzer 主窗口（简体中文 UI 骨架）。

布局目标：
- 左侧控制面板：测试设置、导出报告、采集时长/点数、保存/召回波形、产品信息、开始按钮。
- 中部：双波形区域（左：三元电池；右：刀片电池）。
- 底部：关键指标 KPI（电压、温度），各自大号数字显示。

后续会把 LR8450 的实时采集与分析（温升比对、电压压降、mX+b、mAh）接入到相应槽函数。
"""

from __future__ import annotations

from typing import Optional

from PySide6.QtCore import Qt, Signal, QTimer
from PySide6.QtWidgets import (
    QWidget,
    QMainWindow,
    QHBoxLayout,
    QVBoxLayout,
    QGridLayout,
    QGroupBox,
    QLabel,
    QPushButton,
    QFormLayout,
    QLineEdit,
    QSpinBox,
    QComboBox,
    QApplication,
    QStatusBar,
    QDialog,
    QRadioButton,
    QMessageBox,
)

import pyqtgraph as pg
import numpy as np

from battery_analyzer.core.lr8450_client import LR8450Client
from battery_analyzer.core.analysis_engine import BatteryAnalysisEngine
from battery_analyzer.core.acquisition_thread import DataAcquisitionThread
from battery_analyzer.ui.dialogs.channel_config_dialog import ChannelConfigDialog
from battery_analyzer.ui.dialogs.device_connect_dialog import DeviceConnectDialog


class KPIWidget(QWidget):
    """KPI 数字显示（标题 + 数值+单位在同一框内）。"""

    def __init__(self, title: str, unit: str, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.unit = unit

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        self.title_label = QLabel(title)
        self.title_label.setObjectName("kpiTitle")
        self.title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.title_label.setProperty("class", "kpiTitle")

        # 数值+单位放在同一个标签内
        self.value_label = QLabel(f"0.00{unit}")
        self.value_label.setObjectName("valueLabel")
        self.value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.value_label.setProperty("class", "kpi")
        self.value_label.setStyleSheet(
            "background-color: #0b213f; "
            "border: 1px solid #2b4b7d; "
            "border-radius: 6px; "
            "padding: 8px 12px; "
            "color: #98d2ff; "
            "font-weight: bold; "
            "font-size: 28px;"
        )

        layout.addWidget(self.title_label)
        layout.addWidget(self.value_label)

    def set_value(self, value_text: str) -> None:
        """设置数值（自动附加单位）。"""
        self.value_label.setText(f"{value_text}{self.unit}")


class WaveformPair(QWidget):
    """双波形显示面板（每个图表有双Y轴：左侧电压，右侧温度）。"""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        layout = QGridLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(12)

        # 左：三元电池
        self.left_title = QLabel("三元电池波形 (Ternary Battery Waveform)")
        self.left_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.left_title.setStyleSheet("color: #6dd5ed; font-size: 14px; font-weight: bold;")
        
        self.left_plot = self._create_dual_axis_plot()
        
        # 右：刀片电池
        self.right_title = QLabel("刀片电池波形 (Blade Battery Waveform)")
        self.right_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.right_title.setStyleSheet("color: #6dd5ed; font-size: 14px; font-weight: bold;")
        
        self.right_plot = self._create_dual_axis_plot()

        layout.addWidget(self.left_title, 0, 0)
        layout.addWidget(self.right_title, 0, 1)
        layout.addWidget(self.left_plot, 1, 0)
        layout.addWidget(self.right_plot, 1, 1)
        layout.setColumnStretch(0, 1)
        layout.setColumnStretch(1, 1)

    def _create_dual_axis_plot(self) -> pg.PlotWidget:
        """创建带双Y轴的图表（左：电压，右：温度）。"""
        plot_widget = pg.PlotWidget()
        plot_widget.setBackground("#0c1830")
        
        # 隐藏左下角的 AutoRange 按钮（A按钮）
        plot_widget.plotItem.hideButtons()
        
        # 左Y轴：电压（V）
        plot_widget.setLabel('left', '电压 V', color='#33c1ff', **{'font-size': '11pt'})
        plot_widget.setYRange(0, 10)
        plot_widget.getAxis('left').setPen(pg.mkPen('#33c1ff', width=2))
        plot_widget.getAxis('left').setTextPen('#33c1ff')
        
        # 底部X轴：时间（Time）
        plot_widget.setLabel('bottom', '时间 Time', units='s', color='#6dd5ed', **{'font-size': '11pt'})
        plot_widget.setXRange(0, 300)
        plot_widget.getAxis('bottom').setPen(pg.mkPen('#4a90e2', width=2))
        plot_widget.getAxis('bottom').setTextPen('#6dd5ed')
        
        # 右Y轴：温度（T）
        viewbox_temp = pg.ViewBox()
        plot_widget.plotItem.scene().addItem(viewbox_temp)
        plot_widget.plotItem.getAxis('right').linkToView(viewbox_temp)
        viewbox_temp.setXLink(plot_widget.plotItem)
        
        # 配置右轴
        plot_widget.plotItem.showAxis('right')
        plot_widget.setLabel('right', '温度 T', units='°C', color='#ffb347', **{'font-size': '11pt'})
        plot_widget.getAxis('right').setPen(pg.mkPen('#ffb347', width=2))
        plot_widget.getAxis('right').setTextPen('#ffb347')
        
        # 同步右侧 ViewBox 的几何位置
        def update_views():
            viewbox_temp.setGeometry(plot_widget.plotItem.vb.sceneBoundingRect())
            viewbox_temp.linkedViewChanged(plot_widget.plotItem.vb, viewbox_temp.XAxis)
        
        update_views()
        plot_widget.plotItem.vb.sigResized.connect(update_views)
        
        # 设置温度轴范围
        viewbox_temp.setYRange(0, 260)
        
        # 存储 viewbox 以便后续绘图
        plot_widget.viewbox_temp = viewbox_temp
        
        # 网格
        plot_widget.showGrid(x=True, y=True, alpha=0.15)
        
        return plot_widget


class ControlPanel(QWidget):
    """左侧控制面板。"""

    start_requested = Signal()
    voltage_color_changed = Signal(str)  # 电压颜色改变信号
    temp_color_changed = Signal(str)     # 温度颜色改变信号
    voltage_width_changed = Signal(int)  # 电压线宽改变信号
    temp_width_changed = Signal(int)     # 温度线宽改变信号

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(10)

        # 操作区
        ops = QGroupBox("操作")
        ops_layout = QVBoxLayout(ops)
        self.btn_device_connect = QPushButton("连接设备")
        self.btn_device_connect.setStyleSheet("""
            QPushButton {
                background-color: #0d5aa7;
                color: #ffffff;
                font-weight: bold;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #0c4d91;
            }
        """)
        self.btn_channel_config = QPushButton("通道配置")
        self.btn_report = QPushButton("导出报告")
        self.btn_save = QPushButton("保存波形")
        self.btn_recall = QPushButton("召回波形")
        ops_layout.addWidget(self.btn_device_connect)
        ops_layout.addWidget(self.btn_channel_config)
        ops_layout.addWidget(self.btn_report)
        ops_layout.addWidget(self.btn_save)
        ops_layout.addWidget(self.btn_recall)

        # 通道与显示选项
        vis = QGroupBox("显示设置")
        vis_layout = QVBoxLayout(vis)
        vis_layout.setSpacing(8)
        
        # 电压行：标签 + 像素选择 + 颜色块
        volt_row = QHBoxLayout()
        volt_label = QLabel("电压")
        volt_label.setFixedWidth(40)
        volt_row.addWidget(volt_label)
        self.combo_volt_pix = QComboBox()
        self.combo_volt_pix.addItems(["1像素", "2像素", "3像素", "4像素"])
        self.combo_volt_pix.setCurrentIndex(2)  # 默认3像素
        self.combo_volt_pix.setFixedWidth(80)
        self.combo_volt_pix.currentIndexChanged.connect(self._on_volt_width_changed)
        volt_row.addWidget(self.combo_volt_pix)
        self.volt_color_btn = QPushButton()
        self.volt_color_btn.setFixedSize(50, 22)
        self.volt_color_btn.setStyleSheet("background-color: #ff9933; border: 1px solid #2b4b7d; border-radius: 3px;")
        self.volt_color_btn.clicked.connect(lambda: self._choose_color("volt"))
        volt_row.addWidget(self.volt_color_btn)
        volt_row.addStretch()
        vis_layout.addLayout(volt_row)

        # 温度行：标签 + 像素选择 + 颜色块
        temp_row = QHBoxLayout()
        temp_label = QLabel("温度")
        temp_label.setFixedWidth(40)
        temp_row.addWidget(temp_label)
        self.combo_temp_pix = QComboBox()
        self.combo_temp_pix.addItems(["1像素", "2像素", "3像素", "4像素"])
        self.combo_temp_pix.setCurrentIndex(2)  # 默认3像素
        self.combo_temp_pix.setFixedWidth(80)
        self.combo_temp_pix.currentIndexChanged.connect(self._on_temp_width_changed)
        temp_row.addWidget(self.combo_temp_pix)
        self.temp_color_btn = QPushButton()
        self.temp_color_btn.setFixedSize(50, 22)
        self.temp_color_btn.setStyleSheet("background-color: #66ccff; border: 1px solid #2b4b7d; border-radius: 3px;")
        self.temp_color_btn.clicked.connect(lambda: self._choose_color("temp"))
        temp_row.addWidget(self.temp_color_btn)
        temp_row.addStretch()
        vis_layout.addLayout(temp_row)
        
        # 存储当前颜色
        self.volt_color = "#ff9933"
        self.temp_color = "#66ccff"

        # 产品信息（上下排列，紧凑布局）
        info = QGroupBox("产品信息")
        info_layout = QVBoxLayout(info)
        info_layout.setSpacing(6)
        
        # 产品型号
        info_layout.addWidget(QLabel("产品型号"))
        self.edit_model = QLineEdit("5mm")
        self.edit_model.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.edit_model.setStyleSheet("font-size: 18px; font-weight: bold; color: #33c1ff; padding: 6px;")
        info_layout.addWidget(self.edit_model)
        
        # 产品流水号
        info_layout.addWidget(QLabel("产品流水号"))
        self.edit_sn = QLineEdit("25mm/s")
        self.edit_sn.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.edit_sn.setStyleSheet("font-size: 18px; font-weight: bold; color: #33c1ff; padding: 6px;")
        info_layout.addWidget(self.edit_sn)
        
        # 测试员
        info_layout.addWidget(QLabel("测试员"))
        self.edit_tester = QLineEdit("1dot/s")
        self.edit_tester.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.edit_tester.setStyleSheet("font-size: 18px; font-weight: bold; color: #33c1ff; padding: 6px;")
        info_layout.addWidget(self.edit_tester)

        # 开始按钮
        self.btn_start = QPushButton("开始")
        self.btn_start.setObjectName("startButton")
        self.btn_start.clicked.connect(self.start_requested.emit)

        # 新功能按钮区
        functions = QGroupBox("分析功能")
        func_layout = QVBoxLayout(functions)
        func_layout.setSpacing(8)

        self.btn_mx_plus_b = QPushButton("mX+b")
        self.btn_mx_plus_b.setStyleSheet("""
            QPushButton {
                background-color: #1e3a8a;
                color: #ffffff;
                border: 2px solid #3b82f6;
                border-radius: 6px;
                padding: 10px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #1d4ed8;
                border-color: #60a5fa;
                color: #ffffff;
            }
        """)
        func_layout.addWidget(self.btn_mx_plus_b)

        self.btn_mah_test = QPushButton("mAh 容量测试")
        self.btn_mah_test.setStyleSheet("""
            QPushButton {
                background-color: #059669;
                color: #ffffff;
                border: 2px solid #10b981;
                border-radius: 6px;
                padding: 10px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #047857;
                border-color: #34d399;
                color: #ffffff;
            }
        """)
        func_layout.addWidget(self.btn_mah_test)

        root.addWidget(ops)
        root.addWidget(functions)
        root.addWidget(vis)
        root.addWidget(info)
        root.addStretch(1)
        root.addWidget(self.btn_start)
    
    def _choose_color(self, channel: str) -> None:
        """选择通道颜色。"""
        from PySide6.QtGui import QColor
        from battery_analyzer.ui.color_dialog import SimpleColorDialog
        
        # 获取当前颜色
        current_color = QColor(self.volt_color if channel == "volt" else self.temp_color)
        
        # 打开自定义中文颜色选择对话框
        title = "选择电压颜色" if channel == "volt" else "选择温度颜色"
        color = SimpleColorDialog.get_color_from_dialog(current_color, title, self)
        
        if color and color.isValid():
            color_hex = color.name()
            if channel == "volt":
                self.volt_color = color_hex
                self.volt_color_btn.setStyleSheet(
                    f"background-color: {color_hex}; border: 1px solid #2b4b7d; border-radius: 3px;"
                )
                # 发射信号通知主窗口更新电压曲线颜色
                self.voltage_color_changed.emit(color_hex)
            else:
                self.temp_color = color_hex
                self.temp_color_btn.setStyleSheet(
                    f"background-color: {color_hex}; border: 1px solid #2b4b7d; border-radius: 3px;"
                )
                # 发射信号通知主窗口更新温度曲线颜色
                self.temp_color_changed.emit(color_hex)

    def _on_volt_width_changed(self, index: int) -> None:
        """电压线宽改变"""
        width = index + 1  # 0->1, 1->2, 2->3, 3->4
        self.voltage_width_changed.emit(width)

    def _on_temp_width_changed(self, index: int) -> None:
        """温度线宽改变"""
        width = index + 1  # 0->1, 1->2, 2->3, 3->4
        self.temp_width_changed.emit(width)


class MainWindow(QMainWindow):
    """主窗口。"""

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("电池电压与温升分析软件")
        self.resize(1280, 800)

        # 中央布局（包含顶部标题栏）
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # 顶部标题栏
        title_bar = self._create_title_bar()
        main_layout.addWidget(title_bar)

        # 主内容区域
        content = QWidget()
        hbox = QHBoxLayout(content)
        hbox.setContentsMargins(8, 8, 8, 8)
        hbox.setSpacing(12)

        # 左侧控制
        self.control = ControlPanel()
        self.control.setFixedWidth(260)
        self.control.start_requested.connect(self._on_start)
        self.control.voltage_color_changed.connect(self.update_voltage_color)
        self.control.temp_color_changed.connect(self.update_temp_color)
        self.control.voltage_width_changed.connect(self.update_voltage_width)
        self.control.temp_width_changed.connect(self.update_temp_width)

        # 连接操作按钮
        self.control.btn_device_connect.clicked.connect(self._show_device_connect_dialog)
        self.control.btn_channel_config.clicked.connect(self._show_channel_config_dialog)
        self.control.btn_report.clicked.connect(self._export_report)
        self.control.btn_save.clicked.connect(self._save_waveform)
        self.control.btn_recall.clicked.connect(self._recall_waveform)
        
        # 连接分析功能按钮
        self.control.btn_mx_plus_b.clicked.connect(self._show_mx_plus_b_dialog)
        self.control.btn_mah_test.clicked.connect(self._show_mah_test_dialog)

        # 中部双波形
        self.waveforms = WaveformPair()

        # 底部 KPI 区域（4个KPI：左侧2个三元电池，右侧2个刀片电池）
        bottom = QWidget()
        bottom_layout = QHBoxLayout(bottom)
        bottom_layout.setContentsMargins(0, 0, 0, 0)
        bottom_layout.setSpacing(12)
        
        # 三元电池电压
        self.ternary_voltage_kpi = KPIWidget("三元电池电压\nTernary Battery Voltage", "V")
        bottom_layout.addWidget(self.ternary_voltage_kpi)
        
        # 三元电池温度
        self.ternary_temp_kpi = KPIWidget("三元电池温度\nTernary Battery Temperature", "°C")
        bottom_layout.addWidget(self.ternary_temp_kpi)
        
        # 刀片电池电压
        self.blade_voltage_kpi = KPIWidget("刀片电池电压\nBlade Battery Voltage", "V")
        bottom_layout.addWidget(self.blade_voltage_kpi)
        
        # 刀片电池温度
        self.blade_temp_kpi = KPIWidget("刀片电池温度\nBlade Battery Temperature", "°C")
        bottom_layout.addWidget(self.blade_temp_kpi)

        # 将中部区域组装为上下两块
        center_col = QWidget()
        center_layout = QVBoxLayout(center_col)
        center_layout.setContentsMargins(0, 0, 0, 0)
        center_layout.setSpacing(12)
        center_layout.addWidget(self.waveforms, 1)
        center_layout.addWidget(bottom, 0)

        hbox.addWidget(self.control, 0)
        hbox.addWidget(center_col, 1)

        main_layout.addWidget(content, 1)

        # 状态栏
        status = QStatusBar()
        self.setStatusBar(status)
        status.showMessage("就绪——等待连接与开始测试")

        # 存储曲线对象以便更新颜色
        self.volt_curves = []  # [左图电压曲线, 右图电压曲线]
        self.temp_curves = []  # [左图温度曲线, 右图温度曲线]
        
        # LR8450设备客户端
        self.device_client: Optional[LR8450Client] = None
        self.device_connected = False

        # 数据采集线程
        self.acquisition_thread: Optional[DataAcquisitionThread] = None

        # 分析引擎
        self.analysis_engine = BatteryAnalysisEngine()

        # 通道配置（包含详细参数）
        self.channel_config = {
            'ternary_voltage': {
                'channel': 'CH2_1',
                'type': 'VOLTAGE',
                'range': 10.0,
            },
            'ternary_temp': {
                'channel': 'CH2_3',
                'type': 'TEMPERATURE',
                'range': 500,
                'thermocouple': 'K',
                'int_ext': 'INT',
            },
            'blade_voltage': {
                'channel': 'CH2_4',
                'type': 'VOLTAGE',
                'range': 10.0,
            },
            'blade_temp': {
                'channel': 'CH2_5',
                'type': 'TEMPERATURE',
                'range': 500,
                'thermocouple': 'K',
                'int_ext': 'INT',
            },
        }

        # 数据采集状态
        self.is_running = False
        self.data_index = 0
        self.max_points = 600  # 最多显示600个点（300秒）

        # 数据缓冲区
        self.x_data = []
        self.ternary_volt_data = []
        self.ternary_temp_data = []
        self.blade_volt_data = []
        self.blade_temp_data = []

        # 当前颜色和线宽（从控制面板获取）
        self.current_volt_color = "#ff9933"
        self.current_temp_color = "#66ccff"
        self.current_volt_width = 3  # 默认3像素
        self.current_temp_width = 3  # 默认3像素

        # 定时器：用于虚拟数据模式
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self._update_waveform_virtual)
        self.update_interval_ms = 100  # 更新间隔（毫秒）

        # 预置示例波形（初始静态数据）
        self._plot_demo()

        # 初始化Y轴范围（根据配置的量程）
        self._update_plot_ranges()

    def _plot_demo(self) -> None:
        """绘制示例波形（电压在左Y轴，温度在右Y轴）。"""
        import numpy as np

        x = np.linspace(0, 300, 600)
        
        # 三元电池数据
        y_v_ternary = 5 + 0.3 * np.sin(0.05 * x) + 0.1 * np.random.randn(600)  # 电压 0-10V
        y_t_ternary = 130 + 50 * np.sin(0.02 * x + 1.2) + 5 * np.random.randn(600)  # 温度 0-260°C
        
        # 刀片电池数据
        y_v_blade = 5.2 + 0.25 * np.sin(0.05 * x + 0.4) + 0.08 * np.random.randn(600)
        y_t_blade = 120 + 40 * np.sin(0.02 * x + 2.0) + 4 * np.random.randn(600)

        # 左图：三元电池
        # 电压曲线（左Y轴，主ViewBox）
        volt_curve_left = self.waveforms.left_plot.plot(x, y_v_ternary, pen=pg.mkPen(self.current_volt_color, width=self.current_volt_width), name="电压")
        self.volt_curves.append(volt_curve_left)

        # 温度曲线（右Y轴，viewbox_temp）
        temp_curve_left = pg.PlotCurveItem(x, y_t_ternary, pen=pg.mkPen(self.current_temp_color, width=self.current_temp_width), name="温度")
        self.waveforms.left_plot.viewbox_temp.addItem(temp_curve_left)
        self.temp_curves.append(temp_curve_left)

        # 右图：刀片电池
        # 电压曲线（左Y轴）
        volt_curve_right = self.waveforms.right_plot.plot(x, y_v_blade, pen=pg.mkPen(self.current_volt_color, width=self.current_volt_width), name="电压")
        self.volt_curves.append(volt_curve_right)

        # 温度曲线（右Y轴）
        temp_curve_right = pg.PlotCurveItem(x, y_t_blade, pen=pg.mkPen(self.current_temp_color, width=self.current_temp_width), name="温度")
        self.waveforms.right_plot.viewbox_temp.addItem(temp_curve_right)
        self.temp_curves.append(temp_curve_right)

        # 初始化 KPI
        self.ternary_voltage_kpi.set_value("0.00")
        self.ternary_temp_kpi.set_value("0.00")
        self.blade_voltage_kpi.set_value("0.00")
        self.blade_temp_kpi.set_value("0.00")

    def _create_title_bar(self) -> QWidget:
        """创建顶部标题栏（科技风格）。"""
        title_container = QWidget()
        title_container.setObjectName("titleBar")
        title_container.setFixedHeight(60)
        title_container.setStyleSheet("""
            QWidget#titleBar {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #0a1a35, stop:0.5 #0c1f3f, stop:1 #0a1a35);
                border-bottom: 2px solid #1a4d7d;
            }
            QLabel#mainTitle {
                color: #6dd5ed;
                font-size: 20px;
                font-weight: bold;
                letter-spacing: 2px;
            }
            QPushButton#titleBtn {
                background-color: transparent;
                color: #6dd5ed;
                border: 1px solid #2b5a8a;
                border-radius: 4px;
                padding: 6px 16px;
                font-size: 13px;
            }
            QPushButton#titleBtn:hover {
                background-color: #1a3a5f;
                border-color: #4a8fdd;
            }
        """)

        layout = QHBoxLayout(title_container)
        layout.setContentsMargins(20, 10, 20, 10)

        layout.addStretch()
        
        # 主标题
        title = QLabel("电池电压与温升采集软件")
        title.setObjectName("mainTitle")
        layout.addWidget(title)

        layout.addStretch()

        # 帮助按钮
        btn_help = QPushButton("❓ Help")
        btn_help.setObjectName("titleBtn")
        btn_help.clicked.connect(self._show_help)
        layout.addWidget(btn_help)

        # 退出按钮
        btn_exit = QPushButton("✖ Exit")
        btn_exit.setObjectName("titleBtn")
        btn_exit.clicked.connect(self.close)
        layout.addWidget(btn_exit)

        return title_container

    def _show_help(self) -> None:
        """显示帮助信息。"""
        from PySide6.QtWidgets import QMessageBox
        QMessageBox.information(
            self,
            "帮助",
            "电池电压与温升分析软件\n\n"
            "功能：\n"
            "• 三元电池与刀片电池温升比对\n"
            "• 电池压降采集分析\n"
            "• mX+b 线性校准\n"
            "• mAh 容量测试\n\n"
            "后续将接入 LR8450 设备进行实时数据采集。"
        )

    def update_voltage_color(self, color_hex: str) -> None:
        """更新电压曲线颜色和左Y轴颜色。"""
        self.current_volt_color = color_hex  # 更新当前颜色
        
        # 更新曲线颜色
        for curve in self.volt_curves:
            curve.setPen(pg.mkPen(color_hex, width=2))
        
        # 更新左Y轴（电压轴）颜色
        self.waveforms.left_plot.setLabel('left', '电压 V', color=color_hex, **{'font-size': '11pt'})
        self.waveforms.left_plot.getAxis('left').setPen(pg.mkPen(color_hex, width=2))
        self.waveforms.left_plot.getAxis('left').setTextPen(color_hex)
        
        self.waveforms.right_plot.setLabel('left', '电压 V', color=color_hex, **{'font-size': '11pt'})
        self.waveforms.right_plot.getAxis('left').setPen(pg.mkPen(color_hex, width=2))
        self.waveforms.right_plot.getAxis('left').setTextPen(color_hex)
        
        self.statusBar().showMessage(f"电压曲线和坐标轴颜色已更新为 {color_hex}")
    
    def update_temp_color(self, color_hex: str) -> None:
        """更新温度曲线颜色和右Y轴颜色。"""
        self.current_temp_color = color_hex  # 更新当前颜色
        
        # 更新曲线颜色
        for curve in self.temp_curves:
            curve.setPen(pg.mkPen(color_hex, width=2))
        
        # 更新右Y轴（温度轴）颜色
        self.waveforms.left_plot.setLabel('right', '温度 T', units='°C', color=color_hex, **{'font-size': '11pt'})
        self.waveforms.left_plot.getAxis('right').setPen(pg.mkPen(color_hex, width=2))
        self.waveforms.left_plot.getAxis('right').setTextPen(color_hex)
        
        self.waveforms.right_plot.setLabel('right', '温度 T', units='°C', color=color_hex, **{'font-size': '11pt'})
        self.waveforms.right_plot.getAxis('right').setPen(pg.mkPen(color_hex, width=2))
        self.waveforms.right_plot.getAxis('right').setTextPen(color_hex)
        
        self.statusBar().showMessage(f"温度曲线和坐标轴颜色已更新为 {color_hex}")

    def update_voltage_width(self, width: int) -> None:
        """更新电压曲线线宽"""
        self.current_volt_width = width

        # 更新曲线线宽
        for curve in self.volt_curves:
            curve.setPen(pg.mkPen(self.current_volt_color, width=width))

        self.statusBar().showMessage(f"电压曲线线宽已更新为 {width} 像素")

    def update_temp_width(self, width: int) -> None:
        """更新温度曲线线宽"""
        self.current_temp_width = width

        # 更新曲线线宽
        for curve in self.temp_curves:
            curve.setPen(pg.mkPen(self.current_temp_color, width=width))

        self.statusBar().showMessage(f"温度曲线线宽已更新为 {width} 像素")

    def _show_mx_plus_b_dialog(self) -> None:
        """显示mX+b校准对话框。"""
        dialog = MXPlusBDialog(self)
        dialog.exec()

    def _show_mah_test_dialog(self) -> None:
        """显示mAh容量测试对话框。"""
        dialog = MAHTestDialog(self)
        dialog.exec()

    def _on_data_acquired(self, timestamp: float, data: dict) -> None:
        """处理从采集线程接收到的数据（真实设备数据）

        Args:
            timestamp: 时间戳（秒）
            data: 通道数据字典
        """
        # 提取数据（使用通道名称）
        v_ternary = data.get(self.channel_config['ternary_voltage']['channel'], 0.0)
        t_ternary = data.get(self.channel_config['ternary_temp']['channel'], 0.0)
        v_blade = data.get(self.channel_config['blade_voltage']['channel'], 0.0)
        t_blade = data.get(self.channel_config['blade_temp']['channel'], 0.0)

        # 添加到分析引擎
        self.analysis_engine.add_data_point(v_ternary, t_ternary, v_blade, t_blade, timestamp)

        # 添加到缓冲区
        self.x_data.append(timestamp)
        self.ternary_volt_data.append(v_ternary)
        self.ternary_temp_data.append(t_ternary)
        self.blade_volt_data.append(v_blade)
        self.blade_temp_data.append(t_blade)

        # 限制数据点数量（滚动显示）
        if len(self.x_data) > self.max_points:
            self.x_data.pop(0)
            self.ternary_volt_data.pop(0)
            self.ternary_temp_data.pop(0)
            self.blade_volt_data.pop(0)
            self.blade_temp_data.pop(0)

        # 更新曲线
        if len(self.volt_curves) >= 2 and len(self.temp_curves) >= 2:
            self.volt_curves[0].setData(self.x_data, self.ternary_volt_data)
            self.temp_curves[0].setData(self.x_data, self.ternary_temp_data)
            self.volt_curves[1].setData(self.x_data, self.blade_volt_data)
            self.temp_curves[1].setData(self.x_data, self.blade_temp_data)

        # 更新KPI显示
        self.ternary_voltage_kpi.set_value(f"{v_ternary:.2f}")
        self.ternary_temp_kpi.set_value(f"{t_ternary:.2f}")
        self.blade_voltage_kpi.set_value(f"{v_blade:.2f}")
        self.blade_temp_kpi.set_value(f"{t_blade:.2f}")

    def _on_acquisition_error(self, error_msg: str) -> None:
        """处理采集线程的错误

        Args:
            error_msg: 错误消息
        """
        print(f"⚠️ 采集错误: {error_msg}")
        # 可以选择显示在状态栏或弹窗
        # self.statusBar().showMessage(f"采集错误: {error_msg}")

    def _on_acquisition_status(self, status_msg: str) -> None:
        """处理采集线程的状态变化

        Args:
            status_msg: 状态消息
        """
        print(f"ℹ️ 采集状态: {status_msg}")

    def _update_waveform_virtual(self) -> None:
        """定时更新波形（虚拟数据模式）"""
        # 修正时间戳计算：100ms间隔 = 0.1秒
        t = self.data_index * (self.update_interval_ms / 1000.0)

        # 生成虚拟数据
        v_ternary, t_ternary, v_blade, t_blade = self._generate_virtual_data(t)

        # 添加到分析引擎
        self.analysis_engine.add_data_point(v_ternary, t_ternary, v_blade, t_blade, t)

        # 添加到缓冲区
        self.x_data.append(t)
        self.ternary_volt_data.append(v_ternary)
        self.ternary_temp_data.append(t_ternary)
        self.blade_volt_data.append(v_blade)
        self.blade_temp_data.append(t_blade)

        # 限制数据点数量（滚动显示）
        if len(self.x_data) > self.max_points:
            self.x_data.pop(0)
            self.ternary_volt_data.pop(0)
            self.ternary_temp_data.pop(0)
            self.blade_volt_data.pop(0)
            self.blade_temp_data.pop(0)

        # 更新曲线
        if len(self.volt_curves) >= 2 and len(self.temp_curves) >= 2:
            self.volt_curves[0].setData(self.x_data, self.ternary_volt_data)
            self.temp_curves[0].setData(self.x_data, self.ternary_temp_data)
            self.volt_curves[1].setData(self.x_data, self.blade_volt_data)
            self.temp_curves[1].setData(self.x_data, self.blade_temp_data)

        # 更新KPI显示
        self.ternary_voltage_kpi.set_value(f"{v_ternary:.2f}")
        self.ternary_temp_kpi.set_value(f"{t_ternary:.2f}")
        self.blade_voltage_kpi.set_value(f"{v_blade:.2f}")
        self.blade_temp_kpi.set_value(f"{t_blade:.2f}")

        self.data_index += 1
    
    def _generate_virtual_data(self, t: float) -> tuple[float, float, float, float]:
        """生成虚拟数据（设备未连接时使用）"""
        v_ternary = 5 + 0.3 * np.sin(0.05 * t) + 0.1 * np.random.randn()
        t_ternary = 130 + 50 * np.sin(0.02 * t + 1.2) + 5 * np.random.randn()
        v_blade = 5.2 + 0.25 * np.sin(0.05 * t + 0.4) + 0.08 * np.random.randn()
        t_blade = 120 + 40 * np.sin(0.02 * t + 2.0) + 4 * np.random.randn()
        return v_ternary, t_ternary, v_blade, t_blade
    
    # 槽函数：开始/停止数据采集
    def _on_start(self) -> None:
        """开始数据采集（真实设备或虚拟数据）"""
        if not self.is_running:
            self.is_running = True
            self.data_index = 0
            self.x_data.clear()
            self.ternary_volt_data.clear()
            self.ternary_temp_data.clear()
            self.blade_volt_data.clear()
            self.blade_temp_data.clear()

            # 清空分析引擎数据
            self.analysis_engine.clear_data()

            # 如果设备已连接，使用后台线程采集真实数据
            if self.device_connected and self.device_client:
                # 获取通道列表（使用通道名称）
                channels = [
                    self.channel_config['ternary_voltage']['channel'],
                    self.channel_config['ternary_temp']['channel'],
                    self.channel_config['blade_voltage']['channel'],
                    self.channel_config['blade_temp']['channel'],
                ]

                # 创建并启动采集线程
                self.acquisition_thread = DataAcquisitionThread(
                    device_client=self.device_client,
                    channels=channels,
                    interval_ms=self.update_interval_ms
                )

                # 连接信号
                self.acquisition_thread.data_acquired.connect(self._on_data_acquired)
                self.acquisition_thread.error_occurred.connect(self._on_acquisition_error)
                self.acquisition_thread.status_changed.connect(self._on_acquisition_status)

                # 启动设备采集
                self.device_client.start_acquisition()

                # 启动线程
                self.acquisition_thread.start()

                self.statusBar().showMessage("✓ 真实设备数据采集进行中...")
            else:
                # 使用虚拟数据模式
                self.update_timer.start(self.update_interval_ms)
                self.statusBar().showMessage("虚拟数据采集进行中...")

            self.control.btn_start.setText("停止")
            self.control.btn_start.clicked.disconnect()
            self.control.btn_start.clicked.connect(self._on_stop)

    def _on_stop(self) -> None:
        """停止数据采集"""
        if self.is_running:
            self.is_running = False

            # 停止采集线程
            if self.acquisition_thread and self.acquisition_thread.is_running():
                self.acquisition_thread.stop()
                self.acquisition_thread = None

            # 停止虚拟数据定时器
            if self.update_timer.isActive():
                self.update_timer.stop()

            # 停止设备采集
            if self.device_connected and self.device_client:
                self.device_client.stop_acquisition()

            self.statusBar().showMessage("数据采集已停止")
            self.control.btn_start.setText("开始")
            self.control.btn_start.clicked.disconnect()
            self.control.btn_start.clicked.connect(self._on_start)
    
    def _show_device_connect_dialog(self) -> None:
        """显示设备连接对话框"""
        dialog = DeviceConnectDialog(self)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            params = dialog.get_connection_params()
            self._connect_to_device(params)
    
    def _connect_to_device(self, params: dict) -> None:
        """连接到LR8450设备（支持TCP和USB）

        Args:
            params: 连接参数字典，包含：
                - connection_type: "TCP" 或 "USB"
                - ip_address: TCP IP地址
                - port: TCP端口
                - com_port: COM端口
        """
        connection_type = params['connection_type']

        if connection_type == "TCP":
            ip = params['ip_address']
            port = params['port']
            connection_info = f"IP地址: {ip}\n端口: {port}"
            progress_text = f"正在通过TCP连接设备 {ip}:{port}..."
        else:  # USB
            com_port = params['com_port']
            connection_info = f"COM端口: {com_port}"
            progress_text = f"正在通过USB连接设备 {com_port}..."

        # 显示加载对话框
        from PySide6.QtWidgets import QProgressDialog
        progress = QProgressDialog(progress_text, None, 0, 0, self)
        progress.setWindowTitle("连接中")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        progress.show()
        QApplication.processEvents()

        try:
            # 断开旧连接
            if self.device_client:
                self.device_client.disconnect()

            # 创建新客户端
            self.device_client = LR8450Client(
                connection_type=connection_type,
                ip_address=params['ip_address'],
                port=params['port'],
                com_port=params['com_port']
            )

            # 连接
            if self.device_client.connect():
                self.device_connected = True
                conn_method = "TCP/IP" if connection_type == "TCP" else "USB串口"

                # 更新进度对话框
                progress.setLabelText(f"✓ 设备已连接，正在配置通道...")
                QApplication.processEvents()

                # 准备通道配置（使用当前配置的详细参数）
                channel_configs = [
                    {
                        'channel': self.channel_config['ternary_voltage']['channel'],
                        'type': self.channel_config['ternary_voltage']['type'],
                        'range': self.channel_config['ternary_voltage']['range'],
                    },
                    {
                        'channel': self.channel_config['ternary_temp']['channel'],
                        'type': self.channel_config['ternary_temp']['type'],
                        'range': self.channel_config['ternary_temp']['range'],
                        'thermocouple': self.channel_config['ternary_temp']['thermocouple'],
                        'int_ext': self.channel_config['ternary_temp']['int_ext'],
                    },
                    {
                        'channel': self.channel_config['blade_voltage']['channel'],
                        'type': self.channel_config['blade_voltage']['type'],
                        'range': self.channel_config['blade_voltage']['range'],
                    },
                    {
                        'channel': self.channel_config['blade_temp']['channel'],
                        'type': self.channel_config['blade_temp']['type'],
                        'range': self.channel_config['blade_temp']['range'],
                        'thermocouple': self.channel_config['blade_temp']['thermocouple'],
                        'int_ext': self.channel_config['blade_temp']['int_ext'],
                    },
                ]

                # 提取通道列表
                channels = [cfg['channel'] for cfg in channel_configs]

                # 配置通道（先禁用所有通道，然后只启用需要的通道）
                config_success = self.device_client.configure_channels(
                    channels=channels,
                    disable_others=True,  # 先禁用其他通道，防止数据错乱
                    channel_configs=channel_configs
                )

                # 关闭进度对话框
                progress.close()

                if config_success:
                    self.statusBar().showMessage(f"✓ 设备已通过{conn_method}连接，通道已配置")
                    QMessageBox.information(
                        self,
                        "连接成功",
                        f"成功通过{conn_method}连接到LR8450设备\n\n"
                        f"{connection_info}\n\n"
                        f"已自动配置以下通道：\n"
                        f"• {channels[0]} - 三元电池电压 ({self.channel_config['ternary_voltage']['range']}V)\n"
                        f"• {channels[1]} - 三元电池温度 ({self.channel_config['ternary_temp']['range']}°C, "
                        f"{self.channel_config['ternary_temp']['thermocouple']}型)\n"
                        f"• {channels[2]} - 刀片电池电压 ({self.channel_config['blade_voltage']['range']}V)\n"
                        f"• {channels[3]} - 刀片电池温度 ({self.channel_config['blade_temp']['range']}°C, "
                        f"{self.channel_config['blade_temp']['thermocouple']}型)\n\n"
                        f"现在可以开始测试了！"
                    )
                else:
                    self.statusBar().showMessage(f"⚠️ 设备已连接，但通道配置失败")
                    QMessageBox.warning(
                        self,
                        "通道配置警告",
                        f"设备连接成功，但通道配置失败。\n\n"
                        f"请在设备上手动启用以下通道：\n"
                        f"• {channels[0]}\n"
                        f"• {channels[1]}\n"
                        f"• {channels[2]}\n"
                        f"• {channels[3]}\n\n"
                        f"或者尝试重新连接。"
                    )
            else:
                # 关闭进度对话框
                progress.close()

                self.device_connected = False
                self.statusBar().showMessage("✗ 设备连接失败")

                if connection_type == "TCP":
                    error_msg = (
                        f"无法连接到设备 {ip}:{port}\n\n"
                        f"请检查：\n"
                        f"1. 设备电源是否开启\n"
                        f"2. 网络连接是否正常\n"
                        f"3. IP地址和端口是否正确\n"
                        f"4. 端口8802（SCPI控制端口）"
                    )
                else:
                    error_msg = (
                        f"无法连接到设备 {com_port}\n\n"
                        f"请检查：\n"
                        f"1. 设备电源是否开启\n"
                        f"2. USB线是否连接\n"
                        f"3. 是否已安装HIOKI USB驱动\n"
                        f"4. COM端口是否正确"
                    )

                QMessageBox.warning(self, "连接失败", error_msg)
        except Exception as e:
            # 关闭进度对话框
            progress.close()

            self.device_connected = False
            self.statusBar().showMessage("✗ 连接错误")
            QMessageBox.critical(self, "错误", f"连接失败：{str(e)}")
    
    def _show_channel_config_dialog(self) -> None:
        """显示通道配置对话框"""
        dialog = ChannelConfigDialog(self, current_config=self.channel_config)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.channel_config = dialog.get_config()
            self.statusBar().showMessage("✓ 通道配置已更新")

            # 更新Y轴范围（根据新的量程配置）
            self._update_plot_ranges()

            # 格式化配置信息
            config_info = (
                f"三元电池电压: {self.channel_config['ternary_voltage']['channel']} "
                f"({self.channel_config['ternary_voltage']['range']}V)\n"
                f"三元电池温度: {self.channel_config['ternary_temp']['channel']} "
                f"({self.channel_config['ternary_temp']['range']}°C, "
                f"{self.channel_config['ternary_temp']['thermocouple']}型, "
                f"{self.channel_config['ternary_temp']['int_ext']})\n"
                f"刀片电池电压: {self.channel_config['blade_voltage']['channel']} "
                f"({self.channel_config['blade_voltage']['range']}V)\n"
                f"刀片电池温度: {self.channel_config['blade_temp']['channel']} "
                f"({self.channel_config['blade_temp']['range']}°C, "
                f"{self.channel_config['blade_temp']['thermocouple']}型, "
                f"{self.channel_config['blade_temp']['int_ext']})"
            )

            # 如果设备已连接，立即应用新配置到设备
            if self.device_connected and self.device_client:
                reply = QMessageBox.question(
                    self,
                    "应用配置",
                    f"通道配置已更新：\n\n{config_info}\n\n"
                    f"是否立即应用到已连接的设备？\n"
                    f"（这将重新配置设备通道）",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes
                )

                if reply == QMessageBox.StandardButton.Yes:
                    self._apply_channel_config_to_device()
            else:
                QMessageBox.information(
                    self,
                    "配置成功",
                    f"通道配置已更新：\n\n{config_info}\n\n"
                    f"注意：如果设备已连接，请重新连接以应用新配置。"
                )

    def _apply_channel_config_to_device(self) -> None:
        """应用通道配置到已连接的设备"""
        if not self.device_connected or not self.device_client:
            return

        try:
            # 显示加载对话框
            from PySide6.QtWidgets import QProgressDialog
            progress = QProgressDialog("正在应用通道配置到设备...", None, 0, 0, self)
            progress.setWindowTitle("配置中")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(0)
            progress.show()
            QApplication.processEvents()

            # 准备通道配置
            channel_configs = [
                {
                    'channel': self.channel_config['ternary_voltage']['channel'],
                    'type': self.channel_config['ternary_voltage']['type'],
                    'range': self.channel_config['ternary_voltage']['range'],
                },
                {
                    'channel': self.channel_config['ternary_temp']['channel'],
                    'type': self.channel_config['ternary_temp']['type'],
                    'range': self.channel_config['ternary_temp']['range'],
                    'thermocouple': self.channel_config['ternary_temp']['thermocouple'],
                    'int_ext': self.channel_config['ternary_temp']['int_ext'],
                },
                {
                    'channel': self.channel_config['blade_voltage']['channel'],
                    'type': self.channel_config['blade_voltage']['type'],
                    'range': self.channel_config['blade_voltage']['range'],
                },
                {
                    'channel': self.channel_config['blade_temp']['channel'],
                    'type': self.channel_config['blade_temp']['type'],
                    'range': self.channel_config['blade_temp']['range'],
                    'thermocouple': self.channel_config['blade_temp']['thermocouple'],
                    'int_ext': self.channel_config['blade_temp']['int_ext'],
                },
            ]

            channels = [cfg['channel'] for cfg in channel_configs]

            # 配置通道
            config_success = self.device_client.configure_channels(
                channels=channels,
                disable_others=True,
                channel_configs=channel_configs
            )

            progress.close()

            if config_success:
                self.statusBar().showMessage("✓ 通道配置已成功应用到设备")
                QMessageBox.information(
                    self,
                    "配置成功",
                    f"通道配置已成功应用到设备！\n\n"
                    f"已配置通道：\n"
                    f"• {channels[0]} - 三元电池电压\n"
                    f"• {channels[1]} - 三元电池温度\n"
                    f"• {channels[2]} - 刀片电池电压\n"
                    f"• {channels[3]} - 刀片电池温度"
                )
            else:
                QMessageBox.warning(
                    self,
                    "配置失败",
                    "通道配置应用失败，请尝试重新连接设备。"
                )
        except Exception as e:
            QMessageBox.critical(self, "错误", f"应用配置时出错：{str(e)}")

    def _update_plot_ranges(self) -> None:
        """根据配置的量程更新Y轴范围"""
        # 获取电压量程（两个电池取最大值）
        voltage_range = max(
            self.channel_config['ternary_voltage']['range'],
            self.channel_config['blade_voltage']['range']
        )

        # 获取温度量程（两个电池取最大值）
        temp_range = max(
            self.channel_config['ternary_temp']['range'],
            self.channel_config['blade_temp']['range']
        )

        # 更新三元电池波形的Y轴范围（左图）
        self.waveforms.left_plot.setYRange(0, voltage_range * 1.1, padding=0)  # 左Y轴（电压）
        self.waveforms.left_plot.viewbox_temp.setYRange(0, temp_range * 1.1)  # 右Y轴（温度）

        # 更新刀片电池波形的Y轴范围（右图）
        self.waveforms.right_plot.setYRange(0, voltage_range * 1.1, padding=0)  # 左Y轴（电压）
        self.waveforms.right_plot.viewbox_temp.setYRange(0, temp_range * 1.1)  # 右Y轴（温度）

        # 更新Y轴标签（保持当前颜色）
        self.waveforms.left_plot.setLabel('left', f'电压 (V, 量程: {voltage_range}V)', color=self.current_volt_color, **{'font-size': '11pt'})
        self.waveforms.left_plot.setLabel('right', f'温度 (°C, 量程: {temp_range}°C)', color=self.current_temp_color, **{'font-size': '11pt'})
        self.waveforms.right_plot.setLabel('left', f'电压 (V, 量程: {voltage_range}V)', color=self.current_volt_color, **{'font-size': '11pt'})
        self.waveforms.right_plot.setLabel('right', f'温度 (°C, 量程: {temp_range}°C)', color=self.current_temp_color, **{'font-size': '11pt'})

        self.statusBar().showMessage(f"✓ Y轴范围已更新：电压 0-{voltage_range}V，温度 0-{temp_range}°C")

    def _export_report(self) -> None:
        """导出测试报告"""
        if not self.analysis_engine.ternary_data.timestamps:
            QMessageBox.warning(self, "无数据", "请先进行测试，采集数据后再导出报告")
            return

        # 生成报告数据
        report_data = self.analysis_engine.generate_report_data()

        # 显示分析结果
        temp_compare = report_data.get('对比分析', {}).get('对比', {})

        result_text = "=== 电池测试分析报告 ===\n\n"
        result_text += f"测试时长: {report_data['测试时长']:.1f} 秒\n\n"

        result_text += "【温升对比分析】\n"
        if temp_compare:
            result_text += f"三元电池温升: {temp_compare['三元温升']:.2f}°C\n"
            result_text += f"刀片电池温升: {temp_compare['刀片温升']:.2f}°C\n"
            result_text += f"温升差异: {temp_compare['温升差异']:.2f}°C\n"
            result_text += f"优势电池: {temp_compare['优势电池']}\n\n"

        result_text += f"【mAh容量】\n"
        result_text += f"累计容量: {report_data['mAh容量']:.2f} mAh\n"

        # 询问是否导出到文件
        from PySide6.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            self,
            "分析报告",
            result_text + "\n\n是否导出报告到文件？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes
        )

        if reply == QMessageBox.StandardButton.Yes:
            self._export_report_to_file(report_data, result_text)

    def _save_waveform(self) -> None:
        """保存波形数据到文件"""
        if not self.x_data:
            QMessageBox.warning(self, "无数据", "当前没有波形数据可以保存")
            return

        from PySide6.QtWidgets import QFileDialog
        import json
        from datetime import datetime

        # 选择保存路径
        default_name = f"waveform_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存波形数据",
            default_name,
            "JSON文件 (*.json);;所有文件 (*.*)"
        )

        if not file_path:
            return

        try:
            # 准备保存的数据
            save_data = {
                'timestamp': datetime.now().isoformat(),
                'product_model': self.control.edit_model.text(),
                'product_sn': self.control.edit_sn.text(),
                'tester': self.control.edit_tester.text(),
                'x_data': self.x_data,
                'ternary_voltage': self.ternary_volt_data,
                'ternary_temp': self.ternary_temp_data,
                'blade_voltage': self.blade_volt_data,
                'blade_temp': self.blade_temp_data,
                'channel_config': self.channel_config,
                'analysis_data': self.analysis_engine.generate_report_data() if self.analysis_engine.ternary_data.timestamps else None,
            }

            # 写入文件
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2)

            self.statusBar().showMessage(f"✓ 波形数据已保存到: {file_path}")
            QMessageBox.information(self, "保存成功", f"波形数据已保存到:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"保存波形数据时出错:\n{str(e)}")

    def _recall_waveform(self) -> None:
        """从文件召回波形数据"""
        from PySide6.QtWidgets import QFileDialog
        import json

        # 选择文件
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "召回波形数据",
            "",
            "JSON文件 (*.json);;所有文件 (*.*)"
        )

        if not file_path:
            return

        try:
            # 读取文件
            with open(file_path, 'r', encoding='utf-8') as f:
                load_data = json.load(f)

            # 停止当前采集
            if self.is_running:
                self._on_stop()

            # 恢复数据
            self.x_data = load_data.get('x_data', [])
            self.ternary_volt_data = load_data.get('ternary_voltage', [])
            self.ternary_temp_data = load_data.get('ternary_temp', [])
            self.blade_volt_data = load_data.get('blade_voltage', [])
            self.blade_temp_data = load_data.get('blade_temp', [])

            # 恢复产品信息
            if 'product_model' in load_data:
                self.control.edit_model.setText(load_data['product_model'])
            if 'product_sn' in load_data:
                self.control.edit_sn.setText(load_data['product_sn'])
            if 'tester' in load_data:
                self.control.edit_tester.setText(load_data['tester'])

            # 恢复通道配置
            if 'channel_config' in load_data:
                self.channel_config = load_data['channel_config']

            # 更新波形显示
            if len(self.volt_curves) >= 2 and len(self.temp_curves) >= 2:
                self.volt_curves[0].setData(self.x_data, self.ternary_volt_data)
                self.temp_curves[0].setData(self.x_data, self.ternary_temp_data)
                self.volt_curves[1].setData(self.x_data, self.blade_volt_data)
                self.temp_curves[1].setData(self.x_data, self.blade_temp_data)

            # 更新KPI显示
            if self.ternary_volt_data:
                self.ternary_voltage_kpi.set_value(f"{self.ternary_volt_data[-1]:.2f}")
                self.ternary_temp_kpi.set_value(f"{self.ternary_temp_data[-1]:.2f}")
                self.blade_voltage_kpi.set_value(f"{self.blade_volt_data[-1]:.2f}")
                self.blade_temp_kpi.set_value(f"{self.blade_temp_data[-1]:.2f}")

            self.statusBar().showMessage(f"✓ 波形数据已召回: {file_path}")
            QMessageBox.information(
                self,
                "召回成功",
                f"波形数据已成功召回！\n\n"
                f"文件: {file_path}\n"
                f"时间: {load_data.get('timestamp', '未知')}\n"
                f"数据点数: {len(self.x_data)}"
            )

        except Exception as e:
            QMessageBox.critical(self, "召回失败", f"召回波形数据时出错:\n{str(e)}")

    def _export_report_to_file(self, report_data: dict, result_text: str) -> None:
        """导出报告到文件（支持TXT、Excel、HTML格式）"""
        from PySide6.QtWidgets import QFileDialog, QMessageBox
        from datetime import datetime

        # 选择导出格式和路径
        default_name = f"battery_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "导出测试报告",
            default_name,
            "文本文件 (*.txt);;Excel文件 (*.xlsx);;HTML文件 (*.html);;所有文件 (*.*)"
        )

        if not file_path:
            return

        try:
            if file_path.endswith('.txt'):
                self._export_txt_report(file_path, result_text, report_data)
            elif file_path.endswith('.xlsx'):
                self._export_excel_report(file_path, report_data)
            elif file_path.endswith('.html'):
                self._export_html_report(file_path, result_text, report_data)
            else:
                # 默认导出为TXT
                self._export_txt_report(file_path + '.txt', result_text, report_data)

            self.statusBar().showMessage(f"✓ 报告已导出: {file_path}")
            QMessageBox.information(self, "导出成功", f"测试报告已导出到:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出报告时出错:\n{str(e)}")

    def _export_txt_report(self, file_path: str, result_text: str, report_data: dict) -> None:
        """导出TXT格式报告"""
        from datetime import datetime

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("电池电压与温升测试分析报告\n")
            f.write("=" * 60 + "\n\n")

            f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"产品型号: {self.control.edit_model.text()}\n")
            f.write(f"产品流水号: {self.control.edit_sn.text()}\n")
            f.write(f"测试员: {self.control.edit_tester.text()}\n")
            f.write(f"测试时长: {report_data['测试时长']:.1f} 秒\n\n")

            f.write("-" * 60 + "\n")
            f.write("一、三元电池测试结果\n")
            f.write("-" * 60 + "\n")
            ternary = report_data.get('三元电池', {})
            temp_rise = ternary.get('温升分析', {})
            voltage_drop = ternary.get('压降分析', {})

            if temp_rise:
                f.write("\n【温升分析】\n")
                f.write(f"  初始温度: {temp_rise.get('初始温度', 0):.2f} °C\n")
                f.write(f"  当前温度: {temp_rise.get('当前温度', 0):.2f} °C\n")
                f.write(f"  峰值温度: {temp_rise.get('峰值温度', 0):.2f} °C\n")
                f.write(f"  最低温度: {temp_rise.get('最低温度', 0):.2f} °C\n")
                f.write(f"  温升: {temp_rise.get('温升', 0):.2f} °C\n")
                f.write(f"  平均温度: {temp_rise.get('平均温度', 0):.2f} °C\n")

            if voltage_drop:
                f.write("\n【电压分析】\n")
                f.write(f"  初始电压: {voltage_drop.get('初始电压', 0):.2f} V\n")
                f.write(f"  当前电压: {voltage_drop.get('当前电压', 0):.2f} V\n")
                f.write(f"  最高电压: {voltage_drop.get('最高电压', 0):.2f} V\n")
                f.write(f"  最低电压: {voltage_drop.get('最低电压', 0):.2f} V\n")
                f.write(f"  电压降: {voltage_drop.get('电压降', 0):.2f} V\n")
                f.write(f"  压降率: {voltage_drop.get('压降率', 0):.2f} %\n")
                f.write(f"  平均电压: {voltage_drop.get('平均电压', 0):.2f} V\n")

            f.write("\n" + "-" * 60 + "\n")
            f.write("二、刀片电池测试结果\n")
            f.write("-" * 60 + "\n")
            blade = report_data.get('刀片电池', {})
            temp_rise = blade.get('温升分析', {})
            voltage_drop = blade.get('压降分析', {})

            if temp_rise:
                f.write("\n【温升分析】\n")
                f.write(f"  初始温度: {temp_rise.get('初始温度', 0):.2f} °C\n")
                f.write(f"  当前温度: {temp_rise.get('当前温度', 0):.2f} °C\n")
                f.write(f"  峰值温度: {temp_rise.get('峰值温度', 0):.2f} °C\n")
                f.write(f"  最低温度: {temp_rise.get('最低温度', 0):.2f} °C\n")
                f.write(f"  温升: {temp_rise.get('温升', 0):.2f} °C\n")
                f.write(f"  平均温度: {temp_rise.get('平均温度', 0):.2f} °C\n")

            if voltage_drop:
                f.write("\n【电压分析】\n")
                f.write(f"  初始电压: {voltage_drop.get('初始电压', 0):.2f} V\n")
                f.write(f"  当前电压: {voltage_drop.get('当前电压', 0):.2f} V\n")
                f.write(f"  最高电压: {voltage_drop.get('最高电压', 0):.2f} V\n")
                f.write(f"  最低电压: {voltage_drop.get('最低电压', 0):.2f} V\n")
                f.write(f"  电压降: {voltage_drop.get('电压降', 0):.2f} V\n")
                f.write(f"  压降率: {voltage_drop.get('压降率', 0):.2f} %\n")
                f.write(f"  平均电压: {voltage_drop.get('平均电压', 0):.2f} V\n")

            f.write("\n" + "-" * 60 + "\n")
            f.write("三、对比分析\n")
            f.write("-" * 60 + "\n")
            compare = report_data.get('对比分析', {}).get('对比', {})
            if compare:
                f.write(f"\n三元电池温升: {compare.get('三元温升', 0):.2f} °C\n")
                f.write(f"刀片电池温升: {compare.get('刀片温升', 0):.2f} °C\n")
                f.write(f"温升差异: {compare.get('温升差异', 0):.2f} °C\n")
                f.write(f"优势电池: {compare.get('优势电池', '未知')}\n")

            f.write("\n" + "-" * 60 + "\n")
            f.write("四、容量测试\n")
            f.write("-" * 60 + "\n")
            f.write(f"\n累计容量: {report_data.get('mAh容量', 0):.2f} mAh\n")

            f.write("\n" + "=" * 60 + "\n")
            f.write("报告结束\n")
            f.write("=" * 60 + "\n")

    def _export_excel_report(self, file_path: str, report_data: dict) -> None:
        """导出Excel格式报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from datetime import datetime
        except ImportError:
            QMessageBox.warning(
                self,
                "缺少依赖",
                "导出Excel需要安装 openpyxl 库\n\n请运行: pip install openpyxl"
            )
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "电池测试报告"

        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        # 标题样式
        title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_alignment = Alignment(horizontal='center', vertical='center')

        # 表头样式
        header_font = Font(name='微软雅黑', size=12, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # 边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1

        # 标题
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "电池电压与温升测试分析报告"
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_alignment
        row += 2

        # 基本信息
        ws[f'A{row}'] = "生成时间:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        ws[f'A{row}'] = "产品型号:"
        ws[f'B{row}'] = self.control.edit_model.text()
        row += 1
        ws[f'A{row}'] = "产品流水号:"
        ws[f'B{row}'] = self.control.edit_sn.text()
        row += 1
        ws[f'A{row}'] = "测试员:"
        ws[f'B{row}'] = self.control.edit_tester.text()
        row += 1
        ws[f'A{row}'] = "测试时长:"
        ws[f'B{row}'] = f"{report_data['测试时长']:.1f} 秒"
        row += 2

        # 三元电池数据
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "三元电池测试结果"
        cell.font = header_font
        cell.fill = header_fill
        row += 1

        ternary = report_data.get('三元电池', {})
        temp_rise = ternary.get('温升分析', {})
        voltage_drop = ternary.get('压降分析', {})

        ws[f'A{row}'] = "项目"
        ws[f'B{row}'] = "温升分析"
        ws[f'C{row}'] = "项目"
        ws[f'D{row}'] = "电压分析"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = header_font
        row += 1

        metrics = [
            ('初始温度', '初始电压'),
            ('当前温度', '当前电压'),
            ('峰值温度', '最高电压'),
            ('最低温度', '最低电压'),
            ('温升', '电压降'),
            ('平均温度', '平均电压'),
        ]

        for temp_key, volt_key in metrics:
            ws[f'A{row}'] = temp_key
            ws[f'B{row}'] = f"{temp_rise.get(temp_key, 0):.2f}"
            ws[f'C{row}'] = volt_key
            ws[f'D{row}'] = f"{voltage_drop.get(volt_key, 0):.2f}"
            row += 1

        row += 1

        # 刀片电池数据
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "刀片电池测试结果"
        cell.font = header_font
        cell.fill = header_fill
        row += 1

        blade = report_data.get('刀片电池', {})
        temp_rise = blade.get('温升分析', {})
        voltage_drop = blade.get('压降分析', {})

        ws[f'A{row}'] = "项目"
        ws[f'B{row}'] = "温升分析"
        ws[f'C{row}'] = "项目"
        ws[f'D{row}'] = "电压分析"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = header_font
        row += 1

        for temp_key, volt_key in metrics:
            ws[f'A{row}'] = temp_key
            ws[f'B{row}'] = f"{temp_rise.get(temp_key, 0):.2f}"
            ws[f'C{row}'] = volt_key
            ws[f'D{row}'] = f"{voltage_drop.get(volt_key, 0):.2f}"
            row += 1

        row += 1

        # 对比分析
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "对比分析"
        cell.font = header_font
        cell.fill = header_fill
        row += 1

        compare = report_data.get('对比分析', {}).get('对比', {})
        if compare:
            ws[f'A{row}'] = "三元电池温升"
            ws[f'B{row}'] = f"{compare.get('三元温升', 0):.2f} °C"
            row += 1
            ws[f'A{row}'] = "刀片电池温升"
            ws[f'B{row}'] = f"{compare.get('刀片温升', 0):.2f} °C"
            row += 1
            ws[f'A{row}'] = "温升差异"
            ws[f'B{row}'] = f"{compare.get('温升差异', 0):.2f} °C"
            row += 1
            ws[f'A{row}'] = "优势电池"
            ws[f'B{row}'] = compare.get('优势电池', '未知')
            row += 1

        row += 1
        ws[f'A{row}'] = "累计容量"
        ws[f'B{row}'] = f"{report_data.get('mAh容量', 0):.2f} mAh"

        # 应用边框
        for row_cells in ws.iter_rows(min_row=1, max_row=row, min_col=1, max_col=4):
            for cell in row_cells:
                cell.border = thin_border

        wb.save(file_path)

    def _export_html_report(self, file_path: str, result_text: str, report_data: dict) -> None:
        """导出HTML格式报告"""
        from datetime import datetime

        html_content = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>电池测试分析报告</title>
    <style>
        body {{
            font-family: "Microsoft YaHei", Arial, sans-serif;
            max-width: 1000px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            background-color: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2c3e50;
            text-align: center;
            border-bottom: 3px solid #3498db;
            padding-bottom: 15px;
        }}
        h2 {{
            color: #34495e;
            margin-top: 30px;
            border-left: 4px solid #3498db;
            padding-left: 10px;
        }}
        .info-table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        .info-table th, .info-table td {{
            padding: 12px;
            text-align: left;
            border: 1px solid #ddd;
        }}
        .info-table th {{
            background-color: #3498db;
            color: white;
            font-weight: bold;
        }}
        .info-table tr:nth-child(even) {{
            background-color: #f2f2f2;
        }}
        .highlight {{
            background-color: #fff3cd;
            padding: 15px;
            border-left: 4px solid #ffc107;
            margin: 15px 0;
        }}
        .footer {{
            text-align: center;
            margin-top: 30px;
            color: #7f8c8d;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>电池电压与温升测试分析报告</h1>

        <table class="info-table">
            <tr>
                <th>项目</th>
                <th>内容</th>
            </tr>
            <tr>
                <td>生成时间</td>
                <td>{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</td>
            </tr>
            <tr>
                <td>产品型号</td>
                <td>{self.control.edit_model.text()}</td>
            </tr>
            <tr>
                <td>产品流水号</td>
                <td>{self.control.edit_sn.text()}</td>
            </tr>
            <tr>
                <td>测试员</td>
                <td>{self.control.edit_tester.text()}</td>
            </tr>
            <tr>
                <td>测试时长</td>
                <td>{report_data['测试时长']:.1f} 秒</td>
            </tr>
        </table>

        <h2>一、三元电池测试结果</h2>
        <table class="info-table">
            <tr>
                <th colspan="2">温升分析</th>
                <th colspan="2">电压分析</th>
            </tr>
"""

        ternary = report_data.get('三元电池', {})
        temp_rise = ternary.get('温升分析', {})
        voltage_drop = ternary.get('压降分析', {})

        metrics = [
            ('初始温度', '初始电压'),
            ('当前温度', '当前电压'),
            ('峰值温度', '最高电压'),
            ('最低温度', '最低电压'),
            ('温升', '电压降'),
            ('平均温度', '平均电压'),
        ]

        for temp_key, volt_key in metrics:
            html_content += f"""
            <tr>
                <td>{temp_key}</td>
                <td>{temp_rise.get(temp_key, 0):.2f} °C</td>
                <td>{volt_key}</td>
                <td>{voltage_drop.get(volt_key, 0):.2f} V</td>
            </tr>
"""

        html_content += """
        </table>

        <h2>二、刀片电池测试结果</h2>
        <table class="info-table">
            <tr>
                <th colspan="2">温升分析</th>
                <th colspan="2">电压分析</th>
            </tr>
"""

        blade = report_data.get('刀片电池', {})
        temp_rise = blade.get('温升分析', {})
        voltage_drop = blade.get('压降分析', {})

        for temp_key, volt_key in metrics:
            html_content += f"""
            <tr>
                <td>{temp_key}</td>
                <td>{temp_rise.get(temp_key, 0):.2f} °C</td>
                <td>{volt_key}</td>
                <td>{voltage_drop.get(volt_key, 0):.2f} V</td>
            </tr>
"""

        compare = report_data.get('对比分析', {}).get('对比', {})

        html_content += f"""
        </table>

        <h2>三、对比分析</h2>
        <div class="highlight">
            <p><strong>三元电池温升:</strong> {compare.get('三元温升', 0):.2f} °C</p>
            <p><strong>刀片电池温升:</strong> {compare.get('刀片温升', 0):.2f} °C</p>
            <p><strong>温升差异:</strong> {compare.get('温升差异', 0):.2f} °C</p>
            <p><strong>优势电池:</strong> {compare.get('优势电池', '未知')}</p>
        </div>

        <h2>四、容量测试</h2>
        <table class="info-table">
            <tr>
                <th>项目</th>
                <th>数值</th>
            </tr>
            <tr>
                <td>累计容量</td>
                <td>{report_data.get('mAh容量', 0):.2f} mAh</td>
            </tr>
        </table>

        <div class="footer">
            <p>电池电压与温升采集软件 v0.1.0</p>
            <p>报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>
"""

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)


class MXPlusBDialog(QDialog):
    """mX+b线性校准对话框。"""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("mX+b 线性校准")
        self.setFixedSize(400, 300)

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # 说明文本
        desc = QLabel("线性校准公式: Y = mX + b")
        desc.setStyleSheet("color: #6dd5ed; font-size: 14px; font-weight: bold;")
        layout.addWidget(desc)

        # 参数输入区域
        params_group = QGroupBox("校准参数")
        params_layout = QFormLayout(params_group)

        self.edit_m = QLineEdit("1.0")
        self.edit_m.setPlaceholderText("斜率 m")
        params_layout.addRow("斜率 m:", self.edit_m)

        self.edit_b = QLineEdit("0.0")
        self.edit_b.setPlaceholderText("截距 b")
        params_layout.addRow("截距 b:", self.edit_b)

        layout.addWidget(params_group)

        # 通道选择
        channel_group = QGroupBox("应用通道")
        channel_layout = QVBoxLayout(channel_group)
        channel_layout.setSpacing(8)

        self.radio_ternary = QRadioButton("三元电池")
        self.radio_ternary.setChecked(True)
        self.radio_blade = QRadioButton("刀片电池")

        channel_layout.addWidget(self.radio_ternary)
        channel_layout.addWidget(self.radio_blade)

        layout.addWidget(channel_group)

        # 按钮区域
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        button_layout.addWidget(btn_cancel)

        btn_apply = QPushButton("应用校准")
        btn_apply.clicked.connect(self._apply_calibration)
        btn_apply.setStyleSheet("""
            QPushButton {
                background-color: #059669;
                color: #ffffff;
                border: none;
                padding: 10px 20px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #047857;
            }
        """)
        button_layout.addWidget(btn_apply)

        layout.addLayout(button_layout)

    def _apply_calibration(self) -> None:
        """应用校准参数。"""
        try:
            m = float(self.edit_m.text())
            b = float(self.edit_b.text())

            battery_type = "ternary" if self.radio_ternary.isChecked() else "blade"
            channel_name = "三元电池" if self.radio_ternary.isChecked() else "刀片电池"

            # 应用校准到分析引擎
            if hasattr(self.parent(), 'analysis_engine'):
                self.parent().analysis_engine.set_mx_plus_b(battery_type, m, b)
            
            QMessageBox.information(
                self,
                "校准应用成功",
                f"已对{channel_name}应用线性校准:\nY = {m}X + {b}"
            )

            self.accept()

        except ValueError:
            QMessageBox.warning(self, "输入错误", "请输入有效的数字参数")


class MAHTestDialog(QDialog):
    """mAh容量测试对话框。"""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("mAh 容量测试")
        self.setFixedSize(450, 400)

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # 标题
        title = QLabel("毫安时容量测试")
        title.setStyleSheet("color: #6dd5ed; font-size: 16px; font-weight: bold;")
        layout.addWidget(title)

        # 测试参数
        params_group = QGroupBox("测试参数")
        params_layout = QFormLayout(params_group)

        self.edit_test_current = QLineEdit("1000")
        self.edit_test_current.setPlaceholderText("测试电流 (mA)")
        params_layout.addRow("测试电流 (mA):", self.edit_test_current)

        self.edit_test_voltage = QLineEdit("3.7")
        self.edit_test_voltage.setPlaceholderText("测试电压 (V)")
        params_layout.addRow("测试电压 (V):", self.edit_test_voltage)

        self.edit_test_duration = QLineEdit("3600")
        self.edit_test_duration.setPlaceholderText("测试时长 (秒)")
        params_layout.addRow("测试时长 (秒):", self.edit_test_duration)

        layout.addWidget(params_group)

        # 通道选择
        channel_group = QGroupBox("测试通道")
        channel_layout = QVBoxLayout(channel_group)
        channel_layout.setSpacing(8)

        self.radio_ternary_mah = QRadioButton("三元电池")
        self.radio_ternary_mah.setChecked(True)
        self.radio_blade_mah = QRadioButton("刀片电池")

        channel_layout.addWidget(self.radio_ternary_mah)
        channel_layout.addWidget(self.radio_blade_mah)

        layout.addWidget(channel_group)

        # 实时数据显示
        display_group = QGroupBox("测试结果")
        display_layout = QFormLayout(display_group)

        self.label_current = QLabel("0.00 mA")
        self.label_capacity = QLabel("0.00 mAh")
        self.label_voltage = QLabel("0.00 V")

        display_layout.addRow("实时电流:", self.label_current)
        display_layout.addRow("累计容量:", self.label_capacity)
        display_layout.addRow("实时电压:", self.label_voltage)

        layout.addWidget(display_group)

        # 按钮区域
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        button_layout.addWidget(btn_cancel)

        self.btn_start_test = QPushButton("开始测试")
        self.btn_start_test.setStyleSheet("""
            QPushButton {
                background-color: #dc2626;
                color: #ffffff;
                border: none;
                padding: 10px 20px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #b91c1c;
            }
        """)
        self.btn_start_test.clicked.connect(self._start_capacity_test)
        button_layout.addWidget(self.btn_start_test)

        layout.addLayout(button_layout)

        # 定时器用于模拟测试过程
        self.test_timer = QTimer()
        self.test_timer.timeout.connect(self._update_test_display)
        self.test_active = False
        self.test_time = 0

    def _start_capacity_test(self) -> None:
        """开始容量测试。"""
        try:
            current = float(self.edit_test_current.text())
            voltage = float(self.edit_test_voltage.text())
            duration = float(self.edit_test_duration.text())

            if current <= 0 or voltage <= 0 or duration <= 0:
                raise ValueError("参数必须为正数")

            if self.radio_ternary_mah.isChecked():
                channel = "三元电池"
            else:
                channel = "刀片电池"

            # 启动分析引擎的mAh测试
            if hasattr(self.parent(), 'analysis_engine'):
                self.parent().analysis_engine.start_mah_test(current)
            
            # 开始模拟测试
            self.test_active = True
            self.test_time = 0
            self.test_current = current
            self.test_voltage = voltage
            self.test_duration = duration

            self.btn_start_test.setText("测试中...")
            self.btn_start_test.setEnabled(False)

            # 每秒更新显示
            self.test_timer.start(1000)

            QMessageBox.information(
                self,
                "测试开始",
                f"开始对{channel}进行容量测试\n电流: {current}mA, 电压: {voltage}V, 时长: {duration}s"
            )

        except ValueError as e:
            QMessageBox.warning(self, "参数错误", str(e))

    def _update_test_display(self) -> None:
        """更新测试显示。"""
        self.test_time += 1

        # 从分析引擎获取实时容量
        if hasattr(self.parent(), 'analysis_engine'):
            capacity = self.parent().analysis_engine.update_mah_capacity()
        else:
            capacity = (self.test_current * self.test_time) / 3600

        # 显示当前值
        self.label_current.setText(f"{self.test_current:.2f} mA")
        self.label_capacity.setText(f"{capacity:.2f} mAh")
        self.label_voltage.setText(f"{self.test_voltage:.2f} V")

        # 检查测试完成
        if self.test_time >= self.test_duration:
            self._finish_test()

    def _finish_test(self) -> None:
        """完成测试。"""
        self.test_active = False
        self.test_timer.stop()

        final_capacity = (self.test_current * self.test_duration) / 3600

        QMessageBox.information(
            self,
            "测试完成",
            f"容量测试完成！\n最终容量: {final_capacity:.2f} mAh"
        )

        self.btn_start_test.setText("重新测试")
        self.btn_start_test.setEnabled(True)

        # 重置显示
        self.label_current.setText("0.00 mA")
        self.label_capacity.setText("0.00 mAh")
        self.label_voltage.setText("0.00 V")
